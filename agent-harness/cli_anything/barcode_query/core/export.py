"""Export functionality for barcode queries"""
import os
import json
from pathlib import Path
from datetime import datetime


class ExportManager:
    """Manages export of barcode data"""

    def __init__(self):
        self.barcode_dir = Path(__file__).parent.parent.parent.parent / 'barcode'
        self.data_file = self.barcode_dir / 'barcode_data.json'

    def _load_data(self):
        if self.data_file.exists():
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def export_xlsx(self, barcodes=None, export_all=False, output='export_result.xlsx'):
        """Export barcodes to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return {'success': False, 'error': 'openpyxl 未安装，请运行: pip install openpyxl'}

        barcode_dir = Path(__file__).parent.parent.parent.parent / 'barcode'
        archived_dir = barcode_dir / 'archived'
        export_path = barcode_dir / output

        # Get barcodes to export
        if export_all:
            barcodes = []
            for d in [barcode_dir, archived_dir]:
                if d.exists():
                    for f in os.listdir(d):
                        if f.endswith('.html'):
                            barcodes.append(f.replace('.html', ''))

        if not barcodes:
            return {'success': False, 'error': '没有可导出的条码'}

        # Export fields order
        export_fields = [
            ('newname1_sr1', '条码号'),
            ('remark', '备注'),
            ('newisclosed1_sr2', '结单状态'),
            ('SHIPSTATUS1', '装箱单状态'),
            ('newproductidName1_sr2', '机型'),
            ('ProductNumber1', '物料编码'),
            ('myproductdealer1_sr5', '所属经销商'),
            ('newdealername1_sr2', '服务经销商'),
            ('newstationidName1', '服务站'),
            ('typestr1_sr2', '服务类型'),
            ('statustr1_sr2', '服务单状态'),
            ('servno1_sr2', '服务单号'),
            ('name1_sr2', '客户'),
            ('newtelephone1_sr2', '电话'),
            ('newaddress1_sr2', '地址'),
            ('zxd1', '装箱单号'),
            ('shipdate1', '发货日期'),
            ('newerpshipno1', '发货单号'),
            ('newordsalesorderidName1', '订单号'),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "条码查询结果"

        # Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = [label for _, label in export_fields]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data
        data = self._load_data()
        for row_idx, barcode in enumerate(barcodes, 2):
            for col_idx, (field_id, _) in enumerate(export_fields, 1):
                if field_id == 'remark':
                    value = data.get(barcode, {}).get('remark', '')
                else:
                    value = ''  # Would need to parse HTML for actual values
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Column widths
        for col_idx in range(1, len(export_fields) + 1):
            if col_idx <= 26:
                col_letter = chr(64 + col_idx)
            else:
                col_letter = chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
            ws.column_dimensions[col_letter].width = 18

        wb.save(export_path)
        return {'success': True, 'filename': str(export_path), 'count': len(barcodes)}
