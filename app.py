#!/usr/bin/env python3
"""Flask Web 应用 - 条码查询结果展示"""
import os
import sys

# 在导入任何其他模块之前设置环境变量来禁用 asyncio
os.environ['EVENTLET_NO_GREENDNS'] = '1'
os.environ['ASYNCIO_CORE_EVENT_LOOP'] = '0'

import re
import json
import time
import builtins
import html as html_mod
import platform
import socket
import threading
import queue
import uuid
import shutil
import hmac
from collections import OrderedDict
from urllib import request as urlrequest, error as urlerror
from flask import Flask, render_template, request, jsonify, send_from_directory, Response, session, redirect, has_request_context
from datetime import datetime

from cluster.config import ClusterConfig
from cluster.services import build_cluster_services

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
except Exception:
    pass

_ORIGINAL_PRINT = builtins.print

def _console_safe_text(value):
    return ('' if value is None else str(value)).replace('\xa0', ' ')

def _safe_print(*args, **kwargs):
    safe_args = [_console_safe_text(arg) for arg in args]
    try:
        _ORIGINAL_PRINT(*safe_args, **kwargs)
    except UnicodeEncodeError:
        file = kwargs.get("file") or sys.stdout
        sep = kwargs.get("sep", " ")
        end = kwargs.get("end", "\n")
        sep = " " if sep is None else str(sep)
        end = "\n" if end is None else str(end)
        text = sep.join(safe_args) + end
        encoding = getattr(file, "encoding", None) or "utf-8"
        text = text.encode(encoding, errors="replace").decode(encoding, errors="replace")
        try:
            file.write(text)
            if kwargs.get("flush"):
                file.flush()
        except Exception:
            pass

print = _safe_print

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

try:
    REPORT_IDLE_TIMEOUT_SECONDS = max(60, int(os.environ.get("CRM_REPORT_IDLE_SECONDS", "3600")))
except (TypeError, ValueError):
    REPORT_IDLE_TIMEOUT_SECONDS = 3600
try:
    REPORT_IDLE_CLEANUP_INTERVAL_SECONDS = max(60, int(os.environ.get("CRM_REPORT_IDLE_CHECK_SECONDS", "300")))
except (TypeError, ValueError):
    REPORT_IDLE_CLEANUP_INTERVAL_SECONDS = 300
try:
    QUERY_SLOT_FAILURE_COOLDOWN_SECONDS = max(30, int(os.environ.get("CRM_QUERY_SLOT_FAILURE_COOLDOWN_SECONDS", "300")))
except (TypeError, ValueError):
    QUERY_SLOT_FAILURE_COOLDOWN_SECONDS = 300
try:
    LIBRARY_QUERY_SLOT_RETRY_LIMIT = max(1, int(os.environ.get("CRM_LIBRARY_QUERY_SLOT_RETRY_LIMIT", "3")))
except (TypeError, ValueError):
    LIBRARY_QUERY_SLOT_RETRY_LIMIT = 3

APP_STARTED_AT = time.time()
RUNTIME_BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(__file__)
RESOURCE_BASE_DIR = getattr(sys, "_MEIPASS", RUNTIME_BASE_DIR)
CRM_CONFIG_PATH = os.path.join(RUNTIME_BASE_DIR, "config.json")
IS_DESKTOP_APP = os.environ.get("CRM_DESKTOP_APP") == "1"
SERVER_CLUSTER_CONFIG = ClusterConfig.from_env()
CLUSTER_SERVICES = None
CLUSTER_SERVICES_LOCK = threading.Lock()
STARTUP_LOGIN_AUTO_CHECK = os.environ.get("CRM_STARTUP_LOGIN_AUTO_CHECK", "1") != "0"
try:
    STARTUP_LOGIN_CHECK_DELAY_SECONDS = max(0, int(os.environ.get("CRM_STARTUP_LOGIN_CHECK_DELAY_SECONDS", "2")))
except (TypeError, ValueError):
    STARTUP_LOGIN_CHECK_DELAY_SECONDS = 2
try:
    STARTUP_LOGIN_CHECK_STAGGER_SECONDS = max(0, int(os.environ.get("CRM_STARTUP_LOGIN_CHECK_STAGGER_SECONDS", "3")))
except (TypeError, ValueError):
    STARTUP_LOGIN_CHECK_STAGGER_SECONDS = 3

def load_crm_config():
    config_paths = [
        CRM_CONFIG_PATH,
    ]
    if os.path.exists("/.dockerenv"):
        config_paths.append(os.path.join(RESOURCE_BASE_DIR, "config.docker.example.json"))
    config_paths.append(os.path.join(RESOURCE_BASE_DIR, "config.example.json"))
    for path in config_paths:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError("未找到 config.json 或示例配置文件")


def _get_cluster_services():
    global CLUSTER_SERVICES
    if not SERVER_CLUSTER_CONFIG.enabled:
        return None
    if CLUSTER_SERVICES is None:
        with CLUSTER_SERVICES_LOCK:
            if CLUSTER_SERVICES is None:
                CLUSTER_SERVICES = build_cluster_services(SERVER_CLUSTER_CONFIG)
    return CLUSTER_SERVICES

class CRMSession:
    def __init__(self, session_dir=None):
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.session_dir = session_dir
        self.lock = threading.Lock()
        self.logged_in = False
        self.needs_navigation = True  # 标记是否需要导航到报表页面
        self.last_report_error = ""
        self.report_last_used_at = 0

    def _browser_crash_message(self):
        return "CRM 浏览器页面已崩溃，已自动关闭当前会话，请重新登录 CRM 后再操作"

    def _is_browser_crash_error(self, error):
        text = str(error)
        return any(key in text for key in [
            "Target crashed",
            "Target page, context or browser has been closed",
            "Browser has been closed",
            "Page crashed",
            "browser has been closed",
            "context has been closed",
        ])

    def _handle_browser_exception(self, error):
        if not self._is_browser_crash_error(error):
            return ""
        self._close_browser()
        self.logged_in = False
        self.needs_navigation = True
        self.last_report_error = ""
        return self._browser_crash_message()

    def _goto(self, url, timeout=60000, wait_until="domcontentloaded"):
        try:
            self.page.goto(url, wait_until=wait_until, timeout=timeout)
            return True, ""
        except Exception as e:
            crash_message = self._handle_browser_exception(e)
            if crash_message:
                return False, crash_message
            error_text = str(e)
            if "Timeout" in error_text or "timeout" in error_text:
                try:
                    body_text = self.page.inner_text("body", timeout=3000)
                    if body_text and len(body_text.strip()) > 20:
                        print(f"  [WARN] 页面加载超时但已有内容，继续执行: {url}")
                        return True, ""
                except Exception:
                    pass
                return False, f"打开 CRM 页面超时，请检查云服务器到 CRM 网站的网络连通性：{_brief_batch_error(error_text, 240)}"
            return False, error_text

    def is_alive(self):
        try:
            if self.context and self.page:
                self.page.url
                return True
        except Exception:
            pass
        return False

    def _ensure_browser(self):
        """启动或复用浏览器实例"""
        cfg = load_crm_config()
        browser_cfg = cfg.get("browser", {})
        session_dir = self.session_dir or cfg["session"]["state_path"]
        os.makedirs(session_dir, exist_ok=True)

        if self.is_alive():
            try:
                # 浏览器已经打开，保持使用
                return True
            except Exception as e:
                print(f"  [DEBUG] is_alive check failed: {e}")
                self._close_browser()

        # 关闭可能占用 session 的浏览器进程
        self._close_browser()
        self._cleanup_singleton_lock(session_dir)

        try:
            self.playwright = sync_playwright().start()
            self.context = self.playwright.chromium.launch_persistent_context(
                user_data_dir=session_dir,
                headless=browser_cfg.get("headless", True),
                viewport=browser_cfg["viewport"],
                user_agent=browser_cfg.get("user_agent"),
                locale=browser_cfg.get("locale", "zh-CN"),
                timezone_id=browser_cfg.get("timezone_id", "Asia/Shanghai"),
                args=browser_cfg.get("args", [])
            )
        except Exception as e:
            error_msg = str(e)
            print(f"  [DEBUG] Browser launch error: {error_msg[:200]}")
            if any(key in error_msg for key in [
                "ProcessSingleton",
                "Failed to create a ProcessSingleton",
                "profile appears to be in use",
                "Chromium has locked the profile",
                "SingletonLock",
            ]):
                # 锁文件导致的错误，清理后重试
                self._close_browser()
                self._cleanup_singleton_lock(session_dir)
                self.playwright = sync_playwright().start()
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=browser_cfg.get("headless", True),
                    viewport=browser_cfg["viewport"],
                    user_agent=browser_cfg.get("user_agent"),
                    locale=browser_cfg.get("locale", "zh-CN"),
                    timezone_id=browser_cfg.get("timezone_id", "Asia/Shanghai"),
                    args=browser_cfg.get("args", [])
                )
            else:
                raise

        self.page = self.context.pages[0] if self.context.pages else None
        if not self.page:
            return False
        ok, message = self._goto(cfg["website"]["url"], timeout=60000)
        if not ok:
            print(f"  [错误] 打开 CRM 首页失败: {message}")
            return False
        time.sleep(3)
        if self._is_current_page_logged_in():
            self.logged_in = True
            self.needs_navigation = True
        else:
            self.logged_in = False
        return True

    def _cleanup_singleton_lock(self, session_dir):
        """删除 Chromium 残留的单实例锁文件。"""
        singleton_names = {"SingletonLock", "SingletonSocket", "SingletonCookie"}
        profile_lock_paths = {
            os.path.abspath(os.path.join(session_dir, "LOCK")),
            os.path.abspath(os.path.join(session_dir, "Default", "LOCK")),
        }
        for root, _, files in os.walk(session_dir):
            for filename in files:
                lock_path = os.path.join(root, filename)
                if filename not in singleton_names and os.path.abspath(lock_path) not in profile_lock_paths:
                    continue
                if not os.path.lexists(lock_path):
                    continue
                try:
                    os.remove(lock_path)
                    print(f"  [DEBUG] 已清理浏览器锁文件: {lock_path}")
                except Exception as e:
                    print(f"  [DEBUG] 清理浏览器锁文件失败: {lock_path} {e}")

    def _is_current_page_logged_in(self):
        try:
            if not self.page:
                return False
            url = self.page.url.lower()
            body_text = self.page.inner_text("body")
            return (
                "login" not in url and
                (
                    "退出" in body_text or
                    "注销" in body_text or
                    "首页" in body_text or
                    "报表" in body_text or
                    "home" in url
                )
            )
        except Exception:
            return False

    def _wait_until_logged_in(self, timeout=18):
        end = time.time() + timeout
        while time.time() < end:
            if self._is_current_page_logged_in():
                self.logged_in = True
                self.needs_navigation = True
                return True
            time.sleep(0.8)
        return False

    def _reset_unfinished_login(self):
        """新登录开始前清掉半截验证码页或异常页，避免只能重启 Docker。"""
        if not self.is_alive():
            self.logged_in = False
            self.needs_navigation = True
            return False

        if self._is_current_page_logged_in():
            self.logged_in = True
            return True

        print("  [登录] 检测到未完成/异常登录状态，重开 CRM 登录入口...")
        self._close_browser()
        self.logged_in = False
        self.needs_navigation = True
        return False

    def login_step1(self, username, password):
        """Step1: 填账号密码 → 点登录 → 点发送验证码（你收到短信）"""
        if not HAS_PLAYWRIGHT:
            return False, "Playwright 未安装"
        with self.lock:
            try:
                if self._reset_unfinished_login():
                    return True, "已登录（会话有效）"
                if not self._ensure_browser():
                    return False, "浏览器启动失败"

                # 检查是否已登录（会话有效）
                time.sleep(1)
                if self._is_current_page_logged_in():
                    self.logged_in = True
                    return True, "已登录（会话有效）"

                # 在登录页，填账号密码
                time.sleep(1)

                # 填用户名
                for selector in [
                    "input[name='username']", "input[name='user']",
                    "input[name='logonUsername']", "#username", "#user", "input[type='text']",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click(); time.sleep(0.3)
                            el.press("Control+a"); time.sleep(0.1)
                            el.type(username, delay=100); time.sleep(0.5)
                            break
                    except:
                        continue

                # 填密码
                for selector in [
                    "input[name='password']", "input[name='pwd']",
                    "#password", "#pwd", "input[type='password']",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click(); time.sleep(0.3)
                            el.press("Control+a"); time.sleep(0.1)
                            el.type(password, delay=100); time.sleep(0.5)
                            break
                    except:
                        continue

                # 点击登录按钮
                for selector in [
                    "button[type='submit']", "input[type='submit']",
                    "#loginBtn", ".login-btn",
                    "button:has-text('登录')", "a:has-text('登录')",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click()
                            time.sleep(2)
                            break
                    except:
                        continue

                # 点击发送验证码
                sent_captcha = False
                for selector in [
                    "button:has-text('发送验证码')", "button:has-text('获取验证码')",
                    "a:has-text('发送验证码')", "a:has-text('获取验证码')",
                    "text=发送验证码",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click()
                            time.sleep(2)
                            sent_captcha = True
                            break
                    except:
                        continue

                for _ in range(20 if sent_captcha else 6):
                    captcha_input, _captcha_scope = self._find_captcha_input()
                    if captcha_input:
                        return True, "captcha_sent" if sent_captcha else "captcha_ready"
                    if self._wait_until_logged_in(timeout=1):
                        return True, "登录成功"
                    time.sleep(0.5)

                if sent_captcha:
                    return False, "验证码已发送，但验证码输入框未出现，请点击重新登录后再获取验证码"
                return False, "未找到发送验证码按钮，也未进入验证码输入页面"

            except Exception as e:
                return False, str(e)

    def _visible(self, el):
        try:
            return el and el.is_visible()
        except Exception:
            return False

    def _page_contexts(self):
        contexts = []
        pages = self.context.pages if self.context else []
        for page in reversed(pages):
            contexts.append((page, page))
            for frame in page.frames:
                if frame != page.main_frame:
                    contexts.append((page, frame))
        return contexts

    def _find_captcha_input(self):
        selectors = [
            "input[placeholder*='验证码']",
            "input[placeholder*='短信']",
            "input[name*='verify']",
            "input[name*='captcha']",
            "input[name*='code']",
            "input[id*='verify']",
            "input[id*='captcha']",
            "input[id*='code']",
        ]
        for page, scope in self._page_contexts():
            for selector in selectors:
                try:
                    el = scope.query_selector(selector)
                    if self._visible(el):
                        self.page = page
                        page.bring_to_front()
                        return el, scope
                except Exception:
                    pass

            try:
                inputs = scope.query_selector_all("input")
            except Exception:
                inputs = []
            for el in inputs:
                if not self._visible(el):
                    continue
                try:
                    attrs = " ".join([
                        el.get_attribute("id") or "",
                        el.get_attribute("name") or "",
                        el.get_attribute("placeholder") or "",
                        el.get_attribute("class") or "",
                    ]).lower()
                    input_type = (el.get_attribute("type") or "text").lower()
                    maxlength = el.get_attribute("maxlength") or ""
                    if input_type in ("password", "hidden"):
                        continue
                    if any(k in attrs for k in ["验证码", "verify", "captcha", "sms", "code", "auth"]):
                        self.page = page
                        page.bring_to_front()
                        return el, scope
                    if input_type in ("text", "tel", "number") and maxlength.isdigit() and int(maxlength) <= 8:
                        self.page = page
                        page.bring_to_front()
                        return el, scope
                except Exception:
                    pass
        return None, None

    def _click_confirm_near_captcha(self, scope, captcha_input=None):
        try:
            buttons = self.page.query_selector_all("button")
            for i, btn in enumerate(buttons):
                try:
                    if btn.is_visible():
                        text = btn.inner_text().strip()
                        if "确" in text and "定" in text:
                            box = btn.bounding_box()
                            if box:
                                x = box["x"] + box["width"] / 2
                                y = box["y"] + box["height"] / 2
                                self.page.mouse.click(x, y)
                            else:
                                btn.click()
                            print(f"  [OK] 已鼠标点击'确定' (索引{i}, 文本='{text}')")
                            return True
                except Exception:
                    pass
        except Exception as e:
            print(f"  [失败] 点击确定: {e}")
        return False

    def login_step2(self, captcha):
        """Step2: 收到验证码 → 等弹窗 → 填入 → 点确定 → 完成登录"""
        if not captcha:
            return False, "验证码不能为空"
        with self.lock:
            try:
                # 确保浏览器仍然活跃（可能是 step1 后重建的页面）
                if not self._ensure_browser():
                    return False, "浏览器启动失败"

                # 等待验证码输入框出现（最多10秒）
                captcha_input = None
                captcha_scope = None
                for _ in range(20):
                    captcha_input, captcha_scope = self._find_captcha_input()
                    if captcha_input:
                        break
                    time.sleep(0.5)

                if not captcha_input:
                    if self._wait_until_logged_in(timeout=5):
                        return True, "登录成功"
                    return False, "验证码输入框未出现，请点击重新登录后再获取验证码"

                captcha_input.click(); time.sleep(0.3)
                captcha_input.press("Control+a"); time.sleep(0.1)
                captcha_input.press("Backspace"); time.sleep(0.1)
                captcha_input.type(captcha, delay=150)
                time.sleep(0.5)

                # 点确定
                if not self._click_confirm_near_captcha(captcha_scope, captcha_input):
                    return False, "验证码已填入，但未找到确定按钮"
                if self._wait_until_logged_in(timeout=30):
                    return True, "登录成功"
                return False, "验证码可能错误，请重试，或点击重新登录"

            except Exception as e:
                return False, str(e)

    def login(self, username, password, captcha=None):
        """统一登录：可选传入验证码（有就自动填入并点确定）"""
        if not HAS_PLAYWRIGHT:
            return False, "Playwright 未安装"
        with self.lock:
            try:
                if not captcha and self._reset_unfinished_login():
                    return True, "已登录（会话有效）"

                if not self._ensure_browser():
                    return False, "浏览器启动失败"

                # 检查是否已登录
                time.sleep(1)
                if self._is_current_page_logged_in():
                    self.logged_in = True
                    return True, "已登录（会话有效）"

                # 在登录页，填账号密码
                time.sleep(1)

                # 填用户名
                for selector in [
                    "input[name='username']", "input[name='user']",
                    "input[name='logonUsername']", "#username", "#user", "input[type='text']",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click(); time.sleep(0.3)
                            el.press("Control+a"); time.sleep(0.1)
                            el.type(username, delay=100); time.sleep(0.5)
                            break
                    except:
                        continue

                # 填密码
                for selector in [
                    "input[name='password']", "input[name='pwd']",
                    "#password", "#pwd", "input[type='password']",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click(); time.sleep(0.3)
                            el.press("Control+a"); time.sleep(0.1)
                            el.type(password, delay=100); time.sleep(0.5)
                            break
                    except:
                        continue

                # 点击登录按钮
                for selector in [
                    "button[type='submit']", "input[type='submit']",
                    "#loginBtn", ".login-btn",
                    "button:has-text('登录')", "a:has-text('登录')",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click()
                            time.sleep(2)
                            break
                    except:
                        continue

                # ── 有验证码：直接填入并提交（step2不再重复点发送验证码） ──
                if captcha:
                    # 等待验证码输入框出现（最多10秒）
                    captcha_input = None
                    for _ in range(20):
                        for selector in [
                            "input[placeholder='验证码']",
                            "input[name='verifyCode']", "input[name='captcha']",
                        ]:
                            try:
                                el = self.page.query_selector(selector)
                                if el and el.is_visible():
                                    captcha_input = el
                                    break
                            except:
                                continue
                        if captcha_input:
                            break
                        time.sleep(0.5)
                    if captcha_input:
                        captcha_input.click(); time.sleep(0.3)
                        captcha_input.press("Control+a"); time.sleep(0.1)
                        captcha_input.press("Backspace"); time.sleep(0.1)
                        captcha_input.type(captcha, delay=150)
                        time.sleep(0.5)

                    # 点确定
                    buttons = self.page.query_selector_all("button")
                    for btn in buttons:
                        try:
                            if btn.is_visible():
                                text = btn.inner_text().strip()
                                if "确" in text and "定" in text:
                                    btn.click()
                                    time.sleep(5)  # 等待页面跳转
                                    break
                        except:
                            continue

                    if self._wait_until_logged_in(timeout=30):
                        return True, "登录成功"
                    return False, "验证码可能错误，请重试"

                # ── 无验证码：先点发送验证码，等验证码输入框出现 ──
                # 点击发送验证码
                for selector in [
                    "button:has-text('发送验证码')", "button:has-text('获取验证码')",
                    "a:has-text('发送验证码')", "a:has-text('获取验证码')",
                    "text=发送验证码",
                ]:
                    try:
                        el = self.page.query_selector(selector)
                        if el and el.is_visible():
                            el.click()
                            time.sleep(2)
                            break
                    except:
                        continue

                # 等待验证码输入框出现
                for _ in range(10):
                    for selector in [
                        "input[placeholder='验证码']",
                        "input[name='verifyCode']", "input[name='captcha']",
                    ]:
                        try:
                            el = self.page.query_selector(selector)
                            if el and el.is_visible():
                                return True, "captcha_required"
                        except:
                            continue
                    time.sleep(0.5)

                # 无验证码，检查是否直接登录成功
                url = self.page.url.lower()
                if "login" not in url:
                    self.logged_in = True
                    return True, "登录成功"

                return True, "please_login_manually"

            except Exception as e:
                return False, str(e)

    def _close_browser(self):
        try:
            if self.context:
                self.context.close()
                self.context = None
            if self.playwright:
                self.playwright.stop()
                self.playwright = None
            self.browser = None
            self.page = None
        except Exception:
            pass

    def logout(self):
        with self.lock:
            self._close_browser()
            self.logged_in = False
            self.needs_navigation = True  # 重置导航标记
            return True

    def cancel_login(self):
        """取消未完成的登录流程，避免下次登录接着半截验证码页面执行。"""
        with self.lock:
            if not self.logged_in:
                self._close_browser()
                self.needs_navigation = True
            return True

    def check_login_status(self):
        """检查浏览器当前是否已登录"""
        with self.lock:
            if not self.is_alive():
                if not self._ensure_browser():
                    return False, "浏览器未启动"
            try:
                url = self.page.url.lower()
                if "login" in url or "登录" in url:
                    self.logged_in = False
                    return False, "仍在登录页，未登录"
                time.sleep(1)
                if self._is_current_page_logged_in():
                    self.logged_in = True
                    return True, "已登录"
                else:
                    self.logged_in = False
                    return False, "无法确认登录状态，请检查浏览器"
            except Exception as e:
                self.logged_in = False
                return False, str(e)

    def switch_to_report_tab(self):
        """切换到报表标签页"""
        try:
            pages = self.context.pages
            print(f"  [INFO] 当前共 {len(pages)} 个标签页")
            if len(pages) == 0:
                return False

            report_page = None
            saw_report_page = False
            for i, p in enumerate(pages):
                try:
                    print(f"    标签页{i+1}: {p.url}")
                    if "/EcoCrystalReports/EcoCrystalRp" in p.url:
                        saw_report_page = True
                        error = self._get_report_page_error(p)
                        if error:
                            self.last_report_error = error
                            print(f"  [关闭] 标签页{i+1} 是报表错误页: {error}")
                            try:
                                p.close()
                            except Exception:
                                pass
                            continue
                        input_box = self._find_barcode_input(p)
                        if input_box:
                            self.last_report_error = ""
                            report_page = p
                            print(f"  [找到] 标签页{i+1} 包含可用报表输入框")
                            break
                        report_page = p
                        print(f"  [找到] 标签页{i+1} 包含报表 URL，继续等待输入框")
                except Exception:
                    pass

            if report_page:
                self.page = report_page
                self.page.bring_to_front()
                print("  [切换] 已切换到报表标签页")
                time.sleep(2)
                return True

            if saw_report_page:
                print("  [失败] 报表标签页存在，但没有可用页面")
                return False

            pages = self.context.pages
            if len(pages) == 1:
                print("  [失败] 没有新标签页被打开")
                return False

            self.page = pages[-1]
            self.page.bring_to_front()
            print("  [切换] 已切换到最后一个标签页")
            time.sleep(2)
            try:
                content = self.page.inner_text("body")
                if "输入 barcode" in content or "确定" in content or "EcoCrystalRp" in self.page.url:
                    print("  [成功] 当前标签页是报表页")
                    return True
                else:
                    print(f"  [警告] 当前标签页可能不是报表，URL: {self.page.url}")
            except Exception as e:
                print(f"  [错误] 检查标签页内容失败: {e}")
            return False
        except Exception as e:
            print(f"  [错误] switch_to_report_tab失败: {e}")
            return False

    def close_report_tab(self):
        """关闭报表标签页，只保留CRM列表页"""
        try:
            pages = self.context.pages
            print(f"  [DEBUG] 关闭前共有 {len(pages)} 个标签页")
            pages_to_close = []
            crm_list_page = None

            for p in pages:
                try:
                    url = p.url
                    if "crmportal.ecowaterchina" in url and "/report/reportlist" in url:
                        crm_list_page = p
                    else:
                        pages_to_close.append(p)
                except Exception:
                    pages_to_close.append(p)

            for p in pages_to_close:
                try:
                    print(f"  [DEBUG] 已关闭标签页: {p.url}")
                    p.close()
                except Exception:
                    pass
            time.sleep(1)

            if crm_list_page:
                self.page = crm_list_page
                self.page.bring_to_front()
                print("  [DEBUG] 已切换到CRM列表页")
                time.sleep(2)
            elif self.context.pages:
                self.page = self.context.pages[0]
                self.page.bring_to_front()
        except Exception as e:
            print(f"  [错误] close_report_tab失败: {e}")

    def _close_query_report_pages_locked(self):
        """只关闭条码查询报表页，保留 CRM 系统页面。"""
        if not self.context:
            return 0
        closed_count = 0
        for p in list(self.context.pages):
            try:
                if "/EcoCrystalReports/EcoCrystalRp" not in p.url:
                    continue
                print(f"  [空闲清理] 关闭条码报表页: {p.url}")
                p.close()
                closed_count += 1
            except Exception:
                pass

        crm_page = None
        for p in self.context.pages:
            try:
                url = p.url
                if "crmportal.ecowaterchina" in url and "/report/reportlist" in url:
                    crm_page = p
                    break
                if not crm_page and "crmportal.ecowaterchina" in url:
                    crm_page = p
            except Exception:
                pass

        if crm_page:
            self.page = crm_page
            try:
                self.page.bring_to_front()
            except Exception:
                pass
        elif closed_count and self.context:
            try:
                self.page = self.context.new_page()
                cfg = load_crm_config()
                self._goto(cfg["website"]["url"], timeout=60000)
            except Exception:
                pass

        if closed_count:
            self.needs_navigation = True
            self.last_report_error = ""
        return closed_count

    def close_idle_report_tabs(self, idle_seconds=REPORT_IDLE_TIMEOUT_SECONDS):
        """查询报表页空闲超过指定时间后关闭，保留 CRM 系统页。"""
        with self.lock:
            if not self.is_alive() or not self.report_last_used_at:
                return False, "没有需要清理的查询报表页"
            idle_for = time.time() - self.report_last_used_at
            if idle_for < idle_seconds:
                return False, f"查询报表空闲 {int(idle_for)} 秒，未达到清理时间"
            closed_count = self._close_query_report_pages_locked()
            if not closed_count:
                return False, "未发现打开的查询报表页"
            return True, f"已关闭 {closed_count} 个空闲查询报表页"

    def shutdown(self):
        with self.lock:
            self._close_browser()
            return True

    def _find_barcode_input(self, page=None):
        try:
            target_page = page or self.page
            inputs = target_page.query_selector_all("input[name='CrystalReportViewer1_p0DiscreteValue']")
            for el in inputs:
                try:
                    if el.is_visible():
                        return el
                except Exception:
                    pass
            return inputs[0] if inputs else None
        except Exception:
            return None

    def _get_report_page_error(self, page=None):
        try:
            target_page = page or self.page
            if "/EcoCrystalReports/EcoCrystalRp" not in target_page.url:
                return ""
            body_text = target_page.inner_text("body", timeout=2000).strip()
        except Exception:
            return ""

        compact_text = re.sub(r"\s+", " ", body_text)
        if (
            "CrystalReportViewer1.ReportSourceID" in body_text or
            "CrystalReportSource" in body_text or
            "找不到由" in body_text
        ):
            return compact_text[:160]
        if compact_text.startswith("错误"):
            return compact_text[:160]
        return ""

    def _find_open_report_page(self, require_input=False, close_errors=False):
        report_page = None
        try:
            for p in reversed(self.context.pages):
                try:
                    if "/EcoCrystalReports/EcoCrystalRp" not in p.url:
                        continue

                    error = self._get_report_page_error(p)
                    if error:
                        self.last_report_error = error
                        print(f"  [关闭] 报表错误页: {error}")
                        if close_errors:
                            try:
                                p.close()
                            except Exception:
                                pass
                        continue

                    input_box = self._find_barcode_input(p)
                    if input_box:
                        self.last_report_error = ""
                        return p
                    if not require_input and not report_page:
                        report_page = p
                except Exception:
                    pass
        except Exception:
            pass
        return report_page

    def _reload_report_for_input(self):
        """刷新当前报表页，尝试回到条码输入界面"""
        try:
            if "/EcoCrystalReports/EcoCrystalRp" not in self.page.url:
                return None
            print("  [复用] 刷新当前报表标签页，等待条码输入框...")
            self.page.reload(wait_until="domcontentloaded", timeout=30000)
            for i in range(20):
                time.sleep(1)
                error = self._get_report_page_error()
                if error:
                    self.last_report_error = error
                    print(f"  [失败] 报表刷新后是错误页: {error}")
                    return None
                input_box = self._find_barcode_input()
                try:
                    if input_box and input_box.is_visible():
                        print(f"  [成功] 刷新后找到条码输入框（等待{i+1}秒）")
                        return input_box
                except Exception:
                    pass
            print("  [失败] 刷新后仍未找到条码输入框")
        except Exception as e:
            print(f"  [错误] 刷新报表页失败: {e}")
        return None

    def _find_input_in_open_report_tabs(self):
        """逐个复用已打开的报表标签页，只有都不可用时才让上层重走流程"""
        try:
            report_pages = []
            for p in self.context.pages:
                try:
                    if "/EcoCrystalReports/EcoCrystalRp" in p.url:
                        report_pages.append(p)
                except Exception:
                    pass

            for p in reversed(report_pages):
                try:
                    self.page = p
                    self.page.bring_to_front()
                    time.sleep(1)
                    error = self._get_report_page_error()
                    if error:
                        self.last_report_error = error
                        print(f"  [关闭] 已打开的报表标签页是错误页: {error}")
                        try:
                            p.close()
                        except Exception:
                            pass
                        continue
                    input_box = self._find_barcode_input()
                    if input_box and input_box.is_visible():
                        print("  [复用] 已在打开的报表标签页找到条码输入框")
                        return input_box

                    input_box = self._reload_report_for_input()
                    if input_box and input_box.is_visible():
                        return input_box
                except Exception as e:
                    print(f"  [错误] 报表标签页不可用: {e}")
            print("  [失败] 所有已打开报表标签页都不可用")
        except Exception as e:
            print(f"  [错误] 检查已打开报表标签页失败: {e}")
        return None

    def _open_report_from_current_list(self):
        """从当前报表列表页双击打开'查询条码所有信息'"""
        try:
            list_page = self.page
            self.page = list_page
            self.page.bring_to_front()
            self.last_report_error = ""
            time.sleep(1)

            pages_before = len(self.context.pages)
            print(f"  [INFO] 打开前有 {pages_before} 个标签页")

            try:
                report_link = self.page.get_by_text("查询条码所有信息", exact=True).first
                report_link.scroll_into_view_if_needed(timeout=5000)
                box = report_link.bounding_box()
                if not box:
                    print("  [失败] 未找到'查询条码所有信息'位置")
                    return False
                x = box['x'] + box['width'] / 2
                y = box['y'] + box['height'] / 2
                self.page.mouse.dblclick(x, y)
                print("  [触发] mouse.dblclick 已触发，等待报表输入框...")
            except Exception as e:
                print(f"  [错误] mouse双击失败: {e}")
                return False

            report_page = None
            for i in range(18):
                time.sleep(1)
                report_page = self._find_open_report_page(require_input=True, close_errors=True)
                if report_page:
                    print(f"  [成功] 已发现可用报表标签页（当前共{len(self.context.pages)}个）")
                    break
                print(f"  [等待] 第{i+1}秒，等待报表标签页出现...")

            if not report_page:
                try:
                    result = self.page.evaluate("""() => {
                        const elements = document.querySelectorAll('*');
                        for (const el of elements) {
                            if (el.textContent.trim() === '查询条码所有信息') {
                                const dblclickEvent = new MouseEvent('dblclick', {
                                    bubbles: true, cancelable: true, view: window
                                });
                                el.dispatchEvent(dblclickEvent);
                                return true;
                            }
                        }
                        return false;
                    }""")
                    if result:
                        print("  [兜底] JS dblclick 已触发，继续等待报表输入框...")
                    else:
                        print("  [失败] JS未找到'查询条码所有信息'文字")
                except Exception as e:
                    print(f"  [错误] JS兜底双击失败: {e}")

                for i in range(15):
                    time.sleep(1)
                    report_page = self._find_open_report_page(require_input=True, close_errors=True)
                    if report_page:
                        print(f"  [成功] 兜底后发现可用报表标签页（当前共{len(self.context.pages)}个）")
                        break
                    print(f"  [等待] 兜底第{i+1}秒，等待报表标签页出现...")

            if not report_page:
                print("  [失败] 未能打开'查询条码所有信息'新标签页")
                return False

            self.page = report_page
            self.page.bring_to_front()
            self.last_report_error = ""
            time.sleep(2)
            return True
        except Exception as e:
            print(f"  [错误] 打开报表失败: {e}")
            return False

    def prepare_next_report(self):
        """优先复用当前CRM报表列表页，必要时再完整导航"""
        try:
            if (
                "crmportal.ecowaterchina" in self.page.url.lower() and
                "/report/reportlist" in self.page.url.lower()
            ):
                print("  [导航] 已在CRM报表列表页，优先复用当前页面...")
                time.sleep(2)

                try:
                    if "查询条码所有信息" in self.page.inner_text("body"):
                        print("  [成功] 当前列表页已有'查询条码所有信息'，直接打开")
                        return self._open_report_from_current_list()
                except Exception:
                    pass

                page2_ok = False
                try:
                    page2_link = self.page.get_by_text("2", exact=True).first
                    if page2_link:
                        page2_link.click()
                        time.sleep(2)
                        if "查询条码所有信息" in self.page.inner_text("body"):
                            print("  [成功] 点击页码'2'成功")
                            page2_ok = True
                except Exception as e:
                    print(f"  [失败] 页码'2'链接方式失败: {e}")

                if not page2_ok:
                    try:
                        next_btn = self.page.query_selector("button:has-text('下一页'), a:has-text('下一页')")
                        if next_btn:
                            next_btn.click()
                            time.sleep(2)
                            if "查询条码所有信息" in self.page.inner_text("body"):
                                print("  [成功] 点击'下一页'成功")
                                page2_ok = True
                    except Exception as e:
                        print(f"  [失败] '下一页'按钮方式失败: {e}")

                if page2_ok:
                    return self._open_report_from_current_list()

                print("  [导航] 当前列表页未找到目标报表，改走完整导航")

            return self.navigate_to_report()
        except Exception as e:
            print(f"  [错误] 准备报表失败: {e}")
            return False

    def navigate_to_report(self):
        """导航到报表查询页面"""
        try:
            cfg = load_crm_config()

            print("  ========== 开始导航 ==========")
            time.sleep(5)
            print(f"  [OK] 当前URL: {self.page.url}")

            if "crmportal.ecowaterchina" not in self.page.url.lower():
                print("  [跳转] 当前不在 CRM 主页，正在跳转...")
                ok, message = self._goto(cfg["website"]["url"], timeout=60000)
                if not ok:
                    print(f"  [失败] 跳转 CRM 主页失败: {message}")
                    return False
                time.sleep(3)

            # 点击报表管理
            print("  [步骤1] 点击'报表管理'")
            report_menu_ok = False
            for selector in ["text=报表管理", "a:has-text('报表管理')"]:
                try:
                    self.page.click(selector, timeout=5000)
                    time.sleep(3)
                    page_text = self.page.inner_text("body")
                    if "报表管理" in page_text:
                        print("  [成功] 点击'报表管理'成功")
                        report_menu_ok = True
                        break
                except Exception as e:
                    print(f"  [失败] selector={selector}, 错误: {e}")
            if not report_menu_ok:
                print("  [失败] 点击'报表管理'失败")
                return False

            # 点击水晶报表查看
            print("  [步骤2] 点击'水晶报表查看'")
            crystal_ok = False
            url_before = self.page.url
            for _ in range(10):
                try:
                    if self.page.is_visible("text=水晶报表查看"):
                        break
                except Exception:
                    pass
                time.sleep(1)
            for selector in ["text=水晶报表查看", "a:has-text('水晶报表查看')"]:
                try:
                    self.page.click(selector, timeout=10000)
                    time.sleep(3)
                    url_after = self.page.url
                    page_text = self.page.inner_text("body")
                    if url_after != url_before or "水晶报表" in page_text or "Crystal" in page_text:
                        print("  [成功] 点击'水晶报表查看'成功")
                        crystal_ok = True
                        break
                except Exception as e:
                    print(f"  [失败] selector={selector}, 错误: {e}")
            if not crystal_ok:
                print("  [失败] 点击'水晶报表查看'失败")
                return False

            # 这里保持在水晶报表列表页，继续翻到第2页。
            # main.py 也是点击“水晶报表查看”后直接在当前页翻页。
            time.sleep(2)

            # 翻到第2页，找到"查询条码所有信息"
            print("  [步骤3] 翻到第2页（查找'查询条码所有信息'）")
            time.sleep(3)
            page2_ok = False
            try:
                if "查询条码所有信息" in self.page.inner_text("body"):
                    page2_ok = True
                    print("  [成功] 已找到'查询条码所有信息'")
            except Exception:
                pass

            if not page2_ok:
                try:
                    spinbuttons = self.page.query_selector_all("spinbutton, input[type='number']")
                    for sb in spinbuttons:
                        try:
                            if sb.get_attribute("value") == "1":
                                sb.fill("2")
                                sb.press("Enter")
                                time.sleep(2)
                                if "查询条码所有信息" in self.page.inner_text("body"):
                                    print("  [成功] 输入页码2成功")
                                    page2_ok = True
                                    break
                        except Exception:
                            pass
                except Exception as e:
                    print(f"  [失败] 页码输入框方式失败: {e}")

            if not page2_ok:
                try:
                    next_btn = self.page.query_selector("button:has-text('下一页'), a:has-text('下一页')")
                    if next_btn:
                        next_btn.click()
                        time.sleep(2)
                        if "查询条码所有信息" in self.page.inner_text("body"):
                            print("  [成功] 点击'下一页'成功")
                            page2_ok = True
                except Exception as e:
                    print(f"  [失败] '下一页'按钮方式失败: {e}")

            if not page2_ok:
                try:
                    page2_link = self.page.get_by_text("2", exact=True).first
                    if page2_link:
                        page2_link.click()
                        time.sleep(2)
                        if "查询条码所有信息" in self.page.inner_text("body"):
                            print("  [成功] 点击页码'2'成功")
                            page2_ok = True
                except Exception as e:
                    print(f"  [失败] 页码'2'链接方式失败: {e}")

            if not page2_ok:
                print("  [失败] 无法自动翻到包含'查询条码所有信息'的页面")
                return False

            # 双击打开"查询条码所有信息"
            print("  [步骤4] 双击打开'查询条码所有信息'报表")
            if not self._open_report_from_current_list():
                return False
            print("  ========== 导航完成 ==========")
            return True

        except Exception as e:
            print(f"  [错误] 导航到报表失败: {e}")
            return False

    def query_barcode(self, barcode, log=None, output_dir=None):
        def emit(message, level='info'):
            if log:
                log(message, level)

        with self.lock:
            if not self.is_alive():
                emit("正在恢复 CRM 浏览器会话", "info")
                if not self._ensure_browser():
                    return False, "浏览器未启动"
            if not self.logged_in and not self._is_current_page_logged_in():
                return False, "CRM 当前未登录，请先登录 CRM"
            closed_count = self._close_query_report_pages_locked() if (
                self.report_last_used_at and time.time() - self.report_last_used_at >= REPORT_IDLE_TIMEOUT_SECONDS
            ) else 0
            if closed_count:
                emit(f"条码报表已超过 {REPORT_IDLE_TIMEOUT_SECONDS // 60} 分钟未使用，已关闭旧报表页", "warn")
            try:
                # 导航到报表页面（只在第一次查询时）
                if self.needs_navigation:
                    emit("准备打开条码报表页面")
                    print("  [导航] 准备打开报表页面...")
                    if not self.prepare_next_report():
                        if self.last_report_error:
                            return False, f"报表页面加载错误: {self.last_report_error}"
                        return False, "打开查询条码所有信息报表失败"
                    self.needs_navigation = False

                # 切换到报表标签页
                emit("切换到条码报表页")
                if not self.switch_to_report_tab():
                    emit("未找到已打开的条码报表页，重新打开条码报表", "warn")
                    self.needs_navigation = True
                    if not self.prepare_next_report():
                        if self.last_report_error:
                            return False, f"报表页面加载错误: {self.last_report_error}"
                        return False, "打开查询条码所有信息报表失败"
                    self.needs_navigation = False
                    if not self.switch_to_report_tab():
                        if self.last_report_error:
                            return False, f"报表页面加载错误: {self.last_report_error}"
                        return False, "未找到报表标签页"
                time.sleep(2)

                # 查找条码输入框
                print(f"  [DEBUG] 当前页面URL: {self.page.url}")
                page_text = self.page.inner_text("body")[:200]
                print(f"  [DEBUG] 页面内容: {page_text}")
                input_box = self._find_barcode_input()

                # 如果找不到，尝试重新导航
                if not input_box or not input_box.is_visible():
                    emit("未找到条码输入框，尝试刷新报表页", "warn")
                    input_box = self._reload_report_for_input()
                    if not input_box or not input_box.is_visible():
                        emit("刷新后仍未找到输入框，检查已打开报表标签", "warn")
                        input_box = self._find_input_in_open_report_tabs()
                        if not input_box or not input_box.is_visible():
                            emit("报表标签不可用，重新打开条码报表", "warn")
                            print("  [导航] 所有查询标签页不可用，关闭后重走流程...")
                            self.close_report_tab()
                            self.needs_navigation = True
                            if not self.prepare_next_report():
                                if self.last_report_error:
                                    return False, f"报表页面加载错误: {self.last_report_error}"
                                return False, "打开查询条码所有信息报表失败"
                            if not self.switch_to_report_tab():
                                if self.last_report_error:
                                    return False, f"报表页面加载错误: {self.last_report_error}"
                                return False, "未找到报表标签页"
                            time.sleep(5)
                            print(f"  [DEBUG] 重新导航后URL: {self.page.url}")
                            input_box = self._find_barcode_input()
                            if not input_box or not input_box.is_visible():
                                return False, "未找到条码输入框"

                has_loading = False
                for _ in range(30):
                    try:
                        imgs = self.page.query_selector_all("img")
                        has_loading = False
                        for img in imgs:
                            try:
                                if img.get_attribute("src") and "wait" in img.get_attribute("src"):
                                    parent = self.page.evaluate("(el) => el.offsetParent !== null", img)
                                    if parent:
                                        has_loading = True
                                        break
                            except:
                                pass
                        if not has_loading:
                            break
                    except:
                        break
                    time.sleep(1)
                if has_loading:
                    emit("报表加载遮罩长时间未消失，关闭旧报表页后换通道或重试", "warn")
                    self.needs_navigation = True
                    self._close_query_report_pages_locked()
                    return False, "报表仍在加载，已关闭旧报表页"

                input_box.click()
                time.sleep(0.3)
                input_box.press("Control+a")
                time.sleep(0.1)
                input_box.press("Backspace")
                time.sleep(0.1)
                emit(f"输入条码：{barcode}")
                input_box.type(barcode, delay=100)
                print(f"  已输入条码: {barcode}")

                emit("提交报表查询")
                print("  提交查询...")
                clicked = False
                try:
                    self.page.evaluate("if(typeof CrystalReportViewer1_submit === 'function') { CrystalReportViewer1_submit(); }")
                    clicked = True
                    time.sleep(1)
                except:
                    pass

                if not clicked:
                    try:
                        confirm_button = self.page.get_by_text("确定", exact=True).first
                        if confirm_button:
                            confirm_button.click()
                            clicked = True
                    except:
                        pass

                if not clicked:
                    try:
                        for link in self.page.query_selector_all("a"):
                            try:
                                if "确定" in (link.inner_text() or "").strip():
                                    link.click()
                                    clicked = True
                                    break
                            except:
                                pass
                    except:
                        pass

                if not clicked:
                    return False, "提交查询失败"

                emit("等待 CRM 生成报表")
                print("  等待报表处理...")
                max_wait = 60
                data_ready = False
                for wait_i in range(max_wait):
                    time.sleep(1)
                    try:
                        js_check = self.page.evaluate("""() => {
                            const iframes = document.querySelectorAll('iframe');
                            for (const iframe of iframes) {
                                try {
                                    const doc = iframe.contentDocument || iframe.contentWindow?.document;
                                    if (doc && doc.body) {
                                        const text = doc.body.innerText || '';
                                        if (text.length > 500 && !text.includes('正在处理') && !text.includes('请稍候')) {
                                            return { ready: true, length: text.length };
                                        }
                                    }
                                } catch(e) {}
                            }
                            return { ready: false, length: 0 };
                        }""")
                        if js_check.get('ready'):
                            data_ready = True
                            emit("报表数据已加载")
                            print(f"  报表已加载完成（{js_check['length']} 字符）")
                            break
                        elif (wait_i + 1) % 10 == 0:
                            emit(f"报表仍在加载，已等待 {wait_i + 1} 秒", "dim")
                            print(f"  ...已等待 {wait_i + 1} 秒，继续等待...")
                    except:
                        pass

                if not data_ready:
                    print("  等待超时，尝试提取可能的数据...")

                html_content = ""
                try:
                    html_content = self.page.evaluate("""() => {
                        const iframes = document.querySelectorAll('iframe');
                        let html = '';
                        for (const iframe of iframes) {
                            try {
                                const doc = iframe.contentDocument || iframe.contentWindow?.document;
                                if (doc && doc.body) html += doc.body.innerHTML;
                            } catch(e) {}
                        }
                        return html;
                    }""")
                except:
                    pass

                if html_content and len(html_content.strip()) > 1000:
                    html_dir = output_dir or BARCODE_DIR
                    is_temporary = bool(output_dir and os.path.abspath(output_dir) != os.path.abspath(BARCODE_DIR))
                    emit("保存临时条码查询结果" if is_temporary else "保存条码查询结果")
                    os.makedirs(html_dir, exist_ok=True)
                    output_file = os.path.join(html_dir, f"{barcode}.html")
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    try:
                        fields = extract_fields_from_html(output_file)
                        refresh_product_library_from_query_fields(barcode, fields, emit)
                    except Exception as e:
                        emit(f"条码匹配自动更新失败：{e}", "warn")
                    self.needs_navigation = False
                    return True, barcode
                else:
                    self.needs_navigation = True
                    self._close_query_report_pages_locked()
                    return False, "查询结果为空"

            except Exception as e:
                crash_message = self._handle_browser_exception(e)
                if crash_message:
                    return False, crash_message
                self.needs_navigation = True
                self._close_query_report_pages_locked()
                return False, str(e)
            finally:
                self.report_last_used_at = time.time()

    def _click_visible_crm_text(self, text, exact=False):
        try:
            clicked = self.page.evaluate("""({ text, exact }) => {
                const clean = (value) => (value || '').replace(/\\s+/g, '').trim();
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const wanted = clean(text);
                const nodes = Array.from(document.querySelectorAll('a,button,[role="menuitem"],li,span,div'))
                    .filter(visible)
                    .map(node => {
                        const nodeText = clean(node.innerText || node.textContent || '');
                        if (!nodeText) return null;
                        const matched = exact ? nodeText === wanted : nodeText.includes(wanted);
                        if (!matched) return null;
                        const clickable = node.closest('a,button,[role="menuitem"],li') || node;
                        const tag = clickable.tagName || '';
                        const href = clickable.getAttribute?.('href') || '';
                        let score = 0;
                        if (['A', 'BUTTON', 'LI'].includes(tag)) score += 4;
                        if (/menu|nav|sidebar|item/i.test(clickable.className || '')) score += 2;
                        if (/service|serve|serv|服务/i.test(href + ' ' + (clickable.className || ''))) score += 2;
                        score -= Math.min(nodeText.length / 20, 5);
                        return { node, clickable, score };
                    })
                    .filter(Boolean)
                    .sort((a, b) => b.score - a.score);
                const target = nodes[0]?.clickable || nodes[0]?.node;
                if (!target) return false;
                target.scrollIntoView({ block: 'center', inline: 'nearest' });
                target.click();
                return true;
            }""", {"text": text, "exact": bool(exact)})
            if clicked:
                time.sleep(1.2)
            return bool(clicked)
        except Exception:
            return False

    def _service_order_list_ready(self):
        try:
            body_text = self.page.inner_text("body", timeout=3000)
            compact = re.sub(r"\s+", "", body_text or "")
            return "服务单" in compact and "服务单号" in compact and any(
                word in compact for word in ["批量结单", "更多筛选", "请输入单号", "客户名称"]
            )
        except Exception:
            return False

    def _service_order_list_url(self):
        cfg = load_crm_config()
        return f"{cfg['website']['url'].rstrip('/')}/#/workOrder/list"

    def _open_service_order_list(self, emit):
        emit("打开 CRM 服务单列表...")
        if not self.is_alive():
            if not self._ensure_browser():
                return False, "浏览器未启动"
        self._close_query_report_pages_locked()
        target_url = self._service_order_list_url()
        if "#/workOrder/list" not in (self.page.url or "") or not self._service_order_list_ready():
            ok, message = self._goto(target_url, timeout=60000)
            if not ok:
                return False, message
        for _ in range(15):
            time.sleep(0.8)
            if self._service_order_list_ready():
                return True, ""
        if not self._is_current_page_logged_in():
            return False, "CRM 当前未登录，请先登录 CRM"
        body = ""
        try:
            body = re.sub(r"\s+", " ", self.page.inner_text("body", timeout=2000))[:240]
        except Exception:
            pass
        return False, f"未进入服务单列表，当前页面：{body or self.page.url}"

    def _set_service_search_keyword(self, service_no):
        return self.page.evaluate("""(serviceNo) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const clean = (text) => (text || '').replace(/\\s+/g, '').trim();
            const setValue = (input, value) => {
                const proto = input.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
                const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
                input.focus();
                if (setter) setter.call(input, value);
                else input.value = value;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
            };
            let target = Array.from(document.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                .filter(visible)
                .find(input => /请输入单号|服务单号|平台订单号|联系人|联系电话/.test(input.getAttribute('placeholder') || ''));
            if (target) {
                target.scrollIntoView({ block: 'center', inline: 'nearest' });
                setValue(target, serviceNo);
                target.setAttribute('data-codex-service-search', '1');
                return true;
            }
            const formItems = Array.from(document.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr, div'))
                .filter(visible);
            const scored = [];
            for (const item of formItems) {
                const text = clean(item.innerText || item.textContent || '');
                const inputs = Array.from(item.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .filter(input => !['hidden', 'checkbox', 'radio'].includes((input.getAttribute('type') || 'text').toLowerCase()));
                for (const input of inputs) {
                    const attrs = clean([
                        input.getAttribute('placeholder') || '',
                        input.getAttribute('name') || '',
                        input.getAttribute('id') || '',
                    ].join(' '));
                    let score = 0;
                    if (text.includes('服务单号')) score += 8;
                    if (text.includes('服务单')) score += 5;
                    if (attrs.includes('服务单号')) score += 8;
                    if (attrs.includes('服务单')) score += 5;
                    if (attrs.includes('搜索') || attrs.includes('查询') || attrs.includes('请输入')) score += 2;
                    if (score) scored.push({ input, score });
                }
            }
            target = scored.sort((a, b) => b.score - a.score)[0]?.input;
            if (!target) {
                target = Array.from(document.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .filter(input => !['hidden', 'checkbox', 'radio'].includes((input.getAttribute('type') || 'text').toLowerCase()))
                    .find(input => /服务单|搜索|查询|请输入/i.test([
                        input.getAttribute('placeholder') || '',
                        input.getAttribute('name') || '',
                        input.getAttribute('id') || '',
                    ].join(' ')));
            }
            if (!target) return false;
            target.scrollIntoView({ block: 'center', inline: 'nearest' });
            setValue(target, serviceNo);
            target.setAttribute('data-codex-service-search', '1');
            return true;
        }""", str(service_no))

    def _click_service_search_button(self):
        clicked = self.page.evaluate("""() => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const clean = (text) => (text || '').replace(/\\s+/g, '').trim();
            const input = document.querySelector('[data-codex-service-search="1"]')
                || Array.from(document.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .find(el => /请输入单号|服务单号|平台订单号|联系人|联系电话/.test(el.getAttribute('placeholder') || ''));
            const buttonText = (btn) => clean([
                btn.innerText || btn.textContent || '',
                btn.getAttribute('title') || '',
                btn.getAttribute('aria-label') || '',
                btn.getAttribute('class') || '',
            ].join(' '));
            const badButton = /重置|清空|更多|新增|删除|导出|导入|下载|返回|取消/;
            const goodButton = /查询|搜索|search|query/i;
            const buttons = Array.from(document.querySelectorAll('button:not([disabled]), a, .el-button'))
                .filter(visible)
                .map(btn => {
                    const text = buttonText(btn);
                    const rect = btn.getBoundingClientRect();
                    const inputRect = input ? input.getBoundingClientRect() : null;
                    let score = 0;
                    if (goodButton.test(text)) score += 100;
                    if (badButton.test(text)) score -= 100;
                    if (inputRect) {
                        const sameRow = Math.abs(rect.top - inputRect.top) < 90;
                        if (sameRow) score += 30;
                        if (sameRow && rect.left > inputRect.left) score += 20;
                        const distance = Math.abs(rect.top - inputRect.top) + Math.abs(rect.left - inputRect.right);
                        score -= Math.min(distance / 80, 30);
                        for (let root = input; root; root = root.parentElement) {
                            if (root === btn || root.contains(btn)) {
                                score += 10;
                                break;
                            }
                        }
                    }
                    return { btn, text, score };
                })
                .sort((a, b) => b.score - a.score);
            const target = buttons.find(row => row.score > 0)?.btn;
            if (target) {
                target.scrollIntoView({ block: 'center', inline: 'nearest' });
                target.click();
                return true;
            }
            if (input) {
                input.focus();
                input.dispatchEvent(new KeyboardEvent('keydown', { key: 'Enter', code: 'Enter', bubbles: true }));
                input.dispatchEvent(new KeyboardEvent('keyup', { key: 'Enter', code: 'Enter', bubbles: true }));
                return true;
            }
            return false;
        }""")
        if clicked:
            time.sleep(1.5)
            return True
        try:
            self.page.keyboard.press("Enter")
            time.sleep(1.5)
            return True
        except Exception:
            return False

    def _service_order_search_snapshot(self, service_no):
        try:
            return self.page.evaluate("""(serviceNo) => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const text = document.body ? (document.body.innerText || '') : '';
                const compact = text.replace(/\\s+/g, '');
                const loading = Array.from(document.querySelectorAll(
                    '.el-loading-mask,.el-loading-spinner,.ivu-spin,.ant-spin,.ant-spin-spinning,[aria-busy="true"],[class*="loading"],[class*="Loading"],img'
                )).some(el => {
                    if (!visible(el)) return false;
                    const src = el.getAttribute('src') || '';
                    const cls = el.getAttribute('class') || '';
                    const alt = el.getAttribute('alt') || '';
                    return /wait|loading|load/i.test(src + ' ' + cls + ' ' + alt);
                });
                const noData = /暂无数据|无数据|暂无记录|没有数据|NoData/i.test(compact);
                const found = compact.includes(String(serviceNo).replace(/\\s+/g, ''));
                return { found, loading, noData };
            }""", str(service_no))
        except Exception:
            return {}

    def _search_service_order(self, service_no):
        if not self._set_service_search_keyword(service_no):
            return False, "未找到服务单搜索输入框"
        if not self._click_service_search_button():
            return False, "未找到服务单查询按钮"
        stable_empty = 0
        for _ in range(35):
            time.sleep(0.8)
            snapshot = self._service_order_search_snapshot(service_no)
            if snapshot.get("found"):
                return True, ""
            if snapshot.get("loading"):
                stable_empty = 0
                continue
            if snapshot.get("noData"):
                stable_empty += 1
                if stable_empty >= 4:
                    return False, "服务单列表没有搜索结果"
            else:
                stable_empty = 0
        return False, "服务单搜索后未找到结果"

    def _open_service_order_detail(self, service_no):
        pages_before = len(self.context.pages) if self.context else 0
        clicked = self.page.evaluate("""(serviceNo) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const clean = (text) => (text || '').replace(/\\s+/g, ' ').trim();
            const candidates = Array.from(document.querySelectorAll('a,button,span,td,div'))
                .filter(visible)
                .filter(el => clean(el.innerText || el.textContent || '') === serviceNo || clean(el.innerText || el.textContent || '').includes(serviceNo))
                .map(el => {
                    const clickable = el.closest('a,button') || el;
                    const style = getComputedStyle(el);
                    let score = 0;
                    if (clickable.tagName === 'A' || clickable.tagName === 'BUTTON') score += 5;
                    if (/rgb\\(.*(37|64|68|83|96|112|196)|blue/i.test(style.color || '')) score += 2;
                    score -= Math.min(clean(el.innerText || el.textContent || '').length / 20, 5);
                    return { el, clickable, score };
                })
                .sort((a, b) => b.score - a.score);
            const target = candidates[0]?.clickable || candidates[0]?.el;
            if (!target) return false;
            target.scrollIntoView({ block: 'center', inline: 'nearest' });
            target.click();
            return true;
        }""", str(service_no))
        if not clicked:
            return False, "未找到可点击的蓝色服务单号"
        for _ in range(15):
            time.sleep(0.8)
            try:
                if len(self.context.pages) > pages_before:
                    self.page = self.context.pages[-1]
                    self.page.bring_to_front()
                body_text = self.page.inner_text("body", timeout=2000)
                current_url = self.page.url or ""
                compact = re.sub(r"\s+", "", body_text or "")
                if "#/workOrder/edit/" in current_url:
                    ready = self.page.evaluate("""(serviceNo) => {
                        const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                        const clean = (text) => (text || '').replace(/\\s+/g, '').trim();
                        const values = Array.from(document.querySelectorAll('input, textarea'))
                            .filter(visible)
                            .map(el => clean(el.value || el.innerText || el.textContent || ''));
                        const labels = Array.from(document.querySelectorAll('.el-form-item__label, .ivu-form-item-label, .ant-form-item-label, label, td, th, div'))
                            .filter(visible)
                            .map(el => clean(el.innerText || el.textContent || ''));
                        const hasServiceNo = values.includes(clean(serviceNo)) || clean(document.body.innerText || '').includes(clean(serviceNo));
                        const hasCloseStatus = labels.some(text => text === '结单状态' || text.includes('结单状态'))
                            && values.some(value => value === '已结单' || value === '未结单');
                        return { hasServiceNo, hasCloseStatus };
                    }""", str(service_no))
                    if (ready or {}).get("hasServiceNo") and (ready or {}).get("hasCloseStatus"):
                        return True, ""
                if ("服务工单" in compact and "结单确认" in compact) or ("服务工单号" in compact and "结单状态" in compact):
                    return True, ""
            except Exception:
                pass
        return False, "点击服务单号后未进入服务单详情"

    def _service_detail_closed_state(self):
        try:
            result = self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const clean = (text) => (text || '').replace(/\\s+/g, '').trim();
                const readFieldValue = (node) => {
                    const input = node.querySelector('input:not([type="hidden"]), textarea');
                    if (input) {
                        return clean(
                            input.value
                            || input.getAttribute('value')
                            || input.getAttribute('placeholder')
                            || input.innerText
                            || input.textContent
                            || ''
                        );
                    }
                    const content = node.querySelector('.el-form-item__content, .ivu-form-item-content, .ant-form-item-control, td:last-child, .value');
                    if (content) return clean(content.innerText || content.textContent || '');
                    return '';
                };
                const readNearLabelValue = (label, item) => {
                    const direct = readFieldValue(item);
                    if (direct && direct !== '结单状态' && direct !== '是否结单') return direct;
                    const labelRect = label.getBoundingClientRect();
                    const candidates = Array.from(document.querySelectorAll('input:not([type="hidden"]), textarea'))
                        .filter(visible)
                        .map(el => ({ el, rect: el.getBoundingClientRect(), value: clean(el.value || el.innerText || el.textContent || '') }))
                        .filter(row => row.value && Math.abs(row.rect.top - labelRect.top) < 60 && row.rect.left > labelRect.left)
                        .sort((a, b) => (Math.abs(a.rect.top - labelRect.top) - Math.abs(b.rect.top - labelRect.top)) || (a.rect.left - b.rect.left));
                    return candidates[0]?.value || direct || '';
                };
                const decide = (value, text) => {
                    const raw = clean(value || '');
                    const full = clean(text || '');
                    const normalized = raw.replace('结单状态', '');
                    if (normalized.includes('未结单')) return { found: true, closed: false, value: '未结单' };
                    if (normalized.includes('已结单')) return { found: true, closed: true, value: '已结单' };
                    if (full.includes('未结单')) return { found: true, closed: false, value: '未结单' };
                    if (full.includes('已结单')) return { found: true, closed: true, value: '已结单' };
                    return { found: false, closed: false, value: normalized || full };
                };
                const labelNodes = Array.from(document.querySelectorAll('.el-form-item__label, .ivu-form-item-label, .ant-form-item-label, label, td, th'))
                    .filter(visible)
                    .filter(el => {
                        const text = clean(el.innerText || el.textContent || '');
                        return text === '结单状态';
                    });
                for (const label of labelNodes) {
                    const item = label.closest('.el-form-item, .ivu-form-item, .ant-form-item, tr, .form-group') || label.parentElement;
                    if (!item || !visible(item)) continue;
                    const state = decide(readNearLabelValue(label, item), item.innerText || item.textContent || '');
                    if (state.found) return { closed: state.closed, value: state.value };
                }
                const nodes = Array.from(document.querySelectorAll('.el-form-item, .ivu-form-item, .ant-form-item, tr'))
                    .filter(visible);
                for (const node of nodes) {
                    const text = clean(node.innerText || node.textContent || '');
                    const label = node.querySelector('.el-form-item__label, .ivu-form-item-label, .ant-form-item-label, label, td:first-child, th:first-child');
                    const labelText = clean((label && (label.innerText || label.textContent || '')) || '');
                    if (labelText !== '结单状态' && text !== '结单状态已结单' && text !== '结单状态未结单') continue;
                    const state = decide(readFieldValue(node), text);
                    if (state.found) return { closed: state.closed, value: state.value };
                }
                return { closed: false, value: '' };
            }""")
            return bool((result or {}).get("closed")), _clean_export_value((result or {}).get("value"))
        except Exception:
            return False, ""

    def _close_current_service_order(self, log=None):
        closed = False
        value = ""
        for _ in range(12):
            closed, value = self._service_detail_closed_state()
            if value:
                break
            time.sleep(0.5)
        if log:
            log(f"详情页结单状态：{value or '未识别'}", "dim")
        if closed:
            return True, "already_closed", value or "已结单"
        if not (self._click_top_button("结单确认") or self._click_top_button("确认结单")):
            msg = self._visible_message()
            return False, "failed", msg or "未找到右上方结单确认按钮"
        return self._wait_service_close_result(log)

    def _wait_service_close_result(self, log=None, timeout=60):
        started = time.time()
        last_message = ""
        last_progress = 0
        confirmed_dialog = False
        failure_words = [
            "失败", "错误", "不能", "不可", "不存在", "异常", "无权限",
            "必填", "请选择", "未找到", "不允许", "无法",
        ]
        while time.time() - started < timeout:
            elapsed = int(time.time() - started)
            msg = self._visible_message()
            if msg:
                last_message = msg
                if any(word in msg for word in failure_words):
                    return False, "failed", msg

            clicked = self._click_dialog_button("确定") or self._click_dialog_button("确认")
            if clicked:
                confirmed_dialog = True
                msg = self._visible_message()
                if msg:
                    last_message = msg
                    if any(word in msg for word in failure_words):
                        return False, "failed", msg

            closed, value = self._service_detail_closed_state()
            if closed:
                return True, "closed", value or "已结单"

            if log and elapsed >= last_progress + 10:
                last_progress = elapsed
                if confirmed_dialog:
                    log(f"等待 CRM 结单结果，已等待 {elapsed} 秒", "dim")
                else:
                    log(f"等待 CRM 结单确认弹窗，已等待 {elapsed} 秒", "dim")
            time.sleep(1)

        if last_message:
            return False, "failed", last_message
        return False, "failed", f"结单确认后 {timeout} 秒内未收到 CRM 成功或失败提示"

    def close_service_orders(self, service_orders, log=None):
        def emit(message, level='info'):
            if log:
                log(message, level)

        with self.lock:
            if not self.is_alive():
                emit("正在恢复 CRM 浏览器会话", "info")
                if not self._ensure_browser():
                    return False, {"error": "浏览器未启动，请先登录 CRM", "results": []}
            if not self.logged_in and not self._is_current_page_logged_in():
                return False, {"error": "CRM 当前未登录，请先登录 CRM", "results": []}
            results = []
            try:
                ok, message = self._open_service_order_list(emit)
                if not ok:
                    return False, {"error": message, "results": results}
                total = len(service_orders or [])
                for index, row in enumerate(service_orders or [], 1):
                    service_no = _clean_export_value(row.get("service_no") if isinstance(row, dict) else row)
                    if not service_no:
                        continue
                    display_label = _service_order_display(row if isinstance(row, dict) else {"service_no": service_no})
                    emit(f"处理服务单 {index}/{total}：{display_label}", "info")
                    last_error = ""
                    result_row = {
                        "service_no": service_no,
                        "barcodes": list(row.get("barcodes") or []) if isinstance(row, dict) else [],
                        "customer_names": list(row.get("customer_names") or []) if isinstance(row, dict) else [],
                        "product_names": list(row.get("product_names") or []) if isinstance(row, dict) else [],
                        "display_label": display_label,
                        "success": False,
                        "status": "failed",
                        "message": "",
                    }
                    for attempt in range(1, 3):
                        if attempt > 1:
                            emit(f"{display_label} 准备重试 {attempt}/2", "warn")
                        ok, message = self._open_service_order_list(emit)
                        if not ok:
                            last_error = message
                            continue
                        ok, message = self._search_service_order(service_no)
                        if not ok:
                            last_error = message
                            continue
                        ok, message = self._open_service_order_detail(service_no)
                        if not ok:
                            last_error = message
                            continue
                        ok, status, message = self._close_current_service_order(emit)
                        if ok:
                            result_row.update({
                                "success": True,
                                "status": status,
                                "message": message or ("已结单" if status == "closed" else "原本已结单"),
                            })
                            level = "success" if status == "closed" else "dim"
                            emit(f"{display_label} {'结单成功' if status == 'closed' else '已是结单状态'}", level)
                            break
                        last_error = message
                    if not result_row["success"]:
                        result_row["message"] = last_error or "服务单结单失败"
                        emit(f"{display_label} 结单失败：{result_row['message']}", "error")
                    results.append(result_row)
                success_count = sum(1 for row in results if row.get("success"))
                failed_count = len(results) - success_count
                return failed_count == 0, {
                    "results": results,
                    "success_count": success_count,
                    "failed_count": failed_count,
                }
            except Exception as e:
                crash_message = self._handle_browser_exception(e)
                return False, {"error": crash_message or str(e), "results": results}

    def _move_create_url(self):
        cfg = load_crm_config()
        base = cfg["website"]["url"].rstrip("/")
        return f"{base}/?#/moveRelated/create"

    def _move_list_url(self):
        cfg = load_crm_config()
        base = cfg["website"]["url"].rstrip("/")
        return f"{base}/?#/moveRelated/list"

    def _has_transfer_create_form(self):
        try:
            body_text = self.page.inner_text("body", timeout=3000)
            return "移库类型" in body_text and bool(self._input_by_label("移库类型"))
        except Exception as e:
            self._handle_browser_exception(e)
            return False

    def _return_to_move_list(self):
        try:
            self.page.goto(self._move_list_url(), wait_until="domcontentloaded", timeout=20000)
            time.sleep(1)
            self.needs_navigation = True
            return True
        except Exception as e:
            self._handle_browser_exception(e)
            self.needs_navigation = True
            return False

    def _open_transfer_create_form(self, emit):
        emit("打开 CRM 移库单新增页面...")
        if not self._return_to_move_list() and not self.is_alive():
            emit(self._browser_crash_message(), "error")
            return False
        try:
            self.page.wait_for_function(
                "() => document.body && /移库单/.test(document.body.innerText || '')",
                timeout=10000
            )
        except Exception:
            pass
        if not self._click_top_button("新增"):
            self.page.goto(self._move_create_url(), wait_until="domcontentloaded", timeout=30000)
        for _ in range(12):
            time.sleep(0.8)
            if self._has_transfer_create_form():
                return True
        body_text = ""
        try:
            body_text = re.sub(r"\s+", " ", self.page.inner_text("body", timeout=2000))[:180]
        except Exception:
            pass
        emit(f"未识别到新增移库表单，当前页面：{body_text or self.page.url}", "warn")
        return False

    def _set_input_by_label(self, label, value):
        return self.page.evaluate("""({ label, value }) => {
            const clean = (text) => (text || '').replace(/\\s+/g, '');
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const wanted = clean(label);
            const candidates = Array.from(document.querySelectorAll('label, span, div, p, td')).reverse();
            for (const node of candidates) {
                if (!visible(node)) continue;
                const text = clean(node.innerText || node.textContent || '');
                if (!text || !text.includes(wanted)) continue;
                const roots = [];
                const closest = node.closest('.el-form-item, .ant-form-item, tr, .form-group');
                if (closest) roots.push(closest);
                for (let root = node.parentElement, i = 0; i < 5 && root; i++, root = root.parentElement) {
                    if (!roots.includes(root)) roots.push(root);
                }
                for (const root of roots) {
                    const input = Array.from(root.querySelectorAll('input:not([disabled]), textarea:not([disabled])')).find(visible);
                    if (!input) continue;
                    input.focus();
                    input.value = value;
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                    return true;
                }
            }
            return false;
        }""", {"label": label, "value": str(value)})

    def _input_by_label(self, label):
        handle = self.page.evaluate_handle("""(label) => {
            const clean = (text) => (text || '').replace(/\\s+/g, '');
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const wanted = clean(label);
            const formItems = Array.from(document.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr'))
                .filter(visible)
                .reverse();
            for (const item of formItems) {
                const labelEl = item.querySelector('.el-form-item__label, label, th, td:first-child');
                const labelText = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                if (!labelText || !labelText.includes(wanted)) continue;
                const input = Array.from(item.querySelectorAll('input:not([disabled]), textarea:not([disabled])')).find(visible);
                if (input) return input;
            }
            const candidates = Array.from(document.querySelectorAll('label, span, div, p, td')).reverse();
            for (const node of candidates) {
                if (!visible(node)) continue;
                const text = clean(node.innerText || node.textContent || '');
                if (!text || !text.includes(wanted)) continue;
                const roots = [];
                const closest = node.closest('.el-form-item, .ant-form-item, tr, .form-group');
                if (closest) roots.push(closest);
                for (let root = node.parentElement, i = 0; i < 5 && root; i++, root = root.parentElement) {
                    if (!roots.includes(root)) roots.push(root);
                }
                for (const root of roots) {
                    const inputs = Array.from(root.querySelectorAll('input:not([disabled]), textarea:not([disabled])'));
                    const input = inputs.find(visible);
                    if (input) return input;
                }
            }
            return null;
        }""", label)
        element = handle.as_element()
        return element if element else None

    def _click_form_control_by_label(self, label):
        try:
            return self.page.evaluate("""(label) => {
                const clean = (text) => (text || '').replace(/\\s+/g, '');
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const wanted = clean(label);
                const items = Array.from(document.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr'))
                    .filter(visible)
                    .reverse();
                for (const item of items) {
                    const labelEl = item.querySelector('.el-form-item__label, label, th, td:first-child');
                    const labelText = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                    if (!labelText || !labelText.includes(wanted)) continue;
                    const target = Array.from(item.querySelectorAll('.el-select, .el-input, input, textarea'))
                        .find(visible);
                    if (!target) return false;
                    target.scrollIntoView({ block: 'center', inline: 'center' });
                    target.click();
                    return true;
                }
                return false;
            }""", label)
        except Exception:
            return False

    def _input_value_by_label(self, label):
        try:
            return self.page.evaluate("""(label) => {
                const clean = (text) => (text || '').replace(/\\s+/g, '');
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const wanted = clean(label);
                const items = Array.from(document.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr'))
                    .filter(visible)
                    .reverse();
                for (const item of items) {
                    const labelEl = item.querySelector('.el-form-item__label, label, th, td:first-child');
                    const labelText = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                    if (!labelText || !labelText.includes(wanted)) continue;
                    const input = Array.from(item.querySelectorAll('input, textarea, .el-select__tags-text')).find(visible);
                    if (!input) return '';
                    return clean(input.value || input.innerText || input.textContent || '');
                }
                return '';
            }""", label)
        except Exception:
            return ""

    def _click_dropdown_text(self, text, timeout=10):
        end = time.time() + timeout
        while time.time() < end:
            try:
                clicked = self.page.evaluate("""(text) => {
                    const nodes = Array.from(document.querySelectorAll(
                        '.el-select-dropdown__item, .el-autocomplete-suggestion li, li, [role="option"]'
                    ));
                    const target = nodes.find(el => {
                        const visible = !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                        return visible && (el.innerText || el.textContent || '').trim().includes(text);
                    });
                    if (!target) return false;
                    target.click();
                    return true;
                }""", text)
                if clicked:
                    time.sleep(0.5)
                    return True
            except Exception:
                pass
            time.sleep(0.5)
        return False

    def _select_input_by_label(self, label, value):
        input_el = self._input_by_label(label)
        if not input_el:
            return False
        try:
            input_el.scroll_into_view_if_needed(timeout=3000)
            input_el.click(timeout=5000)
        except Exception:
            clicked = self.page.evaluate("""(el) => {
                if (!el) return false;
                el.focus();
                el.click();
                return true;
            }""", input_el)
            if not clicked:
                return False
        time.sleep(0.2)
        input_el.press("Control+a")
        input_el.type(str(value), delay=80)
        time.sleep(1.5)
        if self._click_dropdown_text(str(value), timeout=8):
            return True
        if self._click_form_control_by_label(label):
            time.sleep(0.5)
            input_el = self._input_by_label(label)
            if input_el:
                try:
                    input_el.press("Control+a")
                    input_el.type(str(value), delay=80)
                    time.sleep(1.2)
                except Exception:
                    pass
            if self._click_dropdown_text(str(value), timeout=5):
                return True
        input_el.press("ArrowDown")
        time.sleep(0.3)
        input_el.press("Enter")
        time.sleep(0.5)
        return bool(self._input_value_by_label(label))

    def _norm_code(self, value):
        return re.sub(r"\s+", "", _clean_export_value(value)).upper()

    def _click_field_action_by_label(self, label):
        clicked = self.page.evaluate("""(label) => {
            const clean = (text) => (text || '').replace(/\\s+/g, '');
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const wanted = clean(label);
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            dialogs.push(document.body);
            for (const scope of dialogs) {
                const items = Array.from(scope.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr'))
                    .filter(visible)
                    .reverse();
                for (const item of items) {
                    const labelEl = item.querySelector('.el-form-item__label, label, th, td:first-child');
                    const labelText = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                    if (!labelText || !labelText.includes(wanted)) continue;
                    const targets = Array.from(item.querySelectorAll(
                        'button:not([disabled]), a, .el-input__suffix, .el-input-group__append, .el-icon-search, i[class*="search"], svg'
                    )).filter(visible);
                    const target = targets.find(el => !['INPUT', 'TEXTAREA'].includes(el.tagName)) || targets[targets.length - 1];
                    if (target) {
                        target.click();
                        return true;
                    }
                }
            }
            return false;
        }""", label)
        if clicked:
            time.sleep(1)
        return bool(clicked)

    def _set_dialog_input_by_label(self, label, value):
        return self.page.evaluate("""({ label, value }) => {
            const clean = (text) => (text || '').replace(/\\s+/g, '');
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const setValue = (input, nextValue) => {
                const proto = input.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
                const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
                input.focus();
                if (setter) setter.call(input, nextValue);
                else input.value = nextValue;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                input.blur();
            };
            const wanted = clean(label);
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const items = Array.from(dialog.querySelectorAll('.el-form-item, .ant-form-item, .form-group, tr'))
                    .filter(visible)
                    .reverse();
                for (const item of items) {
                    const labelEl = item.querySelector('.el-form-item__label, label, th, td:first-child');
                    const labelText = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                    if (!labelText || !labelText.includes(wanted)) continue;
                    const input = Array.from(item.querySelectorAll('input:not([disabled]), textarea:not([disabled])')).find(visible);
                    if (!input) continue;
                    setValue(input, value);
                    return true;
                }
            }
            return false;
        }""", {"label": label, "value": str(value)})

    def _set_any_dialog_input(self, value):
        return self.page.evaluate("""(value) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const setValue = (input, nextValue) => {
                const proto = input.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
                const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
                input.focus();
                if (setter) setter.call(input, nextValue);
                else input.value = nextValue;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                input.blur();
            };
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const input = Array.from(dialog.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .find(el => {
                        const type = (el.getAttribute('type') || 'text').toLowerCase();
                        return !['hidden', 'checkbox', 'radio'].includes(type);
                    });
                if (!input) continue;
                setValue(input, value);
                return true;
            }
            return false;
        }""", str(value))

    def _clear_dialog_inputs(self):
        try:
            self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const setValue = (input, nextValue) => {
                    const proto = input.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
                    const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
                    if (setter) setter.call(input, nextValue);
                    else input.value = nextValue;
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                };
                const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                    .filter(visible)
                    .reverse();
                const dialog = dialogs[0];
                if (!dialog) return;
                Array.from(dialog.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .forEach(input => {
                        const type = (input.getAttribute('type') || 'text').toLowerCase();
                        if (!['hidden', 'checkbox', 'radio'].includes(type)) setValue(input, '');
                    });
            }""")
        except Exception:
            pass

    def _click_dialog_search_button(self):
        clicked = self.page.evaluate("""() => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const nodes = Array.from(dialog.querySelectorAll(
                    'button:not([disabled]), a, .el-button, .el-input-group__append, .el-icon-search, i[class*="search"], svg'
                )).filter(visible);
                const textTarget = nodes.find(el => {
                    const text = (el.innerText || el.textContent || '').replace(/\\s+/g, '');
                    return text && ['查询', '搜索', '查找', '检索'].some(word => text.includes(word));
                });
                const iconTarget = nodes.find(el => {
                    const cls = el.getAttribute('class') || '';
                    const title = el.getAttribute('title') || '';
                    const aria = el.getAttribute('aria-label') || '';
                    return /search|查询|搜索|查找|检索/i.test(cls + title + aria);
                });
                const target = (textTarget || iconTarget)?.closest?.('button,a,.el-button,.el-input-group__append') || textTarget || iconTarget;
                if (target) {
                    target.click();
                    return true;
                }
            }
            return false;
        }""")
        if clicked:
            time.sleep(1)
            return True
        try:
            self.page.keyboard.press("Enter")
            time.sleep(1)
            return True
        except Exception:
            return False

    def _set_dialog_search_keyword(self, value):
        return self.page.evaluate("""(value) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const setValue = (input, nextValue) => {
                const proto = input.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
                const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
                input.focus();
                if (setter) setter.call(input, nextValue);
                else input.value = nextValue;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
            };
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const inputs = Array.from(dialog.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .filter(input => {
                        const type = (input.getAttribute('type') || 'text').toLowerCase();
                        return !['hidden', 'checkbox', 'radio'].includes(type);
                    });
                const keywordInput = inputs.find(input => /搜索|查找|记录/.test(input.getAttribute('placeholder') || '')) || inputs[0];
                if (!keywordInput) continue;
                setValue(keywordInput, value);
                return true;
            }
            return false;
        }""", str(value))

    def _select_dialog_search_field(self, field_text):
        clicked = self.page.evaluate("""() => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const selects = Array.from(dialog.querySelectorAll('.el-select, [class*="select"]'))
                    .filter(visible);
                const target = selects.find(el => {
                    const input = el.querySelector('input');
                    const placeholder = input?.getAttribute('placeholder') || '';
                    return /请选择|产品|编码|名称/.test(placeholder) || input?.readOnly;
                }) || selects[0];
                if (!target) continue;
                target.scrollIntoView({ block: 'center', inline: 'center' });
                target.click();
                return true;
            }
            return false;
        }""")
        if not clicked:
            return False
        time.sleep(0.5)
        return self._click_dropdown_text(field_text, timeout=5)

    def _search_lookup_dialog(self, keyword, field_text="产品编码"):
        try:
            dialog = self.page.locator(".el-dialog:visible, [role='dialog']:visible, .modal:visible").last
            field_input = dialog.locator("input[placeholder*='请选择']").first
            if field_input.count():
                field_input.click(timeout=3000, force=True)
                time.sleep(0.5)
                if not self._click_dropdown_text(field_text, timeout=5):
                    return False
            keyword_input = dialog.locator("input[placeholder*='搜索'], input[placeholder*='查找'], input").first
            keyword_input.fill(str(keyword), timeout=3000)
            keyword_input.press("Enter")
            time.sleep(0.5)
            search_button = dialog.locator("button:has-text('搜索'), .el-button:has-text('搜索'), button:has-text('查询'), .el-button:has-text('查询')").first
            if search_button.count():
                search_button.click(timeout=3000, force=True)
            else:
                self._click_dialog_search_button()
            time.sleep(1.5)
            return True
        except Exception:
            pass
        if not self._select_dialog_search_field(field_text):
            return False
        time.sleep(0.3)
        if not self._set_dialog_search_keyword(keyword):
            return False
        self._click_dialog_search_button()
        time.sleep(1.5)
        return True

    def _click_product_search_result(self, product_code, product_name=""):
        clicked = self.page.evaluate("""({ productCode, productName }) => {
            const clean = (text) => (text || '').replace(/\\s+/g, ' ').trim();
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            for (const dialog of dialogs) {
                const rows = Array.from(dialog.querySelectorAll('tbody tr, .el-table__body tr, tr')).filter(visible);
                let fallback = null;
                for (const row of rows) {
                    const text = clean(row.innerText || row.textContent || '');
                    if (!text) continue;
                    if (productCode && text.includes(productCode)) {
                        row.click();
                        return true;
                    }
                    if (!fallback && productName && text.includes(productName)) {
                        fallback = row;
                    }
                }
                if (fallback) {
                    fallback.click();
                    return true;
                }
            }
            return false;
        }""", {"productCode": str(product_code or ""), "productName": str(product_name or "")})
        if clicked:
            time.sleep(0.8)
        return bool(clicked)

    def _visible_dialog_text(self):
        try:
            return self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                    .filter(visible)
                    .reverse();
                return dialogs
                    .map(dialog => (dialog.innerText || dialog.textContent || '').replace(/\\s+/g, ' ').trim())
                    .filter(Boolean)
                    .slice(0, 2)
                    .join(' | ')
                    .slice(0, 500);
            }""") or ""
        except Exception:
            return ""

    def _dialog_inputs_snapshot(self):
        try:
            return self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const clean = (text) => (text || '').replace(/\\s+/g, ' ').trim();
                const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                    .filter(visible)
                    .reverse();
                const dialog = dialogs[0];
                if (!dialog) return '';
                return Array.from(dialog.querySelectorAll('input:not([disabled]), textarea:not([disabled])'))
                    .filter(visible)
                    .map((input, idx) => {
                        const item = input.closest('.el-form-item, .ant-form-item, .form-group, tr');
                        const labelEl = item?.querySelector?.('.el-form-item__label, label, th, td:first-child');
                        const label = clean(labelEl ? (labelEl.innerText || labelEl.textContent || '') : '');
                        return `${idx + 1}.label=${label || '-'},placeholder=${input.getAttribute('placeholder') || '-'},name=${input.getAttribute('name') || '-'},value=${input.value || '-'}`;
                    })
                    .join('；');
            }""") or ""
        except Exception:
            return ""

    def _search_product_by_code(self, product_code, product_name=""):
        if not product_code:
            return False, "缺少产品编码，无法按编码搜索产品"
        if not self._click_field_action_by_label("产品名称"):
            return False, "未找到产品名称右侧搜索按钮"
        if self._search_lookup_dialog(product_code, "产品编码"):
            if self._click_product_search_result(product_code, product_name):
                self._click_dialog_button("确定")
                time.sleep(0.8)
                return True, ""
        if product_name and self._search_lookup_dialog(product_name, "产品名称"):
            if self._click_product_search_result(product_code, product_name):
                self._click_dialog_button("确定")
                time.sleep(0.8)
                return True, ""
        search_attempts = [
            ("产品编码", product_code),
            ("物料编码", product_code),
            ("产品代码", product_code),
            ("物料代码", product_code),
            ("编码", product_code),
            ("物料名称", product_name),
            ("产品名称", product_name),
            ("名称", product_name),
        ]
        tried = []
        for label, value in search_attempts:
            value = _clean_export_value(value)
            if not value:
                continue
            tried.append(f"{label}={value}")
            self._clear_dialog_inputs()
            if not self._set_dialog_input_by_label(label, value):
                if label == "编码" and not self._set_any_dialog_input(value):
                    continue
                if label != "编码":
                    continue
            self._click_dialog_search_button()
            time.sleep(1.5)
            if self._click_product_search_result(product_code, product_name):
                self._click_dialog_button("确定")
                time.sleep(0.8)
                return True, ""
        dialog_text = self._visible_dialog_text()
        suffix = f"，弹窗内容：{dialog_text}" if dialog_text else ""
        inputs = self._dialog_inputs_snapshot()
        input_suffix = f"，输入框：{inputs}" if inputs else ""
        return False, f"产品搜索结果未找到编码 {product_code}（已尝试 {'；'.join(tried)}）{input_suffix}{suffix}"

    def _select_product_with_code_check(self, product_name, product_code, emit=None):
        if not self._select_input_by_label("产品名称", product_name):
            return False, "未找到产品名称输入框"
        expected_code = self._norm_code(product_code)
        if not expected_code:
            return True, ""
        time.sleep(0.8)
        actual_code = self._norm_code(self._input_value_by_label("产品编码"))
        if actual_code == expected_code:
            return True, ""
        if emit:
            emit(f"产品名称带出的编码 {actual_code or '空'} 与条码编码 {expected_code} 不一致，改用编码搜索", "warn")
        ok, msg = self._search_product_by_code(expected_code, product_name)
        if not ok:
            return False, msg
        actual_code = self._norm_code(self._input_value_by_label("产品编码"))
        if actual_code and actual_code != expected_code:
            return False, f"编码搜索后仍不一致：当前 {actual_code}，应为 {expected_code}"
        return True, ""

    def _select_transfer_type(self, transfer_type):
        if not self._click_form_control_by_label("移库类型"):
            input_el = self._input_by_label("移库类型")
            if not input_el:
                return False
            try:
                input_el.scroll_into_view_if_needed(timeout=3000)
                input_el.click(timeout=5000)
            except Exception:
                clicked = self.page.evaluate("""(el) => {
                    if (!el) return false;
                    el.focus();
                    el.click();
                    return true;
                }""", input_el)
                if not clicked:
                    return False
        time.sleep(0.5)
        if self._click_dropdown_text(transfer_type, timeout=5):
            time.sleep(0.5)
            return self._input_value_by_label("移库类型") == transfer_type
        input_el = self._input_by_label("移库类型")
        if not input_el:
            return False
        input_el.press("ArrowDown")
        time.sleep(0.2)
        input_el.press("Enter")
        time.sleep(0.5)
        return self._input_value_by_label("移库类型") == transfer_type

    def _click_top_button(self, text):
        try:
            clicked = self.page.evaluate("""(text) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const clean = (value) => (value || '').replace(/\\s+/g, '').trim();
            const wanted = clean(text);
            const candidates = Array.from(document.querySelectorAll('button:not([disabled]), a, [role="button"]:not([disabled])'))
                .filter(visible)
                .filter(el => clean(el.innerText || el.textContent || el.getAttribute('aria-label') || '').includes(wanted))
                .map(el => {
                    const rect = el.getBoundingClientRect();
                    let score = 0;
                    const value = clean(el.innerText || el.textContent || el.getAttribute('aria-label') || '');
                    if (value === wanted) score += 10;
                    if (rect.top < 260) score += 4;
                    if ((el.tagName || '').toUpperCase() === 'BUTTON') score += 3;
                    if (/toolbar|top|header|operate|button/i.test(el.className || '')) score += 1;
                    return { el, score };
                })
                .sort((a, b) => b.score - a.score);
            const target = candidates[0]?.el;
            if (!target) return false;
            target.scrollIntoView({ block: 'center', inline: 'nearest' });
            target.click();
            return true;
            }""", text)
            if clicked:
                time.sleep(1)
                return True
        except Exception:
            pass
        try:
            btn = self.page.locator("button, a, [role='button']").filter(has_text=text).first
            btn.click(timeout=5000, force=True)
            time.sleep(1)
            return True
        except Exception:
            return False

    def _has_top_button(self, text):
        try:
            return bool(self.page.evaluate("""(text) => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                return Array.from(document.querySelectorAll('button, a'))
                    .some(el => visible(el) && (el.innerText || el.textContent || '').trim().includes(text));
            }""", text))
        except Exception:
            return False

    def _scroll_section_into_view(self, section_title):
        try:
            return bool(self.page.evaluate("""(sectionTitle) => {
                const clean = (text) => (text || '').replace(/\\s+/g, '');
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const targetText = clean(sectionTitle);
                const section = Array.from(document.querySelectorAll('*')).find(el => {
                    return visible(el) && clean(el.innerText || el.textContent || '') === targetText;
                });
                if (!section) return false;
                section.scrollIntoView({ block: 'center', inline: 'nearest' });
                return true;
            }""", section_title))
        except Exception:
            return False

    def _prepare_visible_dialog(self):
        try:
            self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                    .filter(visible)
                    .reverse();
                const dialog = dialogs[0];
                if (!dialog) return;
                dialog.scrollIntoView({ block: 'center', inline: 'nearest' });
                for (const node of [dialog, ...Array.from(dialog.querySelectorAll('.el-dialog__body, .el-form, [class*="body"], [class*="content"]'))]) {
                    if (node && typeof node.scrollTop === 'number') node.scrollTop = 0;
                }
            }""")
        except Exception:
            pass

    def _close_visible_dialogs(self):
        try:
            self.page.keyboard.press("Escape")
            time.sleep(0.4)
        except Exception:
            pass
        try:
            self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                    .filter(visible)
                    .reverse();
                for (const dialog of dialogs) {
                    const buttons = Array.from(dialog.querySelectorAll('button, a, .el-dialog__headerbtn'))
                        .filter(visible)
                        .reverse();
                    const target = buttons.find(btn => {
                        const text = (btn.innerText || btn.textContent || '').replace(/\\s+/g, '');
                        const cls = btn.getAttribute('class') || '';
                        return text.includes('取消') || text.includes('关闭') || /close|headerbtn/i.test(cls);
                    });
                    if (target) target.click();
                }
            }""")
            time.sleep(0.8)
        except Exception:
            pass

    def _wait_detail_dialog_closed(self, timeout=6):
        end = time.time() + timeout
        while time.time() < end:
            try:
                open_detail_dialog = self.page.evaluate("""() => {
                    const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                    return Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                        .filter(visible)
                        .some(dialog => /新增|修改|产品名称|产品条码|移库数量/.test(dialog.innerText || dialog.textContent || ''));
                }""")
                if not open_detail_dialog:
                    return True
            except Exception:
                return True
            time.sleep(0.3)
        return False

    def _wait_for_input_by_label(self, label, timeout=6):
        end = time.time() + timeout
        while time.time() < end:
            self._prepare_visible_dialog()
            input_el = self._input_by_label(label)
            if input_el:
                return input_el
            time.sleep(0.4)
        return None

    def _click_section_action(self, section_title, action_text):
        self._scroll_section_into_view(section_title)
        time.sleep(0.3)
        clicked = self.page.evaluate("""({ sectionTitle, actionText }) => {
            const clean = (text) => (text || '').replace(/\\s+/g, '');
            const sections = Array.from(document.querySelectorAll('*')).filter(el => {
                const visible = !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                return visible && clean(el.innerText || el.textContent || '') === clean(sectionTitle);
            });
            for (const section of sections) {
                section.scrollIntoView({ block: 'center', inline: 'nearest' });
                let root = section.parentElement;
                for (let i = 0; i < 6 && root; i++, root = root.parentElement) {
                    const buttons = Array.from(root.querySelectorAll('button, a')).reverse();
                    const target = buttons.find(btn => {
                        const visible = !!(btn.offsetWidth || btn.offsetHeight || btn.getClientRects().length);
                        return visible && (btn.innerText || btn.textContent || '').includes(actionText);
                    });
                    if (target) {
                        target.scrollIntoView({ block: 'center', inline: 'nearest' });
                        target.click();
                        return true;
                    }
                }
            }
            return false;
        }""", {"sectionTitle": section_title, "actionText": action_text})
        if clicked:
            time.sleep(1)
            self._prepare_visible_dialog()
        return clicked

    def _click_dialog_button(self, text):
        clicked = self.page.evaluate("""(text) => {
            const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
            const dialogs = Array.from(document.querySelectorAll('.el-dialog, [role="dialog"], .modal'))
                .filter(visible)
                .reverse();
            dialogs.push(document.body);
            for (const dialog of dialogs) {
                const buttons = Array.from(dialog.querySelectorAll('button, a'));
                const target = buttons.reverse().find(btn => {
                    return visible(btn) && (btn.innerText || btn.textContent || '').trim().includes(text);
                });
                if (target) {
                    target.click();
                    return true;
                }
            }
            return false;
        }""", text)
        if clicked:
            time.sleep(1.2)
        return clicked

    def _visible_message(self):
        try:
            text = self.page.evaluate("""() => {
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
                const selectors = [
                    '.el-message',
                    '.el-notification',
                    '.el-message-box',
                    '.el-dialog',
                    '.el-form-item__error',
                    '.toast',
                    '.ant-message',
                    '.ant-notification',
                    '.ant-modal',
                    '.ivu-message',
                    '.ivu-notice',
                    '.ivu-modal',
                    '[role="alert"]',
                    '[role="alertdialog"]',
                    '[role="dialog"]'
                ];
                const values = [];
                const pushText = (value) => {
                    let text = clean(value)
                        .replace(/^(提示|系统提示)\\s*/, '')
                        .replace(/\\s*(确定|确认|取消|关闭)\\s*$/g, '')
                        .trim();
                    if (!text || text.length < 2) return;
                    if (text.length > 500) text = text.slice(0, 500);
                    if (!values.includes(text)) values.push(text);
                };
                for (const sel of selectors) {
                    for (const el of Array.from(document.querySelectorAll(sel)).filter(visible)) {
                        pushText(el.innerText || el.textContent || '');
                    }
                }
                const buttons = Array.from(document.querySelectorAll('button, a, [role="button"]'))
                    .filter(visible)
                    .filter(el => /^(确定|确认|取消|关闭)$/.test(clean(el.innerText || el.textContent || '')));
                for (const btn of buttons) {
                    let root = btn.parentElement;
                    for (let i = 0; i < 5 && root; i++, root = root.parentElement) {
                        if (!visible(root)) continue;
                        const text = clean(root.innerText || root.textContent || '');
                        if (text && text.length < 500 && /结单|失败|错误|不能|不可|不存在|异常|提示|确认|确定/.test(text)) {
                            pushText(text);
                            break;
                        }
                    }
                }
                return values.join(' | ');
            }""")
            return re.sub(r"\s+", " ", text or "").strip()
        except Exception:
            return ""

    def _form_diagnostics(self):
        try:
            return self.page.evaluate("""() => {
                const clean = (text) => (text || '').replace(/\\s+/g, ' ').trim();
                const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
                const errors = Array.from(document.querySelectorAll('.el-form-item__error, .is-error .el-form-item__label'))
                    .filter(visible)
                    .map(el => clean(el.innerText || el.textContent || ''))
                    .filter(Boolean);
                const fields = Array.from(document.querySelectorAll('.el-form-item, tr, .form-group, .ant-form-item'))
                    .filter(visible)
                    .map(root => {
                        const label = clean((root.querySelector('label, .el-form-item__label, th, td:first-child') || {}).innerText || '');
                        if (!label) return null;
                        const valueNode = root.querySelector('input, textarea, .el-input__inner, .el-select__tags-text, .el-select .el-input__inner');
                        const value = valueNode ? clean(valueNode.value || valueNode.innerText || valueNode.textContent || '') : '';
                        const errorNode = root.querySelector('.el-form-item__error');
                        const error = errorNode ? clean(errorNode.innerText || errorNode.textContent || '') : '';
                        return { label, value, error };
                    })
                    .filter(row => row && (row.label || row.value || row.error))
                    .slice(0, 30);
                return { errors, fields };
            }""")
        except Exception:
            return {}

    def _format_form_diagnostics(self):
        diag = self._form_diagnostics() or {}
        parts = []
        errors = [str(x) for x in diag.get("errors", []) if str(x).strip()]
        if errors:
            parts.append("页面错误：" + " / ".join(errors[:8]))
        field_parts = []
        for row in diag.get("fields", [])[:16]:
            label = str(row.get("label") or "").strip()
            value = str(row.get("value") or "").strip()
            error = str(row.get("error") or "").strip()
            if not label:
                continue
            field = f"{label}={value or '空'}"
            if error:
                field += f"({error})"
            field_parts.append(field)
        if field_parts:
            parts.append("字段：" + "；".join(field_parts))
        return "；".join(parts)

    def _retry_detail_action(self, label, action, emit=None, attempts=3):
        last_msg = ""
        for attempt in range(1, attempts + 1):
            ok, msg = action()
            if ok:
                return True, ""
            last_msg = msg or "未知错误"
            if attempt >= attempts:
                break
            if emit:
                emit(f"{label}失败，准备重试 {attempt}/{attempts - 1}：{last_msg}", "warn")
            self._close_visible_dialogs()
            self._scroll_section_into_view("条码明细" if "条码" in label else "移库明细")
            time.sleep(1)
        return False, last_msg

    def _order_number(self):
        try:
            return self.page.evaluate("""() => {
                const clean = (text) => (text || '').replace(/\\s+/g, '');
                const candidates = Array.from(document.querySelectorAll('label, span, div, p, td'));
                for (const node of candidates) {
                    if (!clean(node.innerText || node.textContent || '').includes('移库单号')) continue;
                    let root = node;
                    for (let i = 0; i < 8 && root; i++, root = root.parentElement) {
                        const input = root.querySelector('input');
                        if (input && input.value) return input.value.trim();
                    }
                }
                return '';
            }""")
        except Exception:
            return ""

    def _save_transfer_header(self):
        if not self._click_top_button("保存"):
            return False, "未找到保存按钮"
        for _ in range(20):
            time.sleep(0.8)
            order_no = self._order_number()
            if order_no:
                return True, order_no
            msg = self._visible_message()
            if msg and any(key in msg for key in ["成功", "保存"]):
                order_no = self._order_number()
                return True, order_no or "已保存"
            if msg and any(key in msg for key in ["失败", "错误", "必填", "请选择", "不能为空"]):
                diag = self._format_form_diagnostics()
                return False, f"{msg}；{diag}" if diag else msg
        msg = self._visible_message() or "保存后未获取到移库单号"
        diag = self._format_form_diagnostics()
        return False, f"{msg}；{diag}" if diag else msg

    def _add_transfer_detail(self, group, emit=None):
        if not self._click_section_action("移库明细", "新增"):
            return False, "未找到移库明细新增按钮"
        if not self._wait_for_input_by_label("产品名称", timeout=8):
            diag = self._format_form_diagnostics()
            return False, f"未找到产品名称输入框{('；' + diag) if diag else ''}"
        ok, msg = self._select_product_with_code_check(
            group.get("product_name", ""),
            group.get("product_code", ""),
            emit
        )
        if not ok:
            return False, f"移库明细产品选择失败：{msg}"
        self._set_input_by_label("移库数量", group["quantity"])
        if not self._click_dialog_button("确定"):
            return False, "移库明细未找到确定按钮"
        if not self._wait_detail_dialog_closed():
            return False, "移库明细确定后弹窗未关闭"
        msg = self._visible_message()
        if msg and any(key in msg for key in ["失败", "错误", "必填", "请选择", "不能为空"]):
            return False, msg
        return True, ""

    def _add_barcode_detail(self, detail, emit=None):
        if not self._click_section_action("条码明细", "新增"):
            return False, "未找到条码明细新增按钮"
        if not self._wait_for_input_by_label("产品名称", timeout=8):
            diag = self._format_form_diagnostics()
            return False, f"条码明细未找到产品名称输入框{('；' + diag) if diag else ''}"
        product_name = detail.get("product_name", "")
        product_code = detail.get("product_code", "")
        if product_name:
            ok, msg = self._select_product_with_code_check(product_name, product_code, emit)
            if not ok:
                return False, f"条码明细产品选择失败：{msg}"
        filled = False
        for label in ["产品条码", "条码"]:
            if self._set_input_by_label(label, detail["barcode"]):
                filled = True
                break
        if not filled:
            return False, "条码明细未找到产品条码输入框"
        if not self._click_dialog_button("确定"):
            return False, "条码明细未找到确定按钮"
        if not self._wait_detail_dialog_closed():
            return False, "条码明细确定后弹窗未关闭"
        msg = self._visible_message()
        if msg and any(key in msg for key in ["失败", "错误", "必填", "请选择", "不能为空", "已安装"]):
            return False, msg
        return True, ""

    def _confirm_transfer(self):
        self._close_visible_dialogs()
        last_msg = ""
        for attempt in range(1, 4):
            if not self._click_top_button("确认"):
                return False, "未找到确认按钮"
            for _ in range(5):
                time.sleep(0.8)
                if not self._visible_message() and not self.page.locator(".el-dialog:visible").count():
                    break
                self._click_dialog_button("确定")

            for _ in range(10):
                time.sleep(0.8)
                msg = self._visible_message()
                if msg:
                    last_msg = msg
                if msg and any(key in msg for key in ["失败", "错误", "必填", "请选择", "不能为空", "已安装"]):
                    return False, msg
                if msg and any(key in msg for key in ["成功", "已确认", "移库成功"]):
                    return True, msg
                if not self._has_top_button("确认"):
                    return True, msg or "确认移库已提交"

            if not self._has_top_button("确认"):
                return True, last_msg or "确认移库已提交"
            self._close_visible_dialogs()

        return False, last_msg or "确认后未检测到成功提示，且确认按钮仍存在"

    def create_transfer(self, summary, distributor, transfer_type="移出", remark="", log=None):
        def emit(message, level='info'):
            if log:
                log(message, level)

        with self.lock:
            if not self.is_alive():
                emit("正在恢复 CRM 浏览器会话", "info")
                if not self._ensure_browser():
                    return False, "浏览器未启动，请先登录 CRM"
            if not self.logged_in and not self._is_current_page_logged_in():
                return False, "CRM 当前未登录，请先登录 CRM"
            try:
                def fail(message):
                    self._return_to_move_list()
                    return False, message

                if not self._open_transfer_create_form(emit):
                    return fail("打开移库单页面后未识别到新增表单")

                if transfer_type not in ("移入", "移出"):
                    transfer_type = "移出"
                emit(f"选择移库类型：{transfer_type}")
                if not self._select_transfer_type(transfer_type):
                    return fail("未找到移库类型输入框，已返回移库单列表")
                emit(f"选择目标分销商：{distributor}")
                if not self._select_input_by_label("分销商", distributor):
                    return fail("未找到分销商输入框，已返回移库单列表")
                if remark:
                    emit("填写备注")
                    self._set_input_by_label("备注", remark)

                emit("保存移库单表头，等待单号...")
                ok, result = self._save_transfer_header()
                if not ok:
                    return fail(result)
                order_no = result
                emit(f"移库单已保存：{order_no}", "success")

                added_products = []
                groups = summary.get("groups", [])
                for idx, group in enumerate(groups, start=1):
                    label = f"添加移库明细 {idx}/{len(groups)}：{group['product_name']} × {group['quantity']}"
                    emit(label)
                    ok, msg = self._retry_detail_action(
                        label,
                        lambda group=group: self._add_transfer_detail(group, emit),
                        emit
                    )
                    if not ok:
                        return fail(f"添加移库明细失败：{msg}")
                    added_products.append({
                        "product_name": group["product_name"],
                        "product_code": group["product_code"],
                        "quantity": group["quantity"],
                    })

                added_barcodes = []
                details = summary.get("details", [])
                for idx, detail in enumerate(details, start=1):
                    label = f"添加条码明细 {idx}/{len(details)}：{detail['barcode']}"
                    emit(label)
                    ok, msg = self._retry_detail_action(
                        label,
                        lambda detail=detail: self._add_barcode_detail(detail, emit),
                        emit
                    )
                    if not ok:
                        return fail(f"添加条码明细失败：{msg}")
                    added_barcodes.append(detail["barcode"])

                frozen_name = frozen_warehouse_name()
                if frozen_warehouse_save_only() and distributor == frozen_name:
                    emit(f"目标为{frozen_name}，移库单只保存不确认，等待审批", "success")
                    self._return_to_move_list()
                    return True, {
                        "order_no": order_no,
                        "products": added_products,
                        "barcodes": added_barcodes,
                        "confirmed": False,
                        "pending_approval": True,
                        "message": f"移库单已保存，{frozen_name}需审批，未点击确认",
                    }

                emit("点击确认移库，等待 CRM 提示...")
                ok, msg = self._confirm_transfer()
                if not ok:
                    return fail(f"确认移库失败：{msg}")
                emit(msg or "CRM 已提示移库成功", "success")
                self._return_to_move_list()

                return True, {
                    "order_no": order_no,
                    "products": added_products,
                    "barcodes": added_barcodes,
                    "confirmed": True,
                    "pending_approval": False,
                    "message": msg or self._visible_message(),
                }
            except Exception as e:
                crash_message = self._handle_browser_exception(e)
                if crash_message:
                    return False, crash_message
                return False, str(e)

class CRMWorker:
    """把所有 Playwright 操作固定到同一个线程里执行。"""
    def __init__(self, slot_id="default", session_dir=None):
        self.slot_id = slot_id
        self.session_dir = session_dir
        self.tasks = queue.Queue()
        self.state_lock = threading.Lock()
        self.browser_running = False
        self.remembered_logged_in_cache = _slot_remembered_logged_in(slot_id)
        self.logged_in_cache = False
        self.current_task = ""
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def _update_state(self, session):
        try:
            browser_running = session.is_alive()
            logged_in = session.logged_in
        except Exception:
            browser_running = False
            logged_in = False
        current_logged_in = bool(browser_running and logged_in)
        with self.state_lock:
            self.browser_running = browser_running
            self.logged_in_cache = current_logged_in
            self.remembered_logged_in_cache = bool(logged_in)
        _save_slot_logged_in(self.slot_id, logged_in)

    def _run(self):
        session = CRMSession(self.session_dir)
        while True:
            method_name, args, kwargs, result_queue = self.tasks.get()
            try:
                with self.state_lock:
                    self.current_task = method_name
                if method_name == "shutdown":
                    result = session.shutdown()
                    self._update_state(session)
                    result_queue.put((True, result))
                    return
                result = getattr(session, method_name)(*args, **kwargs)
                self._update_state(session)
                result_queue.put((True, result))
            except Exception as e:
                crash_message = session._handle_browser_exception(e)
                self._update_state(session)
                result_queue.put((False, crash_message or str(e)))
            finally:
                with self.state_lock:
                    self.current_task = ""

    def _call(self, method_name, *args, **kwargs):
        result_queue = queue.Queue(maxsize=1)
        self.tasks.put((method_name, args, kwargs, result_queue))
        success, result = result_queue.get()
        if success:
            return result
        raise RuntimeError(result)

    def is_alive(self):
        with self.state_lock:
            return self.browser_running

    @property
    def logged_in(self):
        with self.state_lock:
            return self.logged_in_cache

    @property
    def remembered_logged_in(self):
        with self.state_lock:
            return self.remembered_logged_in_cache

    @property
    def busy(self):
        with self.state_lock:
            return bool(self.current_task)

    def login(self, username, password, captcha=None):
        return self._call("login", username, password, captcha)

    def login_step1(self, username, password):
        return self._call("login_step1", username, password)

    def login_step2(self, captcha):
        return self._call("login_step2", captcha)

    def logout(self):
        result = self._call("logout")
        with self.state_lock:
            self.browser_running = False
            self.logged_in_cache = False
            self.remembered_logged_in_cache = False
        _save_slot_logged_in(self.slot_id, False)
        return result

    def cancel_login(self):
        result = self._call("cancel_login")
        with self.state_lock:
            if not self.logged_in_cache:
                self.browser_running = False
        return result

    def check_login_status(self):
        return self._call("check_login_status")

    def query_barcode(self, barcode, log=None, output_dir=None):
        if log:
            log(f"CRM 查询任务已加入队列：{barcode}", "dim")
        return self._call("query_barcode", barcode, log, output_dir)

    def close_service_orders(self, service_orders, log=None):
        if log:
            log(f"CRM 服务单结单任务已加入队列：{len(service_orders or [])} 个服务单", "dim")
        return self._call("close_service_orders", service_orders, log)

    def close_idle_report_tabs(self, idle_seconds=REPORT_IDLE_TIMEOUT_SECONDS):
        return self._call("close_idle_report_tabs", idle_seconds)

    def create_transfer(self, summary, distributor, transfer_type="移出", remark="", log=None):
        return self._call("create_transfer", summary, distributor, transfer_type, remark, log)

    def shutdown(self):
        try:
            result = self._call("shutdown")
        except Exception:
            result = False
        with self.state_lock:
            self.browser_running = False
            self.logged_in_cache = False
        return result

def _positive_int_env(name, default):
    try:
        value = int(os.environ.get(name, default))
    except (TypeError, ValueError):
        value = default
    return max(1, value)

def _runtime_data_base_dir():
    return os.environ.get("CRM_DATA_DIR", RUNTIME_BASE_DIR)

def _runtime_config_dir():
    return os.path.join(_runtime_data_base_dir(), "config")

def _runtime_config_path():
    return os.path.join(_runtime_config_dir(), "runtime_config.json")

def _migrate_root_config_file(filename):
    source = os.path.join(_runtime_data_base_dir(), filename)
    target = os.path.join(_runtime_config_dir(), filename)
    if os.path.abspath(source) == os.path.abspath(target) or not os.path.exists(source):
        return target
    try:
        os.makedirs(os.path.dirname(target), exist_ok=True)
        if not os.path.exists(target):
            shutil.copy2(source, target)
            message = "已迁移配置文件"
        else:
            message = "已清理旧配置文件"
        os.remove(source)
        print(f"  [DATA] {message}: {source} -> {target}")
    except Exception as e:
        print(f"  [DATA] 根目录配置文件迁移失败: {source} -> {target}: {e}")
    return target

def _normalize_worker_count(value, default=2):
    try:
        count = int(value)
    except (TypeError, ValueError):
        count = default
    return max(1, min(count, 10))

def _load_worker_count_config():
    defaults = {
        "query_workers": _normalize_worker_count(os.environ.get("CRM_QUERY_WORKERS"), 2),
        "transfer_workers": _normalize_worker_count(os.environ.get("CRM_TRANSFER_WORKERS"), 2),
    }
    config_path = _runtime_config_path()
    _migrate_root_config_file("runtime_config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                defaults["query_workers"] = _normalize_worker_count(data.get("query_workers"), defaults["query_workers"])
                defaults["transfer_workers"] = _normalize_worker_count(data.get("transfer_workers"), defaults["transfer_workers"])
        except Exception:
            pass
    return defaults

def _crm_session_base_dir():
    env_base = os.environ.get("CRM_SESSION_BASE")
    if env_base:
        return env_base
    try:
        return load_crm_config()["session"]["state_path"]
    except Exception:
        return os.path.join(RUNTIME_BASE_DIR, "session")

CRM_SLOT_STATE_FILE = os.path.join(_runtime_config_dir(), "crm_slot_state.json")
crm_slot_state_lock = threading.Lock()

def _load_crm_slot_state():
    try:
        _migrate_root_config_file("crm_slot_state.json")
        with crm_slot_state_lock:
            if not os.path.exists(CRM_SLOT_STATE_FILE):
                return {}
            with open(CRM_SLOT_STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _slot_remembered_logged_in(slot_id):
    row = _load_crm_slot_state().get(slot_id) or {}
    return bool(row.get("logged_in"))

def _save_slot_logged_in(slot_id, logged_in):
    try:
        with crm_slot_state_lock:
            data = {}
            if os.path.exists(CRM_SLOT_STATE_FILE):
                try:
                    with open(CRM_SLOT_STATE_FILE, "r", encoding="utf-8") as f:
                        loaded = json.load(f)
                    if isinstance(loaded, dict):
                        data = loaded
                except Exception:
                    data = {}
            data[slot_id] = {
                "logged_in": bool(logged_in),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            os.makedirs(os.path.dirname(CRM_SLOT_STATE_FILE), exist_ok=True)
            with open(CRM_SLOT_STATE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

class CRMWorkerPool:
    def __init__(self):
        worker_config = _load_worker_count_config()
        self.query_count = worker_config["query_workers"]
        self.transfer_count = worker_config["transfer_workers"]
        self.session_base = _crm_session_base_dir()
        self.workers = {}
        self.pool_lock = threading.Lock()
        self.query_slots = self._make_slots("query", self.query_count)
        self.transfer_slots = self._make_slots("transfer", self.transfer_count)

    def _ensure_worker(self, slot_id):
        if slot_id not in self.workers:
            session_dir = os.path.join(self.session_base, slot_id)
            self.workers[slot_id] = CRMWorker(slot_id, session_dir)
        return self.workers[slot_id]

    def _make_slots(self, prefix, count):
        slots = []
        for index in range(1, count + 1):
            slot_id = f"{prefix}-{index}"
            self._ensure_worker(slot_id)
            slots.append(slot_id)
        return slots

    def default_slot(self, kind="query"):
        with self.pool_lock:
            return (self.transfer_slots if kind == "transfer" else self.query_slots)[0]

    def get(self, slot_id=None, kind="query"):
        with self.pool_lock:
            default_slot = (self.transfer_slots if kind == "transfer" else self.query_slots)[0]
            slot_id = (slot_id or "").strip() or default_slot
            worker = self.workers.get(slot_id)
            if not worker:
                worker = self.workers[default_slot]
            return worker

    def normalize_slot(self, slot_id=None, kind="query"):
        worker = self.get(slot_id, kind)
        return worker.slot_id

    def slots_payload(self):
        with self.pool_lock:
            query_slots = list(self.query_slots)
            transfer_slots = list(self.transfer_slots)
        def rows(slot_ids, kind):
            return [{
                "id": slot_id,
                "label": f"{'移库' if kind == 'transfer' else '查询'}{index}",
                "kind": kind,
                "browser_running": self.workers[slot_id].is_alive(),
                "logged_in": self.workers[slot_id].logged_in,
                "remembered_logged_in": self.workers[slot_id].remembered_logged_in,
                "busy": self.workers[slot_id].busy,
            } for index, slot_id in enumerate(slot_ids, 1)]
        return {
            "query": rows(query_slots, "query"),
            "transfer": rows(transfer_slots, "transfer"),
            "defaults": {
                "query": query_slots[0],
                "transfer": transfer_slots[0],
            }
        }

    def resize(self, query_count, transfer_count):
        query_count = _normalize_worker_count(query_count, self.query_count)
        transfer_count = _normalize_worker_count(transfer_count, self.transfer_count)
        with self.pool_lock:
            self.query_count = query_count
            self.transfer_count = transfer_count
            self.query_slots = self._make_slots("query", self.query_count)
            self.transfer_slots = self._make_slots("transfer", self.transfer_count)
        return self.slots_payload()

    def shutdown(self):
        with self.pool_lock:
            workers = list(self.workers.values())
        for worker in workers:
            try:
                worker.shutdown()
            except Exception:
                pass

crm_pool = CRMWorkerPool()

def _desktop_startup_login_check_loop():
    """桌面应用启动后验证上次记住的 CRM 登录，避免显示假登录状态。"""
    if not IS_DESKTOP_APP or not STARTUP_LOGIN_AUTO_CHECK or not HAS_PLAYWRIGHT:
        return
    if STARTUP_LOGIN_CHECK_DELAY_SECONDS:
        time.sleep(STARTUP_LOGIN_CHECK_DELAY_SECONDS)

    try:
        with crm_pool.pool_lock:
            slots = (
                [("query", slot_id) for slot_id in crm_pool.query_slots] +
                [("transfer", slot_id) for slot_id in crm_pool.transfer_slots]
            )
        for kind, slot_id in slots:
            worker = crm_pool.get(slot_id, kind)
            if worker.logged_in or not worker.remembered_logged_in:
                continue
            try:
                print(f"  [启动检测] 正在验证 {slot_id} 上次 CRM 登录状态")
                success, message = worker.check_login_status()
                if success:
                    print(f"  [启动检测] {slot_id} CRM 会话有效")
                else:
                    print(f"  [启动检测] {slot_id} CRM 会话无效: {message}")
            except Exception as e:
                print(f"  [启动检测] {slot_id} CRM 会话检测失败: {e}")
            if STARTUP_LOGIN_CHECK_STAGGER_SECONDS:
                time.sleep(STARTUP_LOGIN_CHECK_STAGGER_SECONDS)
    except Exception as e:
        print(f"  [启动检测] CRM 启动检测失败: {e}")

def _idle_report_cleanup_loop():
    while True:
        time.sleep(REPORT_IDLE_CLEANUP_INTERVAL_SECONDS)
        try:
            with crm_pool.pool_lock:
                query_slots = list(crm_pool.query_slots)
            for slot_id in query_slots:
                worker = crm_pool.get(slot_id, "query")
                if not worker.is_alive():
                    continue
                closed, message = worker.close_idle_report_tabs(REPORT_IDLE_TIMEOUT_SECONDS)
                if closed:
                    print(f"  [空闲清理] {slot_id}: {message}")
        except Exception as e:
            print(f"  [空闲清理] 检查查询报表页失败: {e}")

threading.Thread(target=_idle_report_cleanup_loop, daemon=True).start()
threading.Thread(target=_desktop_startup_login_check_loop, daemon=True).start()
crm_session = crm_pool.get(kind="query")

DEFAULT_BATCH_RETRY_LIMIT = 5
MAX_BATCH_RETRY_LIMIT = 5

BATCH_LOG_LIMIT = 5000

batch_job_lock = threading.Lock()
batch_jobs = {}
latest_batch_job_by_slot = {}

library_query_lock = threading.Lock()
library_query_job = {
    'running': False,
    'done': False,
    'success': False,
    'barcode': '',
    'slot_id': '',
    'slot_label': '',
    'error': '',
    'logs': [],
    'started_at': '',
    'finished_at': '',
}

transfer_job_lock = threading.Lock()
transfer_jobs = {}
latest_transfer_job_by_slot = {}

service_close_job_lock = threading.Lock()
service_close_jobs = {}
latest_service_close_job_by_slot = {}

summary_job_lock = threading.Lock()
summary_jobs = {}
latest_summary_job_by_slot = {}

bulk_login_job_lock = threading.Lock()
bulk_login_jobs = {}
latest_bulk_login_job_by_scope = {}

def _empty_batch_job(slot_id=None, barcodes=None, retry_limit=DEFAULT_BATCH_RETRY_LIMIT):
    return {
        'job_id': uuid.uuid4().hex,
        'slot_id': slot_id or crm_pool.default_slot("query"),
        'running': False,
        'stop_requested': False,
        'barcodes': list(barcodes or []),
        'total': len(barcodes or []),
        'current': 0,
        'success': 0,
        'failed': 0,
        'retry_limit': retry_limit,
        'log_seq': 0,
        'logs': [],
        'results': [],
        'started_at': '',
        'finished_at': '',
    }

def _empty_transfer_job(slot_id=None, summary=None, distributor='', transfer_type='', remark=''):
    return {
        'job_id': uuid.uuid4().hex,
        'slot_id': slot_id or crm_pool.default_slot("transfer"),
        'running': False,
        'done': False,
        'success': False,
        'error': '',
        'result': None,
        'summary': summary,
        'distributor': distributor,
        'transfer_type': transfer_type,
        'remark': remark,
        'log_seq': 0,
        'logs': [],
        'started_at': '',
        'finished_at': '',
    }

def _empty_summary_job(slot_id=None):
    return {
        'job_id': uuid.uuid4().hex,
        'slot_id': slot_id or crm_pool.default_slot("transfer"),
        'running': False,
        'done': False,
        'success': False,
        'error': '',
        'summary': None,
        'log_seq': 0,
        'logs': [],
        'started_at': '',
        'finished_at': '',
    }

def _empty_service_close_job(slot_id=None, orders=None):
    orders = list(orders or [])
    return {
        'job_id': uuid.uuid4().hex,
        'slot_id': slot_id or crm_pool.default_slot("query"),
        'running': False,
        'done': False,
        'success': False,
        'error': '',
        'orders': orders,
        'total': len(orders),
        'current': 0,
        'closed_count': 0,
        'already_closed_count': 0,
        'failed_count': 0,
        'results': [],
        'missing': [],
        'no_service': [],
        'log_seq': 0,
        'logs': [],
        'started_at': '',
        'finished_at': '',
    }

def _empty_bulk_login_job(scope, slots=None):
    return {
        'job_id': uuid.uuid4().hex,
        'scope': scope,
        'running': False,
        'done': False,
        'success': False,
        'error': '',
        'username': '',
        'password': '',
        'stop_requested': False,
        'captcha': '',
        'step1_done': False,
        'log_seq': 0,
        'logs': [],
        'slots': [
            {
                'id': slot['id'],
                'kind': slot['kind'],
                'label': slot['label'],
                'status': 'pending',
                'message': '',
            }
            for slot in (slots or [])
        ],
        'started_at': '',
        'finished_at': '',
    }

batch_job = _empty_batch_job()
transfer_job = _empty_transfer_job()
summary_job = _empty_summary_job()
service_close_job = _empty_service_close_job()

def _normalize_retry_limit(value):
    try:
        retry_limit = int(value)
    except (TypeError, ValueError):
        retry_limit = DEFAULT_BATCH_RETRY_LIMIT
    return max(0, min(retry_limit, MAX_BATCH_RETRY_LIMIT))

def _brief_batch_error(error, limit=240):
    text = re.sub(r'\s+', ' ', str(error or '')).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."

def _safe_log_text(message):
    return str(message or '').replace('\xa0', ' ')

def format_duration_seconds(seconds):
    seconds = max(0, int(seconds or 0))
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    if hours:
        return f"{hours}小时{minutes}分{secs}秒"
    if minutes:
        return f"{minutes}分{secs}秒"
    return f"{secs}秒"

def _append_job_log_unlocked(job, message, level='dim', limit=300):
    job['log_seq'] = int(job.get('log_seq') or 0) + 1
    job['logs'].append({
        'id': job['log_seq'],
        'time': datetime.now().strftime('%H:%M:%S'),
        'message': _safe_log_text(message),
        'level': level,
    })
    job['logs'] = job['logs'][-limit:]

def _append_batch_log_unlocked(message, level='dim'):
    _append_job_log_unlocked(batch_job, message, level, BATCH_LOG_LIMIT)

def _batch_log(message, level='dim'):
    with batch_job_lock:
        _append_batch_log_unlocked(message, level)

def _job_log(lock, jobs, job_id, message, level='dim', limit=300):
    with lock:
        job = jobs.get(job_id)
        if job:
            _append_job_log_unlocked(job, message, level, limit)

def _batch_job_log(job_id, message, level='dim'):
    _job_log(batch_job_lock, batch_jobs, job_id, message, level, BATCH_LOG_LIMIT)

def _summary_job_log(job_id, message, level='dim'):
    _job_log(summary_job_lock, summary_jobs, job_id, message, level, 500)

def _transfer_job_log(job_id, message, level='dim'):
    _job_log(transfer_job_lock, transfer_jobs, job_id, message, level, 500)

def _service_close_job_log(job_id, message, level='dim'):
    _job_log(service_close_job_lock, service_close_jobs, job_id, message, level, 1000)

def _bulk_login_job_log(job_id, message, level='dim'):
    _job_log(bulk_login_job_lock, bulk_login_jobs, job_id, message, level, 1000)

def _slot_logged_in_message(message):
    return message in {'已登录（会话有效）', '登录成功', '登录成功（页面已跳转）'}

def _bulk_login_slots_for_scope(scope):
    payload = crm_pool.slots_payload()
    if scope == "query":
        slots = payload.get("query", [])
    elif scope == "transfer":
        slots = payload.get("transfer", [])
    else:
        scope = "all"
        slots = [*(payload.get("query", [])), *(payload.get("transfer", []))]
    return scope, [slot for slot in slots if not slot.get("logged_in")]

def _bulk_login_slot_snapshot(job):
    slots = list(job.get('slots') or [])
    waiting = [slot for slot in slots if slot.get('status') == 'waiting_captcha']
    active = [slot for slot in slots if slot.get('status') in {'pending', 'opening', 'submitting_captcha'}]
    success_count = sum(1 for slot in slots if slot.get('status') == 'logged_in')
    failed_count = sum(1 for slot in slots if slot.get('status') == 'failed')
    return slots, waiting, active, success_count, failed_count

def _finalize_bulk_login_job_if_ready(job_id):
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return
        if job.get('done'):
            return
        slots, waiting, active, success_count, failed_count = _bulk_login_slot_snapshot(job)
        if job.get('step1_done') and not waiting and not active:
            job['running'] = False
            job['done'] = True
            job['success'] = failed_count == 0
            job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if failed_count:
                job['error'] = f'仍有 {failed_count} 个通道未登录'
            _append_job_log_unlocked(
                job,
                f"批量登录完成，成功 {success_count} 个，失败 {failed_count} 个",
                'success' if failed_count == 0 else 'warn',
                1000,
            )

def _update_bulk_login_slot(job_id, slot_id, status, message=''):
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return None
        for slot in job.get('slots') or []:
            if slot.get('id') == slot_id:
                slot['status'] = status
                slot['message'] = message
                return dict(slot)
    return None

def _submit_bulk_login_slot(job_id, slot, captcha):
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return False
        username = job.get('username') or ''
        password = job.get('password') or ''
        target = next((row for row in job.get('slots') or [] if row.get('id') == slot['id']), None)
        if not target or target.get('status') != 'waiting_captcha':
            return False
        target['status'] = 'submitting_captcha'
        target['message'] = '正在提交验证码'
    _bulk_login_job_log(job_id, f"{slot['label']} 提交验证码", 'info')
    try:
        worker = crm_pool.get(slot['id'], slot.get('kind') or 'query')
        success, message = worker.login_step2(captcha)
        if (
            not success
            and username
            and password
            and "验证码输入框未出现" in str(message or "")
        ):
            _bulk_login_job_log(job_id, f"{slot['label']} 验证码框未出现，正在重开登录页后重试", 'warn')
            step1_success, step1_message = worker.login_step1(username, password)
            if step1_success and _slot_logged_in_message(step1_message):
                success, message = True, step1_message
            elif step1_success:
                success, message = worker.login_step2(captcha)
            else:
                success, message = False, step1_message
    except Exception as e:
        success, message = False, str(e)
    if success:
        _update_bulk_login_slot(job_id, slot['id'], 'logged_in', message or '登录成功')
        _bulk_login_job_log(job_id, f"{slot['label']} 登录成功", 'success')
        return True
    message_text = str(message or '验证码提交失败')
    keep_waiting = any(text in message_text for text in ["验证码可能错误", "验证码不能为空", "验证码已填入"])
    _update_bulk_login_slot(job_id, slot['id'], 'waiting_captcha' if keep_waiting else 'failed', message_text)
    _bulk_login_job_log(job_id, f"{slot['label']} 验证码提交失败：{message or '未知错误'}", 'error')
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if job:
            job['captcha'] = ''
    return False

def _submit_bulk_login_pending(job_id):
    while True:
        with bulk_login_job_lock:
            job = bulk_login_jobs.get(job_id)
            if not job or not job.get('captcha'):
                return
            captcha = job['captcha']
            pending = [dict(slot) for slot in job.get('slots') or [] if slot.get('status') == 'waiting_captcha']
        if not pending:
            _finalize_bulk_login_job_if_ready(job_id)
            return
        for slot in pending:
            _submit_bulk_login_slot(job_id, slot, captcha)
        _finalize_bulk_login_job_if_ready(job_id)

def _run_bulk_login_job(job_id, username, password):
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        slots = [dict(slot) for slot in (job or {}).get('slots') or []]
    _bulk_login_job_log(job_id, f"开始批量登录 {len(slots)} 个 CRM 通道", 'info')
    for slot in slots:
        with bulk_login_job_lock:
            job = bulk_login_jobs.get(job_id)
            if not job or job.get('stop_requested'):
                break
        _update_bulk_login_slot(job_id, slot['id'], 'opening', '正在打开登录页')
        _bulk_login_job_log(job_id, f"打开 {slot['label']} 登录页", 'info')
        success, message = False, ""
        for attempt in range(2):
            try:
                worker = crm_pool.get(slot['id'], slot.get('kind') or 'query')
                success, message = worker.login_step1(username, password)
            except Exception as e:
                success, message = False, str(e)
            if success:
                break
            if attempt == 0:
                _bulk_login_job_log(job_id, f"{slot['label']} 登录页未准备好，准备重试一次：{message or '未知错误'}", 'warn')
                time.sleep(2)

        if not success:
            _update_bulk_login_slot(job_id, slot['id'], 'failed', message or '登录失败')
            _bulk_login_job_log(job_id, f"{slot['label']} 登录失败：{message or '未知错误'}", 'error')
            continue

        if _slot_logged_in_message(message):
            _update_bulk_login_slot(job_id, slot['id'], 'logged_in', message or '已登录')
            _bulk_login_job_log(job_id, f"{slot['label']} 已登录", 'success')
            continue

        _update_bulk_login_slot(job_id, slot['id'], 'waiting_captcha', message or '等待验证码')
        _bulk_login_job_log(job_id, f"{slot['label']} 已进入验证码步骤", 'success')
        with bulk_login_job_lock:
            job = bulk_login_jobs.get(job_id)
            captcha = (job or {}).get('captcha') or ''
        if captcha:
            _submit_bulk_login_slot(job_id, slot, captcha)

    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if job:
            job['step1_done'] = True
            if job.get('stop_requested'):
                job['running'] = False
                job['done'] = True
                job['error'] = '批量登录已取消'
                job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                _append_job_log_unlocked(job, '批量登录已取消', 'warn', 1000)
    _submit_bulk_login_pending(job_id)
    _finalize_bulk_login_job_if_ready(job_id)

def _library_query_log(message, level='dim'):
    with library_query_lock:
        library_query_job['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'message': message,
            'level': level,
        })
        library_query_job['logs'] = library_query_job['logs'][-300:]

def _run_library_query_job(barcode, worker=None, slot_id='', slot_label=''):
    if is_disassembly_barcode(barcode):
        _library_query_log(f"已跳过拆机条码：{barcode}，CRM 不查询", 'warn')
        with library_query_lock:
            library_query_job['running'] = False
            library_query_job['done'] = True
            library_query_job['success'] = True
            library_query_job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            library_query_job['error'] = ''
        return
    _library_query_log(f"开始查询条码：{barcode}", 'info')
    attempted_slots = set()
    last_error = ""
    max_attempts = max(1, min(LIBRARY_QUERY_SLOT_RETRY_LIMIT, len(crm_pool.query_slots)))
    for attempt in range(1, max_attempts + 1):
        try:
            if not worker:
                worker, slot_id, slot_label, error = _select_idle_query_worker_desc(attempted_slots)
                if error:
                    last_error = error
                    break
                _library_query_log(f"已切换查询通道：{slot_label}", 'info')
            attempted_slots.add(slot_id or worker.slot_id)
            with library_query_lock:
                library_query_job['slot_id'] = slot_id or worker.slot_id
                library_query_job['slot_label'] = slot_label or _query_slot_label(worker.slot_id)
            if attempt > 1:
                _library_query_log(f"使用 {slot_label or _query_slot_label(worker.slot_id)} 重试查询（第 {attempt}/{max_attempts} 个通道）", 'warn')

            actual_worker = worker
            actual_slot_id = slot_id or actual_worker.slot_id
            existing_paths = existing_barcode_result_paths(barcode)
            had_metadata = barcode_metadata_exists(barcode)
            success, result = actual_worker.query_barcode(barcode, _library_query_log, TEMP_QUERY_DIR)
            if success:
                _mark_query_slot_healthy(actual_slot_id)
                with library_query_lock:
                    library_query_job['running'] = False
                    library_query_job['done'] = True
                    library_query_job['success'] = True
                    library_query_job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    library_query_job['error'] = ''
                removed = delete_temporary_query_result(
                    barcode,
                    _library_query_log,
                    keep_paths=existing_paths,
                    keep_metadata=had_metadata,
                )
                if not removed and (existing_paths or had_metadata):
                    _library_query_log("该条码原本已在结果管理中，保留原查询文件", 'dim')
                _library_query_log(f"条码查询完成：{barcode}", 'success')
                return

            last_error = _brief_batch_error(result, 800)
            _mark_query_slot_unhealthy(actual_slot_id, result)
            _library_query_log(f"{slot_label or _query_slot_label(actual_slot_id)} 查询失败：{last_error}", 'error')
            if attempt < max_attempts:
                _library_query_log("将自动换一个未使用的查询通道重试", 'warn')
                worker = None
                slot_id = ''
                slot_label = ''
                continue
            break
        except Exception as e:
            actual_slot_id = slot_id or getattr(worker, 'slot_id', '')
            last_error = _brief_batch_error(e, 800)
            _mark_query_slot_unhealthy(actual_slot_id, e)
            _library_query_log(f"{slot_label or _query_slot_label(actual_slot_id)} 查询出错：{last_error}", 'error')
            if attempt < max_attempts:
                _library_query_log("将自动换一个未使用的查询通道重试", 'warn')
                worker = None
                slot_id = ''
                slot_label = ''
                continue
            break

    with library_query_lock:
        library_query_job['running'] = False
        library_query_job['done'] = True
        library_query_job['success'] = False
        library_query_job['error'] = last_error or '条码查询失败'
        library_query_job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _library_query_log(f"条码查询失败：{last_error or '未找到可用查询通道'}", 'error')

def _transfer_log(message, level='dim'):
    with transfer_job_lock:
        transfer_job['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'message': message,
            'level': level,
        })
        transfer_job['logs'] = transfer_job['logs'][-300:]

def _summary_log(message, level='dim'):
    with summary_job_lock:
        summary_job['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'message': message,
            'level': level,
        })
        summary_job['logs'] = summary_job['logs'][-300:]

def _finish_summary_job(success, error='', summary=None):
    with summary_job_lock:
        summary_job['running'] = False
        summary_job['done'] = True
        summary_job['success'] = bool(success)
        summary_job['error'] = error
        summary_job['summary'] = summary
        summary_job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def _summary_error_from_result(summary):
    if summary.get('missing'):
        return '部分条码没有查询结果，也没有匹配到条码前缀，请先维护条码匹配或查询一次该产品条码'
    if summary.get('incomplete'):
        return '部分条码缺少产品名称或产品编码，无法自动汇总'
    return ''

def _exclude_unmatched_transfer_barcodes(summary):
    missing = [_clean_export_value(x) for x in summary.get('missing', []) if _clean_export_value(x)]
    if not missing:
        return []
    missing_set = set(missing)
    summary['details'] = [
        row for row in summary.get('details', [])
        if _clean_export_value(row.get('barcode')) not in missing_set
    ]
    summary['groups'] = [
        row for row in summary.get('groups', [])
        if row.get('barcodes')
    ]
    summary['missing'] = []
    summary['total'] = len(summary.get('details', []))
    excluded = summary.setdefault('excluded', [])
    for barcode in missing:
        if barcode not in excluded:
            excluded.append(barcode)
    summary['excluded_unmatched'] = missing
    return missing

def _log_transfer_summary_products(log, summary):
    groups = summary.get('groups') or []
    total = len(groups)
    for index, group in enumerate(groups, 1):
        prefixes = "、".join([
            _clean_export_value(prefix)
            for prefix in group.get('matched_prefixes', [])
            if _clean_export_value(prefix)
        ]) or "-"
        product_model = _clean_export_value(group.get('product_model')) or _clean_export_value(group.get('product_name')) or "-"
        product_name = _clean_export_value(group.get('product_name')) or "-"
        product_code = _clean_export_value(group.get('product_code')) or "-"
        quantity = group.get('quantity') or 0
        if product_model and product_model != product_name:
            label = f"产品型号 {product_model}，产品名称 {product_name}"
        else:
            label = f"产品名称 {product_name}"
        log(
            f"汇总产品 {index}/{total}：前缀 {prefixes}，{label}，编码 {product_code}，数量 {quantity}",
            'success'
        )

def _run_summary_job(job_id, worker, barcodes, transfer_type, distributor, excluded=None):
    def log(message, level='dim'):
        _summary_job_log(job_id, message, level)

    def finish(success, error='', summary=None):
        with summary_job_lock:
            job = summary_jobs.get(job_id)
            if not job:
                return
            job['running'] = False
            job['done'] = True
            job['success'] = bool(success)
            job['error'] = error
            job['summary'] = summary
            job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        barcodes, filtered = filter_disassembly_barcodes(barcodes)
        excluded = list(excluded or []) + filtered
        if excluded:
            log(f"已排除拆机条码 {len(excluded)} 个，不查询不移库：{', '.join(excluded[:10])}", 'warn')
        if not barcodes:
            error = '输入的条码都是拆机条码，无需汇总或移库'
            log(error, 'warn')
            finish(False, error, {'total': 0, 'details': [], 'groups': [], 'missing': [], 'incomplete': [], 'blocked': [], 'excluded': excluded})
            return
        log(f"开始汇总预览，共 {len(barcodes)} 个条码", 'info')
        log("开始检查条码匹配前缀和已有查询结果", 'info')
        auto_library = {'queried': [], 'failed': []}
        representatives = _missing_product_library_representatives(barcodes)
        if representatives:
            items = "，".join([f"{prefix}←{barcode}" for prefix, barcode in representatives.items()])
            log(f"发现 {len(representatives)} 个缺失前缀：{items}", 'info')
            log("将自动逐个查询代表条码补充条码匹配；离开本页面不会停止后台汇总，可返回移库页查看日志", 'dim')
            ready, ready_message = _crm_ready_for_auto_query(worker)
            if not ready:
                log(ready_message, 'error')
                finish(False, ready_message)
                return
            auto_library = ensure_product_library_for_barcodes(barcodes, log, worker)
            if auto_library.get('failed'):
                failed_items = "，".join([
                    f"{row.get('prefix')}←{row.get('barcode')}"
                    for row in auto_library.get('failed', [])
                ])
                log(f"仍有前缀自动补充失败：{failed_items}", 'warn')
        else:
            log("条码匹配已覆盖所有条码前缀，不需要自动查询", 'success')

        log("开始按产品名称和编码汇总移库明细", 'info')
        summary = build_transfer_summary(barcodes, transfer_type, distributor)
        summary['excluded'] = excluded
        summary['auto_library'] = auto_library
        excluded_unmatched = _exclude_unmatched_transfer_barcodes(summary)
        if excluded_unmatched:
            log(
                f"已临时排除 {len(excluded_unmatched)} 个查不到产品信息的条码：{', '.join(excluded_unmatched[:10])}",
                'warn'
            )
        if not summary.get('groups'):
            error = '本次没有可移库条码，未匹配到产品信息的条码已临时排除'
            log(error, 'error')
            finish(False, error, summary)
            return
        error = _summary_error_from_result(summary)
        if error:
            log(error, 'error')
            finish(False, error, summary)
            return
        log(f"汇总完成：产品 {len(summary.get('groups', []))} 条，条码 {summary.get('total', 0)} 个", 'success')
        _log_transfer_summary_products(log, summary)
        finish(True, '', summary)
    except Exception as e:
        error = _brief_batch_error(e, 800)
        log(f"汇总预览出错：{error}", 'error')
        finish(False, error)

def _run_transfer_job(job_id, worker, summary, distributor, transfer_type, remark):
    def log(message, level='dim'):
        _transfer_job_log(job_id, message, level)

    log(f"开始提交移库：{transfer_type}，分销商 {distributor}", 'info')
    try:
        success, result = worker.create_transfer(summary, distributor, transfer_type, remark, log)
        with transfer_job_lock:
            job = transfer_jobs.get(job_id)
            if not job:
                return
            job['running'] = False
            job['done'] = True
            job['success'] = bool(success)
            job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if success:
                job['result'] = result
                job['error'] = ''
            else:
                job['error'] = _brief_batch_error(result, 800)
        if success:
            save_distributor_history(distributor)
            order_no = result.get('order_no') if isinstance(result, dict) else ''
            if isinstance(result, dict) and result.get('pending_approval'):
                log(f"移库单已保存待审批：{order_no or '已保存'}", 'success')
            else:
                _apply_transfer_local_dealer(summary, transfer_type, distributor)
                log(f"移库完成：{order_no or '已完成'}", 'success')
        else:
            log(f"移库失败：{result}", 'error')
    except Exception as e:
        with transfer_job_lock:
            job = transfer_jobs.get(job_id)
            if job:
                job['running'] = False
                job['done'] = True
                job['success'] = False
                job['error'] = _brief_batch_error(e, 800)
                job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log(f"移库出错：{e}", 'error')

def _run_service_close_job(job_id, workers, orders):
    def log(message, level='dim'):
        _service_close_job_log(job_id, message, level)

    worker_entries = workers if isinstance(workers, list) else [workers]
    normalized_workers = []
    for entry in worker_entries:
        if isinstance(entry, (list, tuple)) and len(entry) >= 3:
            normalized_workers.append((entry[0], entry[1], entry[2]))
        else:
            slot_id = getattr(entry, "slot_id", "") or ""
            normalized_workers.append((entry, slot_id, _query_slot_label(slot_id) if slot_id else "查询通道"))

    started = time.time()
    log(f"开始批量结单：共 {len(orders or [])} 个服务单，使用 {len(normalized_workers)} 个查询通道", "info")
    try:
        total = len(orders or [])
        order_queue = queue.Queue()
        for index, row in enumerate(orders or [], 1):
            order_queue.put((index, row))
        rows_by_index = [None] * total
        worker_errors = []
        result_lock = threading.Lock()

        def mark_processed():
            with service_close_job_lock:
                job = service_close_jobs.get(job_id)
                if job:
                    job['current'] = min(int(job.get('total') or total), int(job.get('current') or 0) + 1)

        def run_worker(worker, slot_id, slot_label):
            while True:
                try:
                    index, row = order_queue.get_nowait()
                except queue.Empty:
                    return
                service_no = _clean_export_value(row.get("service_no") if isinstance(row, dict) else row)
                try:
                    display_label = _service_order_display(row if isinstance(row, dict) else {"service_no": service_no})
                    log(f"{slot_label} 处理服务单 {index}/{total}：{display_label}", "info")

                    def worker_log(message, level='dim'):
                        text = str(message or "")
                        if text.startswith("处理服务单 1/1"):
                            return
                        if text.startswith("CRM 服务单结单任务已加入队列"):
                            return
                        log(f"{slot_label} {text}", level)

                    ok, result = worker.close_service_orders([row], worker_log)
                    result = result if isinstance(result, dict) else {"error": str(result), "results": []}
                    result_rows = result.get("results") or []
                    result_row = result_rows[0] if result_rows else {
                        "service_no": service_no,
                        "success": False,
                        "status": "failed",
                        "message": result.get("error") or "服务单结单失败",
                    }
                    if isinstance(row, dict):
                        for key in ("barcodes", "customer_names", "product_names"):
                            if row.get(key) and not result_row.get(key):
                                result_row[key] = list(row.get(key) or [])
                        if not result_row.get("display_label"):
                            result_row["display_label"] = _service_order_display(result_row or row)
                    if not ok and result.get("error") and not result_row.get("message"):
                        result_row["message"] = result.get("error")
                    with result_lock:
                        rows_by_index[index - 1] = result_row
                except Exception as e:
                    error = _brief_batch_error(e, 800)
                    with result_lock:
                        rows_by_index[index - 1] = {
                            "service_no": service_no,
                            "success": False,
                            "status": "failed",
                            "message": error,
                        }
                        worker_errors.append(error)
                    log(f"{slot_label} {service_no} 结单出错：{error}", "error")
                finally:
                    mark_processed()
                    order_queue.task_done()

        threads = [
            threading.Thread(target=run_worker, args=(worker, slot_id, slot_label), daemon=True)
            for worker, slot_id, slot_label in normalized_workers
        ]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()

        rows = [row for row in rows_by_index if row]
        closed_count = 0
        already_closed_count = 0
        failed_count = 0
        order_map = {
            _clean_export_value(row.get("service_no")): row
            for row in (orders or [])
            if isinstance(row, dict)
        }
        for row in rows:
            service_no = _clean_export_value(row.get("service_no"))
            if not service_no:
                continue
            source_row = order_map.get(service_no) or {}
            row["barcodes"] = list(source_row.get("barcodes") or row.get("barcodes") or [])
            row["customer_names"] = list(source_row.get("customer_names") or row.get("customer_names") or [])
            row["product_names"] = list(source_row.get("product_names") or row.get("product_names") or [])
            row["display_label"] = _service_order_display(row)
            if row.get("success"):
                if row.get("status") == "already_closed":
                    already_closed_count += 1
                else:
                    closed_count += 1
                _record_service_closed_for_barcodes(service_no, (order_map.get(service_no) or {}).get("barcodes") or [])
            else:
                failed_count += 1

        general_error = "; ".join(worker_errors[:3])
        with service_close_job_lock:
            job = service_close_jobs.get(job_id)
            if not job:
                return
            job['running'] = False
            job['done'] = True
            job['success'] = bool(failed_count == 0 and not general_error)
            job['error'] = _brief_batch_error(general_error, 800) if general_error else ''
            job['results'] = rows
            job['current'] = total
            job['closed_count'] = closed_count
            job['already_closed_count'] = already_closed_count
            job['failed_count'] = failed_count
            job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elapsed = format_duration_seconds(time.time() - started)
        if failed_count:
            log(f"批量结单完成：新结单 {closed_count} 个，原本已结单 {already_closed_count} 个，失败 {failed_count} 个，总耗时 {elapsed}", "warn")
        else:
            log(f"批量结单完成：新结单 {closed_count} 个，原本已结单 {already_closed_count} 个，总耗时 {elapsed}", "success")
    except Exception as e:
        error = _brief_batch_error(e, 800)
        with service_close_job_lock:
            job = service_close_jobs.get(job_id)
            if job:
                job['running'] = False
                job['done'] = True
                job['success'] = False
                job['error'] = error
                job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log(f"批量结单出错：{error}", "error")

def _batch_stop_requested(job_id, idx):
    with batch_job_lock:
        job = batch_jobs.get(job_id)
        if not job or not job['stop_requested']:
            return False
        job['running'] = False
        job['stop_requested'] = False
        job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _batch_job_log(job_id, f"已停止，停在第 {idx} 个条码", 'warn')
    return True

def _run_batch_job(job_id, worker, barcodes, retry_limit=DEFAULT_BATCH_RETRY_LIMIT, excluded=None):
    retry_limit = _normalize_retry_limit(retry_limit)
    barcodes, filtered = filter_disassembly_barcodes(barcodes)
    excluded = list(excluded or []) + filtered
    retry_text = f"，失败最多重试 {retry_limit} 次" if retry_limit else ""
    if excluded:
        _batch_job_log(job_id, f"已排除拆机条码 {len(excluded)} 个，不查询：{', '.join(excluded[:10])}", 'warn')
    if not barcodes:
        with batch_job_lock:
            job = batch_jobs.get(job_id)
            if job:
                job['running'] = False
                job['stop_requested'] = False
                job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _batch_job_log(job_id, "输入的条码都是拆机条码，无需查询", 'warn')
        return
    _batch_job_log(job_id, f"开始批量查询 {len(barcodes)} 个条码{retry_text}...", 'info')
    if not worker.logged_in:
        if worker.remembered_logged_in:
            _batch_job_log(job_id, "正在验证上次 CRM 登录状态...", 'info')
            success, message = worker.check_login_status()
            if success and worker.logged_in:
                _batch_job_log(job_id, "上次 CRM 登录状态有效，继续查询", 'success')
            else:
                error = f"CRM 当前未登录，请先登录 CRM（{message or '会话未恢复'}）"
                with batch_job_lock:
                    job = batch_jobs.get(job_id)
                    if job:
                        job['failed'] = len(barcodes)
                        job['running'] = False
                        job['stop_requested'] = False
                        job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        job['results'].extend({
                            'barcode': barcode,
                            'success': False,
                            'attempts': 0,
                            'error': error,
                        } for barcode in barcodes)
                _batch_job_log(job_id, error, 'error')
                return
        else:
            error = "CRM 当前未登录，请先登录 CRM"
            with batch_job_lock:
                job = batch_jobs.get(job_id)
                if job:
                    job['failed'] = len(barcodes)
                    job['running'] = False
                    job['stop_requested'] = False
                    job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    job['results'].extend({
                        'barcode': barcode,
                        'success': False,
                        'attempts': 0,
                        'error': error,
                    } for barcode in barcodes)
            _batch_job_log(job_id, error, 'error')
            return
    for idx, barcode in enumerate(barcodes, start=1):
        with batch_job_lock:
            job = batch_jobs.get(job_id)
            if not job:
                return
            if job['stop_requested']:
                job['running'] = False
                job['stop_requested'] = False
                job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                stopped_at = idx
                should_stop = True
            else:
                job['current'] = idx
                should_stop = False
        if should_stop:
            _batch_job_log(job_id, f"已停止，停在第 {stopped_at} 个条码", 'warn')
            return

        success = False
        result = ""
        attempts = retry_limit + 1
        for attempt in range(1, attempts + 1):
            if _batch_stop_requested(job_id, idx):
                return
            if attempt == 1:
                _batch_job_log(job_id, f"正在查询第 {idx}/{len(barcodes)} 个：{barcode}", 'info')
            else:
                _batch_job_log(job_id, f"{barcode} 第 {attempt - 1}/{retry_limit} 次重试中...", 'warn')

            success, result = worker.query_barcode(barcode, lambda msg, level='dim': _batch_job_log(job_id, msg, level))
            if success:
                success, publish_error = publish_cluster_query_result(barcode, worker.slot_id)
                if not success:
                    result = publish_error
                break

            if attempt <= retry_limit:
                _batch_job_log(
                    job_id,
                    f"{barcode} 查询失败，将重试 {attempt}/{retry_limit}: {_brief_batch_error(result)}",
                    'warn'
                )
                time.sleep(2)

        with batch_job_lock:
            job = batch_jobs.get(job_id)
            if not job:
                return
            if success:
                update_barcode_query_slot(barcode, job.get('slot_id') or worker.slot_id)
                job['success'] += 1
                job['results'].append({
                    'barcode': barcode,
                    'success': True,
                    'attempts': attempt,
                    'view_url': f'/barcode/{barcode}.html',
                })
                _append_job_log_unlocked(
                    job,
                    f"✓ {barcode} 查询成功" + (f"（重试第 {attempt - 1} 次）" if attempt > 1 else ""),
                    'success',
                    BATCH_LOG_LIMIT
                )
            else:
                job['failed'] += 1
                job['results'].append({
                    'barcode': barcode,
                    'success': False,
                    'attempts': attempts,
                    'error': _brief_batch_error(result),
                })
                _append_job_log_unlocked(
                    job,
                    f"✗ {barcode} 查询失败（已重试 {retry_limit} 次）: {_brief_batch_error(result)}",
                    'error',
                    BATCH_LOG_LIMIT
                )

    with batch_job_lock:
        job = batch_jobs.get(job_id)
        if not job:
            return
        job['running'] = False
        job['stop_requested'] = False
        job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        success_count = job['success']
        failed_count = job['failed']
    _batch_job_log(
        job_id,
        f"批量查询完成，成功 {success_count} 个，失败 {failed_count} 个",
        'success'
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(
    __name__,
    template_folder=os.path.join(RESOURCE_BASE_DIR, "templates"),
    static_folder=os.path.join(RESOURCE_BASE_DIR, "static"),
    static_url_path="/static",
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "crm-barcode-query-local-secret")

DATA_BASE_DIR = _runtime_data_base_dir()
CONFIG_DIR = _runtime_config_dir()
BARCODE_DIR = os.path.join(DATA_BASE_DIR, "barcode")
ARCHIVE_DIR = os.path.join(BARCODE_DIR, "archived")
DATA_FILE = os.path.join(CONFIG_DIR, "barcode_data.json")
PRODUCT_LIBRARY_FILE = os.path.join(CONFIG_DIR, "product_library.json")
ACCOUNTS_FILE = os.path.join(CONFIG_DIR, "accounts.json")
DISTRIBUTOR_HISTORY_FILE = os.path.join(CONFIG_DIR, "distributor_history.json")
DISTRIBUTOR_HISTORY_DELETED_FILE = os.path.join(CONFIG_DIR, "distributor_history_deleted.json")
RESULTS_DIR = os.path.join(DATA_BASE_DIR, "results")
TEMP_QUERY_DIR = os.path.join(DATA_BASE_DIR, "temp_queries")
RUNTIME_CONFIG_FILE = _runtime_config_path()
CLUSTER_NODES_FILE = os.path.join(CONFIG_DIR, "cluster_nodes.json")
CRM_CREDENTIALS_FILE = os.path.join(CONFIG_DIR, "crm_credentials.json")
crm_credentials_lock = threading.Lock()
DEFAULT_OWN_DEALER_NAME = "江西省天麓工贸有限公司"
DEFAULT_FROZEN_WAREHOUSE_NAME = "江西天麓冻结仓库"
OWN_DEALER_NAME = DEFAULT_OWN_DEALER_NAME
FROZEN_WAREHOUSE_NAME = DEFAULT_FROZEN_WAREHOUSE_NAME

def _env_text(name, default=""):
    return str(os.environ.get(name) or default).strip()

def _node_identity():
    hostname = socket.gethostname()
    node_id = _env_text("CRM_NODE_ID", hostname)
    node_name = _env_text("CRM_NODE_NAME", node_id)
    return {
        "id": node_id,
        "name": node_name,
        "role": _env_text("CRM_NODE_ROLE", "standalone"),
        "cluster_id": _env_text("CRM_CLUSTER_ID", "default"),
        "hostname": hostname,
        "version": _env_text("CRM_APP_VERSION", _env_text("GITHUB_SHA", "")),
        "python": platform.python_version(),
        "platform": platform.platform(),
        "started_at": datetime.fromtimestamp(APP_STARTED_AT).strftime("%Y-%m-%d %H:%M:%S"),
        "uptime_seconds": int(time.time() - APP_STARTED_AT),
    }

def _check_directory_writable(path):
    try:
        os.makedirs(path, exist_ok=True)
        marker = os.path.join(path, ".healthcheck-write-test")
        with open(marker, "w", encoding="utf-8") as f:
            f.write(datetime.now().isoformat())
        os.remove(marker)
        return True, ""
    except Exception as e:
        return False, str(e)

def _readiness_checks():
    checks = {}
    config_ok = True
    config_error = ""
    try:
        load_crm_config()
    except Exception as e:
        config_ok = False
        config_error = str(e)
    checks["config"] = {"ok": config_ok, "message": config_error}

    for key, path in {
        "data_dir": DATA_BASE_DIR,
        "config_dir": CONFIG_DIR,
        "barcode_dir": BARCODE_DIR,
        "results_dir": RESULTS_DIR,
        "session_base": _crm_session_base_dir(),
    }.items():
        ok, message = _check_directory_writable(path)
        checks[key] = {"ok": ok, "path": path, "message": message}

    pool_ok = bool(getattr(crm_pool, "query_slots", None) and getattr(crm_pool, "transfer_slots", None))
    checks["worker_pool"] = {
        "ok": pool_ok,
        "query_workers": len(getattr(crm_pool, "query_slots", []) or []),
        "transfer_workers": len(getattr(crm_pool, "transfer_slots", []) or []),
    }
    return checks

def _node_status_payload(include_checks=False):
    slots = crm_pool.slots_payload()
    query_slots = slots.get("query") or []
    transfer_slots = slots.get("transfer") or []
    payload = {
        "success": True,
        "node": _node_identity(),
        "storage": {
            "data_dir": DATA_BASE_DIR,
            "config_dir": CONFIG_DIR,
            "barcode_dir": BARCODE_DIR,
            "results_dir": RESULTS_DIR,
            "session_base": _crm_session_base_dir(),
            "database_configured": bool(_env_text("DATABASE_URL")),
            "r2_configured": bool(_env_text("R2_BUCKET") and _env_text("R2_ENDPOINT_URL")),
        },
        "crm": {
            "query_total": len(query_slots),
            "query_logged_in": sum(1 for row in query_slots if row.get("logged_in")),
            "query_busy": sum(1 for row in query_slots if row.get("busy")),
            "transfer_total": len(transfer_slots),
            "transfer_logged_in": sum(1 for row in transfer_slots if row.get("logged_in")),
            "transfer_busy": sum(1 for row in transfer_slots if row.get("busy")),
            "slots": slots,
        },
    }
    if include_checks:
        checks = _readiness_checks()
        payload["checks"] = checks
        payload["ready"] = all((row or {}).get("ok") for row in checks.values())
    return payload

def _directory_has_files(path):
    if not os.path.isdir(path):
        return False
    for _root, _dirs, files in os.walk(path):
        if files:
            return True
    return False

def _copy_missing_tree(source, target):
    copied = 0
    for root, _dirs, files in os.walk(source):
        rel_root = os.path.relpath(root, source)
        target_root = target if rel_root == "." else os.path.join(target, rel_root)
        os.makedirs(target_root, exist_ok=True)
        for filename in files:
            src = os.path.join(root, filename)
            dst = os.path.join(target_root, filename)
            if os.path.exists(dst):
                continue
            shutil.copy2(src, dst)
            copied += 1
    return copied

def _migrate_legacy_runtime_data():
    if os.environ.get("CRM_DISABLE_DATA_MIGRATION") in ("1", "true", "yes"):
        return
    legacy_pairs = [
        (
            [
                os.environ.get("CRM_LEGACY_BARCODE_DIR", "/app/legacy/barcode"),
                "/app/barcode",
                os.path.join(RESOURCE_BASE_DIR, "barcode"),
            ],
            BARCODE_DIR,
        ),
        (
            [
                os.environ.get("CRM_LEGACY_RESULTS_DIR", "/app/legacy/results"),
                "/app/results",
                os.path.join(RESOURCE_BASE_DIR, "results"),
            ],
            RESULTS_DIR,
        ),
    ]
    for sources, target in legacy_pairs:
        for source in sources:
            if not source or os.path.abspath(source) == os.path.abspath(target):
                continue
            if not os.path.isdir(source) or not _directory_has_files(source):
                continue
            try:
                copied = _copy_missing_tree(source, target)
                if copied:
                    print(f"  [DATA] 已补充旧数据目录: {source} -> {target}，新增 {copied} 个文件")
            except Exception as e:
                print(f"  [DATA] 迁移旧数据目录失败: {source} -> {target}: {e}")

_migrate_legacy_runtime_data()

def _migrate_config_files_from_barcode_dir():
    os.makedirs(CONFIG_DIR, exist_ok=True)
    for filename in (
        "runtime_config.json",
        "crm_slot_state.json",
        "crm_credentials.json",
    ):
        _migrate_root_config_file(filename)
    for filename in (
        "barcode_data.json",
        "product_library.json",
        "accounts.json",
        "distributor_history.json",
        "distributor_history_deleted.json",
    ):
        source = os.path.join(BARCODE_DIR, filename)
        target = os.path.join(CONFIG_DIR, filename)
        if not os.path.exists(source):
            continue
        if os.path.exists(target):
            try:
                os.remove(source)
                print(f"  [DATA] 已清理旧配置文件: {source}")
            except Exception as e:
                print(f"  [DATA] 旧配置文件清理失败: {source}: {e}")
            continue
        try:
            shutil.copy2(source, target)
            os.remove(source)
            print(f"  [DATA] 已迁移配置文件: {source} -> {target}")
        except Exception as e:
            print(f"  [DATA] 配置文件迁移失败: {source} -> {target}: {e}")

_migrate_config_files_from_barcode_dir()

def _runtime_text_value(value, default):
    value = str(value or "").replace("\xa0", " ").strip()
    return value or default

def _runtime_bool_value(value, default=True):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        text = value.strip().lower()
        if text in ("1", "true", "yes", "on", "是", "开启"):
            return True
        if text in ("0", "false", "no", "off", "否", "关闭"):
            return False
    return default

def load_runtime_config():
    defaults = {
        "query_workers": _normalize_worker_count(os.environ.get("CRM_QUERY_WORKERS"), 2),
        "transfer_workers": _normalize_worker_count(os.environ.get("CRM_TRANSFER_WORKERS"), 2),
        "own_dealer_name": _runtime_text_value(os.environ.get("CRM_OWN_DEALER_NAME"), DEFAULT_OWN_DEALER_NAME),
        "frozen_warehouse_name": _runtime_text_value(os.environ.get("CRM_FROZEN_WAREHOUSE_NAME"), DEFAULT_FROZEN_WAREHOUSE_NAME),
        "frozen_warehouse_save_only": _runtime_bool_value(os.environ.get("CRM_FROZEN_WAREHOUSE_SAVE_ONLY"), True),
    }
    services = _get_cluster_services()
    if services:
        data = services.catalog.get_runtime_config(
            f"node:{SERVER_CLUSTER_CONFIG.node_id}"
        ) or services.catalog.get_runtime_config("global") or {}
        defaults["query_workers"] = _normalize_worker_count(data.get("query_workers"), defaults["query_workers"])
        defaults["transfer_workers"] = _normalize_worker_count(data.get("transfer_workers"), defaults["transfer_workers"])
        defaults["own_dealer_name"] = _runtime_text_value(data.get("own_dealer_name"), defaults["own_dealer_name"])
        defaults["frozen_warehouse_name"] = _runtime_text_value(data.get("frozen_warehouse_name"), defaults["frozen_warehouse_name"])
        defaults["frozen_warehouse_save_only"] = _runtime_bool_value(data.get("frozen_warehouse_save_only"), defaults["frozen_warehouse_save_only"])
        return defaults
    if os.path.exists(RUNTIME_CONFIG_FILE):
        try:
            with open(RUNTIME_CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                defaults["query_workers"] = _normalize_worker_count(data.get("query_workers"), defaults["query_workers"])
                defaults["transfer_workers"] = _normalize_worker_count(data.get("transfer_workers"), defaults["transfer_workers"])
                defaults["own_dealer_name"] = _runtime_text_value(data.get("own_dealer_name"), defaults["own_dealer_name"])
                defaults["frozen_warehouse_name"] = _runtime_text_value(data.get("frozen_warehouse_name"), defaults["frozen_warehouse_name"])
                defaults["frozen_warehouse_save_only"] = _runtime_bool_value(data.get("frozen_warehouse_save_only"), defaults["frozen_warehouse_save_only"])
        except Exception:
            pass
    return defaults

def save_runtime_config(config):
    current = load_runtime_config()
    payload = {
        "query_workers": _normalize_worker_count(config.get("query_workers"), current.get("query_workers", 2)),
        "transfer_workers": _normalize_worker_count(config.get("transfer_workers"), current.get("transfer_workers", 2)),
        "own_dealer_name": _runtime_text_value(config.get("own_dealer_name"), current.get("own_dealer_name", DEFAULT_OWN_DEALER_NAME)),
        "frozen_warehouse_name": _runtime_text_value(config.get("frozen_warehouse_name"), current.get("frozen_warehouse_name", DEFAULT_FROZEN_WAREHOUSE_NAME)),
        "frozen_warehouse_save_only": _runtime_bool_value(config.get("frozen_warehouse_save_only"), current.get("frozen_warehouse_save_only", True)),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    services = _get_cluster_services()
    if services:
        services.catalog.set_runtime_config(
            f"node:{SERVER_CLUSTER_CONFIG.node_id}",
            payload,
        )
        return payload
    os.makedirs(os.path.dirname(RUNTIME_CONFIG_FILE), exist_ok=True)
    with open(RUNTIME_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return payload

def business_config():
    return load_runtime_config()

def _cluster_admin_token():
    return str(os.environ.get("CRM_CLUSTER_ADMIN_TOKEN") or "").strip()

def _cluster_admin_authorized():
    token = _cluster_admin_token()
    if not token:
        return False
    provided = (request.headers.get("X-CRM-Cluster-Token") or "").strip()
    auth_header = (request.headers.get("Authorization") or "").strip()
    if not provided and auth_header.lower().startswith("bearer "):
        provided = auth_header[7:].strip()
    return bool(provided and hmac.compare_digest(provided, token))

def _normalize_node_url(url):
    text = str(url or "").strip()
    return text.rstrip("/")

def _public_node_url():
    explicit = _normalize_node_url(os.environ.get("CRM_PUBLIC_URL"))
    if explicit:
        return explicit
    if has_request_context():
        return request.host_url.rstrip("/")
    return ""

def _sanitize_cluster_node(row):
    if not isinstance(row, dict):
        return None
    node_id = str(row.get("id") or "").strip()
    url = _normalize_node_url(row.get("url"))
    if not node_id or not url:
        return None
    return {
        "id": node_id,
        "name": str(row.get("name") or node_id).strip(),
        "url": url,
        "role": str(row.get("role") or "").strip(),
    }

def _env_cluster_nodes():
    raw = os.environ.get("CRM_CLUSTER_NODES") or ""
    if raw.strip():
        try:
            data = json.loads(raw)
            nodes = [_sanitize_cluster_node(row) for row in (data if isinstance(data, list) else [])]
            return [row for row in nodes if row]
        except Exception:
            return []
    identity = _node_identity()
    url = _public_node_url()
    if not url:
        return []
    return [{
        "id": identity["id"],
        "name": identity["name"],
        "role": identity["role"],
        "url": url,
    }]

def load_cluster_nodes():
    if os.path.exists(CLUSTER_NODES_FILE):
        try:
            with open(CLUSTER_NODES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            nodes = [_sanitize_cluster_node(row) for row in (data if isinstance(data, list) else [])]
            nodes = [row for row in nodes if row]
            if nodes:
                return nodes
        except Exception:
            pass
    return _env_cluster_nodes()

def _cluster_node_status(node):
    local_id = _node_identity().get("id")
    if node.get("id") == local_id:
        payload = _node_status_payload(include_checks=False)
        return {**node, "online": True, "status": payload}
    try:
        req = urlrequest.Request(node["url"] + "/api/node/status", headers={"Accept": "application/json"})
        with urlrequest.urlopen(req, timeout=5) as resp:
            payload = json.loads(resp.read().decode("utf-8", errors="replace"))
        return {**node, "online": True, "status": payload}
    except Exception as e:
        return {**node, "online": False, "error": str(e)}

def _save_local_runtime_config_from_payload(data):
    current = load_runtime_config()
    config = save_runtime_config({
        'query_workers': data.get('query_workers', current.get('query_workers', 2)),
        'transfer_workers': data.get('transfer_workers', current.get('transfer_workers', 2)),
        'own_dealer_name': data.get('own_dealer_name', current.get('own_dealer_name', DEFAULT_OWN_DEALER_NAME)),
        'frozen_warehouse_name': data.get('frozen_warehouse_name', current.get('frozen_warehouse_name', DEFAULT_FROZEN_WAREHOUSE_NAME)),
        'frozen_warehouse_save_only': data.get('frozen_warehouse_save_only', current.get('frozen_warehouse_save_only', True)),
    })
    slots = crm_pool.resize(config['query_workers'], config['transfer_workers'])
    return config, slots

def _post_node_runtime_config(node, data):
    token = _cluster_admin_token()
    if not token:
        return False, "未配置 CRM_CLUSTER_ADMIN_TOKEN，不能远程修改节点", None
    payload = json.dumps({
        "query_workers": data.get("query_workers"),
        "transfer_workers": data.get("transfer_workers"),
    }).encode("utf-8")
    req = urlrequest.Request(
        node["url"] + "/api/runtime-config",
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-CRM-Cluster-Token": token,
        },
    )
    try:
        with urlrequest.urlopen(req, timeout=12) as resp:
            result = json.loads(resp.read().decode("utf-8", errors="replace"))
    except urlerror.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return False, body or str(e), None
    except Exception as e:
        return False, str(e), None
    if not result.get("success"):
        return False, result.get("error") or "节点保存失败", result
    return True, "", result

def _current_cluster_node():
    local_id = _node_identity().get("id")
    for node in load_cluster_nodes():
        if node.get("id") == local_id:
            return node
    return None

def own_dealer_name():
    return business_config().get("own_dealer_name") or DEFAULT_OWN_DEALER_NAME

def frozen_warehouse_name():
    return business_config().get("frozen_warehouse_name") or DEFAULT_FROZEN_WAREHOUSE_NAME

def frozen_warehouse_save_only():
    return bool(business_config().get("frozen_warehouse_save_only", True))

FIELD_IDS = {
    'newisclosed1': '结单状态',
    'SHIPSTATUS1': '装箱单状态',
    'instlled1': '是否安装',
    'newstatus1': '产品状态',
    'typestr1': '服务类型',
    'statustr1': '服务单状态',
    'newname1': '条码号',
    'zxd1': '装箱单号',
    'shipdate1': '发货日期',
    'newerpshipno1': '发货单号',
    'ProductNumber1': '物料编码',
    'newproductidName1': '机型',
    'newordsalesorderidName1': '订单号',
    'servno1': '服务单号',
    'name1': '客户',
    'newaddress1': '地址',
    'newtelephone1': '电话',
    'newstationidName1': '服务站',
    'newdealername1': '服务经销商',
    'dealername1': '所属经销商',
    'buno1': '调拨单号',
    'transstockdate1': '移库日期',
    'newproductname1': '物料名称',
    'newaccountidName1': '装箱单经销商',
    'newtype1': '装箱类型',
    'newdeposit1': '押金',
    'returned1': '返修',
    'newpresaledealername1': '售前经销商',
    'underlinestr1': '线下带货上门',
    'productname1': '物料描述',
    'productdealernumber1': '所属经销商代码',
    'depositdealer1': '押金经销商',
    'depositdealernumber1': '押金经销商代码',
    'isonline1': '线上产品',
    'transdate1': '调拨日期',
    'serno1': '条码',
    'productcode1': '物料编码',
    'dealercustcode1': '经销商代码',
    'dealercustname1': '经销商名称',
    'distributorcustcode1': '分销商代码',
    'distributorcustname1': '分销商名称',
    'type1': '移库类型',
    'state1': '状态',
    'statuscode1': '是否可用',
    'installDate1': '安装日期',
    'myproductdealer1': '所属经销商',
    'customer1': '客户',
    'newphone1': '电话',
    'newaddress1': '地址',
    'installdate1': '安装日期',
    'warrantyid1': '保卡条码',
    'scandate1': '扫描时间',
    'returndepositid1': '押金单号',
    'returndepositamount1': '押金金额',
    'servdate1': '服务单日期',
    'returndealer1': '押金经销商',
    'returndate1': '返还日期',
    'returnstatus1': '状态',
    'adjustid1': '库存调整单号',
    'adjustdate1': '调整日期',
    'adjustdirection1': '交易方向',
    'adjuststatus1': '状态',
    'adjustdealer1': '经销商',
    'moveid1': '移机单号',
    'movedate1': '移机日期',
    'oldcontact1': '原联系人',
    'oldtelephone1': '原联系电话',
    'oldaddress1': '原联系地址',
    'newdealer1': '关联经销商',
    'olddealer1': '原经销商',
}

FILTER_FIELDS = {
    'myproductdealer1_sr5': '归属经销商',
    'newdealername1_sr2': '服务经销商',
    'newisclosed1_sr2': '是否结单',
}

SUBREPORT_NAMES = {
    1: '装箱单',
    2: '服务单',
    3: '保卡扫描',
    4: '押金返还',
    5: '产品档案',
    6: '库存调整',
    7: '调拨单',
    8: '移库单',
    9: '移机单',
    10: '库存状态',
}

SUBREPORT_FIELD_MAP = {
    1: ['SHIPSTATUS1', 'zxd1', 'shipdate1', 'newerpshipno1', 'ProductNumber1',
        'newproductidName1', 'newordsalesorderidName1', 'newname1', 'newproductname1',
        'newtype1', 'newdeposit1', 'newaccountidName1', 'newpayaccountidName1'],
    2: ['servno1', 'underlinestr1', 'statustr1', 'newproductname1', 'newisclosed1',
        'typestr1', 'newproductidName1', 'newaddress1', 'newtelephone1', 'name1',
        'newstationidName1', 'newdealername1', 'newpresaledealername1'],
    3: ['warrantyid1', 'scandate1'],
    4: ['returndepositid1', 'returndepositamount1', 'servno1', 'servdate1',
        'newdealername1', 'returndealer1', 'returndate1', 'returnstatus1'],
    5: ['newname1', 'newproductname1', 'productname1', 'instlled1', 'returned1',
        'newdeposit1', 'isonline1', 'myproductdealer1', 'productdealernumber1',
        'depositdealernumber1', 'depositdealer1', 'customer1', 'newphone1',
        'newaddress1', 'installdate1'],
    6: ['adjustid1', 'adjustdate1', 'serno1', 'productname1',
        'adjustdirection1', 'adjuststatus1', 'adjustdealer1'],
    7: ['buno1', 'transdate1', 'serno1', 'productcode1', 'productname1',
        'accountincustcustname1', 'accountoutcustcode1', 'accountoutcustname1', 'transstatus1'],
    8: ['buno1', 'transstockdate1', 'serno1', 'productcode1', 'productname1',
        'dealercustcode1', 'dealercustname1', 'distributorcustcode1',
        'distributorcustname1', 'type1', 'state1'],
    9: ['moveid1', 'movedate1', 'serno1', 'productcode1', 'productname1',
        'oldcontact1', 'oldtelephone1', 'oldaddress1', 'newcontact1',
        'newtelephone1', 'newaddress1', 'newdealer1', 'olddealer1'],
    10: ['newname1', 'newproductname1', 'newproductidName1', 'newstatus1',
         'statuscode1', 'dealername1', 'newdeposit1'],
}

def extract_fields_from_html(filepath):
    result = {}
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()

        sr_pattern = re.compile(r'<div id="Subreport(\d+)"')
        sr_matches = list(sr_pattern.finditer(html))

        field_pattern_cache = {}
        for field_id in FIELD_IDS:
            field_pattern_cache[field_id] = rf'id="{field_id}"[^>]*>.*?<span[^>]*>([^<]+)</span>'

        for i, m in enumerate(sr_matches):
            sr_num = int(m.group(1))
            if sr_num not in SUBREPORT_FIELD_MAP:
                continue

            sr_start = m.start()
            sr_end = sr_matches[i + 1].start() if i + 1 < len(sr_matches) else len(html)
            sr_body = html[sr_start:sr_end]

            sr_key = f'sr{sr_num}'
            raw_field_values = {}

            for fid in SUBREPORT_FIELD_MAP[sr_num]:
                if fid not in field_pattern_cache:
                    continue
                pattern = field_pattern_cache[fid]
                matches = re.findall(pattern, sr_body, re.DOTALL)
                vals = []
                for match in matches:
                    value = html_mod.unescape(match).strip()
                    if value and value not in ['¥.00', '.00', '', ' ', '***']:
                        vals.append(value)
                if vals:
                    raw_field_values[fid] = vals

            if not raw_field_values:
                continue

            record_count = max(len(v) for v in raw_field_values.values())

            if record_count == 1:
                result[sr_key] = {fid: raw_field_values[fid][0] for fid in raw_field_values}
            else:
                result[sr_key] = [
                    {fid: raw_field_values.get(fid, [''])[j] if j < len(raw_field_values.get(fid, [''])) else '' for fid in raw_field_values}
                    for j in range(record_count)
                ]

    except Exception:
        pass

    return result

def _get_field(fields, field_id):
    import re as re_mod
    m = re_mod.search(r'_sr(\d+)$', field_id)
    if not m:
        return fields.get(field_id, '')
    sr_num = m.group(1)
    sr_key = f'sr{sr_num}'
    real_fid = field_id.rsplit('_sr', 1)[0]
    sub = fields.get(sr_key, {})
    if isinstance(sub, list) and sub:
        sub = sub[0]
    if isinstance(sub, dict):
        return sub.get(real_fid, '')
    return ''

def _clean_export_value(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return html_mod.unescape(str(value)).replace('\xa0', ' ').strip()

def is_disassembly_barcode(barcode):
    barcode = _clean_export_value(barcode)
    return bool(barcode) and barcode[0].upper() in ("A", "C")

def filter_disassembly_barcodes(barcodes):
    kept = []
    excluded = []
    for barcode in barcodes or []:
        barcode = _clean_export_value(barcode)
        if not barcode:
            continue
        if is_disassembly_barcode(barcode):
            excluded.append(barcode)
        else:
            kept.append(barcode)
    return kept, excluded

def normalize_input_barcodes(barcodes):
    normalized = OrderedDict()
    for barcode in barcodes or []:
        barcode = _clean_export_value(barcode)
        if barcode:
            normalized[barcode] = True
    return list(normalized.keys())

def _records_for_export(fields, sr_key):
    raw = fields.get(sr_key)
    if isinstance(raw, list):
        records = raw
    elif isinstance(raw, dict):
        records = [raw]
    else:
        return []

    result = []
    seen = set()
    for record in records:
        if not isinstance(record, dict):
            continue
        cleaned = OrderedDict()
        for key, value in record.items():
            cell_value = _clean_export_value(value)
            if cell_value:
                cleaned[key] = cell_value
        if not cleaned:
            continue
        signature = json.dumps(cleaned, ensure_ascii=False, sort_keys=True)
        if signature in seen:
            continue
        seen.add(signature)
        result.append(cleaned)
    return result

def _prepare_export_rows(selected_barcodes):
    export_rows = []
    for item in selected_barcodes:
        fields = item.get('fields') or {}
        records_by_sr = {}
        for sr_num in sorted(SUBREPORT_FIELD_MAP):
            sr_key = f'sr{sr_num}'
            records = _records_for_export(fields, sr_key)
            if records:
                records_by_sr[sr_key] = records

        barcode = _clean_export_value(item.get('barcode')) or _get_field(fields, 'newname1_sr1')
        export_rows.append({
            'barcode': barcode,
            'remark': _clean_export_value(item.get('remark')),
            'time': _clean_export_value(item.get('time') or item.get('archiveTime')),
            'records_by_sr': records_by_sr,
        })
    return export_rows

def _build_export_columns(export_rows):
    columns = [
        {'group': '基础信息', 'label': '条码', 'type': 'base', 'key': 'barcode'},
        {'group': '基础信息', 'label': '备注', 'type': 'base', 'key': 'remark'},
        {'group': '基础信息', 'label': '查询时间', 'type': 'base', 'key': 'time'},
    ]

    for sr_num in sorted(SUBREPORT_FIELD_MAP):
        sr_key = f'sr{sr_num}'
        max_count = max((len(row['records_by_sr'].get(sr_key, [])) for row in export_rows), default=0)
        if max_count == 0:
            continue

        present_fields = OrderedDict()
        for row in export_rows:
            for record in row['records_by_sr'].get(sr_key, []):
                for field_id in record:
                    present_fields[field_id] = True

        preferred = [field_id for field_id in SUBREPORT_FIELD_MAP.get(sr_num, []) if field_id in present_fields]
        extras = [field_id for field_id in present_fields if field_id not in preferred]
        field_ids = preferred + extras
        group_name = SUBREPORT_NAMES.get(sr_num, f'子报表{sr_num}')

        for record_index in range(max_count):
            group_label = f'{group_name} {record_index + 1}' if max_count > 1 else group_name
            for field_id in field_ids:
                columns.append({
                    'group': group_label,
                    'label': FIELD_IDS.get(field_id, field_id),
                    'type': 'field',
                    'sr_key': sr_key,
                    'record_index': record_index,
                    'field_id': field_id,
                })
    return columns

def _export_cell_value(row, column):
    if column['type'] == 'base':
        return row.get(column['key'], '')
    records = row['records_by_sr'].get(column['sr_key'], [])
    if column['record_index'] >= len(records):
        return ''
    return records[column['record_index']].get(column['field_id'], '')

def _display_width(value):
    text = str(value or '')
    return sum(2 if ord(ch) > 127 else 1 for ch in text)

def _suggest_column_width(label, values):
    max_width = _display_width(label)
    for value in values:
        max_width = max(max_width, _display_width(value))
    if '地址' in label:
        return min(max(max_width + 2, 24), 42)
    if any(word in label for word in ['名称', '经销商', '客户', '机型', '物料描述', '备注']):
        return min(max(max_width + 2, 18), 34)
    if any(word in label for word in ['日期', '时间']):
        return min(max(max_width + 2, 14), 20)
    return min(max(max_width + 2, 12), 28)

def _service_closed_filter_value(fields):
    sub = (fields or {}).get('sr2', {})
    records = sub if isinstance(sub, list) else ([sub] if isinstance(sub, dict) else [])
    return '已结单' if any((row or {}).get('newisclosed1') == '已结单' for row in records) else '未结单'

def _get_filter_value(item, field_id):
    fields = item.get('fields') or {}
    if field_id == 'newisclosed1_sr2':
        return _service_closed_filter_value(fields)
    m = re.search(r'_sr(\d+)$', field_id)
    if m:
        sr_key = f"sr{m.group(1)}"
        real_fid = field_id.rsplit('_sr', 1)[0]
        sub = fields.get(sr_key, {})
        if isinstance(sub, list):
            values = []
            for row in sub:
                value = _clean_export_value((row or {}).get(real_fid, ''))
                if value and value not in values:
                    values.append(value)
            return values
    return _get_field(fields, field_id)

def get_filter_options(barcodes):
    options = {}
    for field_id, label in FILTER_FIELDS.items():
        values = set()
        for b in barcodes:
            val = _get_filter_value(b, field_id)
            vals = val if isinstance(val, list) else [val]
            for item in vals:
                if item:
                    values.add(item)
        options[field_id] = {
            'label': label,
            'field_id': field_id,
            'options': sorted(values)
        }
    return options

def _first_record(fields, sr_key):
    sub = fields.get(sr_key, {})
    if isinstance(sub, list):
        return sub[0] if sub else {}
    if isinstance(sub, dict):
        return sub
    return {}

def product_prefix_from_barcode(barcode):
    barcode = _clean_export_value(barcode)
    if len(barcode) > 10:
        return barcode[:-10]
    return barcode[:2]

def load_product_library():
    services = _get_cluster_services()
    if services:
        return {
            row["prefix"]: row
            for row in services.catalog.list_product_rules()
        }
    if os.path.exists(PRODUCT_LIBRARY_FILE):
        try:
            with open(PRODUCT_LIBRARY_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            pass
    return {}

def save_product_library(data):
    services = _get_cluster_services()
    if services:
        services.catalog.replace_product_rules(list(data.values()))
        return
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(PRODUCT_LIBRARY_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def upsert_product_library(prefix, product_code, product_name, source_barcode=""):
    prefix = _clean_export_value(prefix)
    product_code = _clean_export_value(product_code)
    product_name = _clean_export_value(product_name)
    if not prefix or not product_code or not product_name:
        return False
    data = load_product_library()
    data[prefix] = {
        'prefix': prefix,
        'product_code': product_code,
        'product_name': product_name,
        'source_barcode': _clean_export_value(source_barcode),
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    save_product_library(data)
    return True

def update_product_library_from_info(info):
    barcode = _clean_export_value(info.get('barcode'))
    prefix = product_prefix_from_barcode(barcode)
    return upsert_product_library(prefix, info.get('product_code'), info.get('product_name'), barcode)

def match_product_library(barcode):
    barcode = _clean_export_value(barcode)
    prefix = product_prefix_from_barcode(barcode)
    data = load_product_library()
    row = data.get(prefix) or {}
    if row:
        return {
            'prefix': prefix,
            'product_code': _clean_export_value(row.get('product_code')),
            'product_name': _clean_export_value(row.get('product_name')),
        }
    return None

def lookup_product_by_barcode(barcode):
    barcode = _clean_export_value(barcode)
    if not barcode:
        return None
    for item in scan_barcodes():
        if item.get('barcode') == barcode:
            info = _barcode_product_info(item)
            if info.get('product_code') and info.get('product_name'):
                update_product_library_from_info(info)
                info['matched_prefix'] = product_prefix_from_barcode(barcode)
                return info
            break
    return _barcode_product_info_from_library(barcode)

def _barcode_product_info(item):
    fields = item.get('fields') or {}
    sr10 = _first_record(fields, 'sr10')
    sr5 = _first_record(fields, 'sr5')
    sr2 = _first_record(fields, 'sr2')
    sr1 = _first_record(fields, 'sr1')

    product_code = (
        _clean_export_value(sr10.get('newproductname1')) or
        _clean_export_value(sr5.get('newproductname1')) or
        _clean_export_value(sr1.get('ProductNumber1'))
    )
    product_name = (
        _clean_export_value(sr10.get('newproductidName1')) or
        _clean_export_value(sr5.get('productname1')) or
        _clean_export_value(sr1.get('newproductidName1')) or
        _clean_export_value(sr1.get('newproductname1'))
    )
    product_model = (
        _clean_export_value(sr10.get('newproductidName1')) or
        _clean_export_value(sr1.get('newproductidName1')) or
        product_name
    )
    current_dealer = (
        _clean_export_value(item.get('currentDealerOverride')) or
        _clean_export_value(sr10.get('dealername1')) or
        _clean_export_value(sr5.get('myproductdealer1')) or
        _clean_export_value(sr1.get('newaccountidName1'))
    )
    installed = _clean_export_value(sr5.get('instlled1'))
    product_status = _clean_export_value(sr10.get('newstatus1'))
    service_dealer = _clean_export_value(sr2.get('newdealername1'))

    info = {
        'barcode': _clean_export_value(item.get('barcode')),
        'product_code': product_code,
        'product_name': product_name,
        'product_model': product_model,
        'current_dealer': current_dealer,
        'service_dealer': service_dealer,
        'installed': installed,
        'product_status': product_status,
        'source': 'query_result',
        'matched_prefix': product_prefix_from_barcode(item.get('barcode')),
    }
    return info


def publish_cluster_query_result(barcode, slot_id):
    services = _get_cluster_services()
    if not services:
        return True, ""
    html_path = os.path.join(BARCODE_DIR, f"{barcode}.html")
    if not os.path.exists(html_path):
        return False, "查询结果文件不存在，无法发布到共享存储"
    try:
        fields = extract_fields_from_html(html_path)
        info = _barcode_product_info({"barcode": barcode, "fields": fields})
        existing = services.catalog.get_barcode(barcode) or {}
        metadata = dict(existing.get("metadata") or {})
        query_time = datetime.now().astimezone()
        metadata.update({
            "querySlotId": slot_id,
            "querySlotLabel": _query_slot_label(slot_id),
            "queryUpdatedAt": query_time.isoformat(),
        })
        services.publish_barcode_result(
            barcode,
            html_path,
            fields,
            {
                "product_name": info.get("product_name") or "",
                "product_code": info.get("product_code") or "",
                "current_dealer": info.get("current_dealer") or "",
                "service_dealer": info.get("service_dealer") or "",
                "query_node_id": SERVER_CLUSTER_CONFIG.node_id,
                "query_slot_id": slot_id,
                "query_updated_at": query_time,
                "metadata": metadata,
                "remark": existing.get("remark") or "",
                "archived": bool(existing.get("archived")),
                "archive_time": existing.get("archive_time"),
                "current_dealer_override": existing.get("current_dealer_override") or "",
                "transfer_updated_at": existing.get("transfer_updated_at"),
                "service_closed": existing.get("service_closed"),
                "latest_service_order": existing.get("latest_service_order") or "",
            },
        )
        return True, ""
    except Exception as error:
        return False, f"共享查询结果发布失败：{error}"

def refresh_product_library_from_query_fields(barcode, fields, log=None):
    info = _barcode_product_info({
        'barcode': barcode,
        'fields': fields,
    })
    prefix = product_prefix_from_barcode(barcode)
    if update_product_library_from_info(info):
        if log:
            log(
                f"条码匹配已按本次查询刷新：前缀 {prefix}，{info.get('product_code')} / {info.get('product_name')}",
                "success"
            )
        return True
    if log:
        log(f"条码匹配未刷新：条码 {barcode} 缺少产品编码或产品名称", "warn")
    return False

def _barcode_product_info_from_library(barcode):
    matched = match_product_library(barcode)
    if not matched:
        return {
            'barcode': _clean_export_value(barcode),
            'product_code': '',
            'product_name': '',
            'product_model': '',
            'current_dealer': '',
            'service_dealer': '',
            'installed': '',
            'product_status': '',
            'source': 'unmatched',
            'matched_prefix': '',
        }
    return {
        'barcode': _clean_export_value(barcode),
        'product_code': matched['product_code'],
        'product_name': matched['product_name'],
        'product_model': matched['product_name'],
        'current_dealer': '',
        'service_dealer': '',
        'installed': '',
        'product_status': '',
        'source': 'product_library',
        'matched_prefix': matched['prefix'],
    }

def _missing_product_library_representatives(selected_barcodes):
    wanted = OrderedDict()
    for barcode in selected_barcodes:
        barcode = _clean_export_value(barcode)
        if is_disassembly_barcode(barcode):
            continue
        if barcode and barcode not in wanted:
            wanted[barcode] = True

    all_items = {item['barcode']: item for item in scan_barcodes()}
    groups = OrderedDict()
    for barcode in wanted:
        if match_product_library(barcode):
            continue

        item = all_items.get(barcode)
        if item:
            info = _barcode_product_info(item)
            if info.get('product_code') and info.get('product_name'):
                update_product_library_from_info(info)
                continue

        prefix = product_prefix_from_barcode(barcode)
        if prefix and prefix not in groups:
            groups[prefix] = barcode
    return groups

def ensure_product_library_for_barcodes(selected_barcodes, log=None, worker=None):
    representatives = _missing_product_library_representatives(selected_barcodes)
    result = {'queried': [], 'failed': []}
    if not representatives:
        return result

    def emit(message, level='info'):
        if log:
            log(message, level)

    total = len(representatives)
    emit(f"发现 {total} 个产品前缀未维护，先各查询 1 个代表条码补充条码匹配")
    for index, (prefix, barcode) in enumerate(representatives.items(), 1):
        emit(f"自动补充 {index}/{total}：前缀 {prefix}，代表条码 {barcode}", "info")
        emit(f"准备查询代表条码 {barcode}，用于补充前缀 {prefix}", "dim")
        existing_paths = existing_barcode_result_paths(barcode)
        had_metadata = barcode_metadata_exists(barcode)
        success, message = (worker or crm_pool.get(kind="query")).query_barcode(barcode, log, TEMP_QUERY_DIR)
        if success:
            delete_temporary_query_result(
                barcode,
                emit,
                keep_paths=existing_paths,
                keep_metadata=had_metadata,
            )
        emit(f"代表条码 {barcode} 查询返回：{'成功' if success else '失败'}", "success" if success else "warn")
        if success and match_product_library(barcode):
            result['queried'].append({'prefix': prefix, 'barcode': barcode})
            matched = match_product_library(barcode) or {}
            emit(
                f"前缀 {prefix} 已写入条码匹配：{matched.get('product_code', '')} / {matched.get('product_name', '')}",
                'success'
            )
        else:
            error = _brief_batch_error(message, 300)
            result['failed'].append({'prefix': prefix, 'barcode': barcode, 'error': error})
            emit(f"前缀 {prefix} 条码匹配补充失败：{error}", 'warn')
    emit(
        f"自动补充完成：成功 {len(result['queried'])} 个，失败 {len(result['failed'])} 个",
        "success" if not result['failed'] else "warn"
    )
    return result

def _crm_ready_for_auto_query(worker=None):
    worker = worker or crm_pool.get(kind="query")
    if worker.logged_in:
        return True, ""
    success, message = worker.check_login_status()
    if success and worker.logged_in:
        return True, ""
    return False, f"CRM 当前未登录，请先登录 CRM（{message or '会话未恢复'}）"

def _query_slot_label(slot_id):
    with crm_pool.pool_lock:
        slots = list(crm_pool.query_slots)
    if slot_id in slots:
        return f"查询{slots.index(slot_id) + 1}"
    return slot_id or "查询通道"

def _query_slot_has_running_batch(slot_id):
    with batch_job_lock:
        job_id = latest_batch_job_by_slot.get(slot_id)
        job = batch_jobs.get(job_id)
        return bool(job and job.get('running'))

def _query_slot_has_running_service_close(slot_id):
    with service_close_job_lock:
        job_id = latest_service_close_job_by_slot.get(slot_id)
        job = service_close_jobs.get(job_id)
        return bool(job and job.get('running'))

query_slot_cooldown_lock = threading.Lock()
query_slot_cooldowns = {}

def _mark_query_slot_healthy(slot_id):
    if not slot_id:
        return
    with query_slot_cooldown_lock:
        query_slot_cooldowns.pop(slot_id, None)

def _mark_query_slot_unhealthy(slot_id, reason=''):
    if not slot_id:
        return
    with query_slot_cooldown_lock:
        query_slot_cooldowns[slot_id] = {
            'until': time.time() + QUERY_SLOT_FAILURE_COOLDOWN_SECONDS,
            'reason': _brief_batch_error(reason, 160),
        }

def _query_slot_cooldown_message(slot_id):
    with query_slot_cooldown_lock:
        row = query_slot_cooldowns.get(slot_id)
        if not row:
            return ''
        remaining = int(row.get('until', 0) - time.time())
        if remaining <= 0:
            query_slot_cooldowns.pop(slot_id, None)
            return ''
        reason = row.get('reason') or '上次查询失败'
        return f"{reason}，冷却 {remaining} 秒"

def _select_idle_query_workers_desc(exclude_slot_ids=None):
    exclude_slot_ids = set(exclude_slot_ids or [])
    with crm_pool.pool_lock:
        slots = list(crm_pool.query_slots)
    workers = []
    logged_in_count = 0
    busy_count = 0
    skipped_cooldown = 0
    for slot_id in reversed(slots):
        if slot_id in exclude_slot_ids:
            continue
        worker = crm_pool.get(slot_id, "query")
        if worker.busy:
            busy_count += 1
            continue
        if not worker.logged_in:
            if not worker.remembered_logged_in:
                continue
            success, _message = worker.check_login_status()
            if not success or not worker.logged_in:
                continue
        logged_in_count += 1
        cooldown_message = _query_slot_cooldown_message(slot_id)
        if cooldown_message:
            skipped_cooldown += 1
            continue
        if _query_slot_has_running_batch(slot_id):
            continue
        if _query_slot_has_running_service_close(slot_id):
            continue
        workers.append((worker, slot_id, _query_slot_label(slot_id)))
    if workers:
        return workers, ""
    if logged_in_count == 0:
        if busy_count:
            return [], "所有已登录查询通道都在查询中，请稍后再试"
        return [], "没有已登录的查询通道，请先到在线查询页登录 CRM"
    if skipped_cooldown:
        return [], "已登录查询通道刚查询失败正在冷却，请稍后再试，或到在线查询页重新登录对应通道"
    return [], "所有已登录查询通道都在查询中，请稍后再试"

def _select_idle_query_worker_desc(exclude_slot_ids=None):
    workers, error = _select_idle_query_workers_desc(exclude_slot_ids)
    if workers:
        worker, slot_id, slot_label = workers[0]
        return worker, slot_id, slot_label, ""
    return None, "", "", error

def _request_slot_id(kind="query"):
    data = request.get_json(silent=True) or {}
    slot_id = (
        request.args.get("slot_id")
        or data.get("slot_id")
        or data.get("slot")
        or ""
    )
    return crm_pool.normalize_slot(slot_id, kind)

def _latest_job_id(mapping, slot_id):
    return mapping.get(slot_id) or ""

def _job_logs_since(job, since):
    logs = list(job.get('logs') or [])
    if since > 0:
        logs = [row for row in logs if int(row.get('id') or 0) > since]
    return logs

def _replace_html_field_values(html, field_ids, value):
    escaped_value = html_mod.escape(_clean_export_value(value), quote=False)
    changed = False
    for field_id in field_ids:
        pattern = re.compile(
            rf'(<div\s+id="{re.escape(field_id)}"[^>]*>.*?<span[^>]*>)([^<]*)(</span>)',
            re.DOTALL
        )

        def repl(match):
            nonlocal changed
            changed = True
            return match.group(1) + escaped_value + match.group(3)

        html = pattern.sub(repl, html)
    return html, changed

def _replace_closed_service_html(html, info):
    closed_map = (info or {}).get("closedServiceNos") or {}
    if not isinstance(closed_map, dict) or not closed_map:
        return html, False
    closed_set = {_clean_export_value(key) for key in closed_map if _clean_export_value(key)}
    if not closed_set:
        return html, False
    row_classes = set()
    serv_pattern = re.compile(
        r'(<div\s+id="servno1"\s+class="([^"]+)"[^>]*>.*?<span[^>]*>)([^<]*)(</span>)',
        re.DOTALL
    )
    for match in serv_pattern.finditer(html):
        if _clean_export_value(html_mod.unescape(match.group(3))) in closed_set:
            row_classes.add(match.group(2))
    if not row_classes:
        return html, False
    changed = False
    escaped_value = html_mod.escape("已结单", quote=False)
    for row_class in row_classes:
        status_pattern = re.compile(
            rf'(<div\s+id="newisclosed1"\s+class="{re.escape(row_class)}"[^>]*>.*?<span[^>]*>)([^<]*)(</span>)',
            re.DOTALL
        )

        def repl(match):
            nonlocal changed
            changed = True
            return match.group(1) + escaped_value + match.group(3)

        html = status_pattern.sub(repl, html)
    return html, changed

def _apply_barcode_html_overrides(html, info):
    changed = False
    dealer = _clean_export_value((info or {}).get('currentDealerOverride'))
    if dealer:
        html, dealer_changed = _replace_html_field_values(html, ["myproductdealer1", "dealername1"], dealer)
        changed = changed or dealer_changed
    html, service_changed = _replace_closed_service_html(html, info)
    changed = changed or service_changed
    return html, changed

def _apply_dealer_to_fields(fields, dealer):
    dealer = _clean_export_value(dealer)
    if not dealer:
        return fields
    if isinstance(fields.get('sr5'), dict):
        fields['sr5']['myproductdealer1'] = dealer
    if isinstance(fields.get('sr10'), dict):
        fields['sr10']['dealername1'] = dealer
    return fields

def _service_rows(fields):
    sub = (fields or {}).get('sr2')
    if isinstance(sub, list):
        return [row for row in sub if isinstance(row, dict)]
    if isinstance(sub, dict):
        return [sub]
    return []

def _parse_service_date(value):
    text = _clean_export_value(value)
    if not text:
        return 0
    normalized = (
        text.replace("年", "-")
        .replace("月", "-")
        .replace("日", "")
        .replace("/", "-")
        .replace(".", "-")
    )
    match = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?", normalized)
    if not match:
        return 0
    year, month, day, hour, minute, second = match.groups()
    try:
        return datetime(
            int(year),
            int(month),
            int(day),
            int(hour or 0),
            int(minute or 0),
            int(second or 0),
        ).timestamp()
    except Exception:
        return 0

def _parse_service_date_from_no(service_no):
    text = _clean_export_value(service_no)
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])", text)
    if not match:
        return 0
    try:
        year, month, day = match.groups()
        return datetime(int(year), int(month), int(day)).timestamp()
    except Exception:
        return 0

def _latest_service_record(fields):
    latest = None
    for index, row in enumerate(_service_rows(fields)):
        service_no = _clean_export_value(row.get("servno1"))
        if not service_no:
            continue
        score = _parse_service_date(row.get("servdate1")) or _parse_service_date_from_no(service_no) or (index + 1)
        if latest is None or score >= latest["score"]:
            latest = {"service_no": service_no, "row": row, "score": score}
    return latest

def _service_row_is_closed(row):
    text = _clean_export_value((row or {}).get("newisclosed1"))
    return text == "已结单"

def _apply_service_close_overrides(fields, info):
    closed_map = (info or {}).get("closedServiceNos") or {}
    if not isinstance(closed_map, dict) or not closed_map:
        return fields
    closed_set = {_clean_export_value(key) for key in closed_map if _clean_export_value(key)}
    if not closed_set:
        return fields
    for row in _service_rows(fields):
        if _clean_export_value(row.get("servno1")) in closed_set:
            row["newisclosed1"] = "已结单"
    return fields

def _record_service_closed_for_barcodes(service_no, barcodes):
    service_no = _clean_export_value(service_no)
    if not service_no:
        return 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data = load_data()
    changed = 0
    for barcode in normalize_input_barcodes(barcodes):
        info = data.get(barcode, {'remark': '', 'archived': False, 'archiveTime': '', 'archivedBy': ''})
        closed_map = info.get("closedServiceNos")
        if not isinstance(closed_map, dict):
            closed_map = {}
        closed_map[service_no] = now
        info["closedServiceNos"] = closed_map
        info["serviceCloseUpdatedAt"] = now
        data[barcode] = info
        changed += 1
    if changed:
        save_data(data)
    return changed

def _barcode_html_path(barcode):
    filename = barcode + '.html'
    for directory in (BARCODE_DIR, ARCHIVE_DIR):
        filepath = os.path.join(directory, filename)
        if os.path.exists(filepath):
            return filepath
    return ''

def _sync_barcode_html_dealer(barcode, dealer):
    filepath = _barcode_html_path(barcode)
    if not filepath:
        return False
    try:
        original_mtime = os.path.getmtime(filepath)
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()
        updated_html, changed = _replace_html_field_values(
            html,
            ["myproductdealer1", "dealername1"],
            dealer
        )
        if changed:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(updated_html)
            os.utime(filepath, (original_mtime, original_mtime))
        return changed
    except Exception:
        return False

def _apply_transfer_local_dealer(summary, transfer_type, distributor):
    new_dealer = own_dealer_name() if transfer_type == "移入" else distributor
    if not new_dealer:
        return
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = load_data()
    for detail in summary.get('details', []):
        barcode = _clean_export_value(detail.get('barcode'))
        if not barcode:
            continue
        info = data.get(barcode, {'remark': '', 'archived': False, 'archiveTime': '', 'archivedBy': ''})
        info['currentDealerOverride'] = new_dealer
        info['transferUpdatedAt'] = now
        info['transferType'] = transfer_type
        info['transferDistributor'] = distributor
        data[barcode] = info
        _sync_barcode_html_dealer(barcode, new_dealer)
    save_data(data)

def build_transfer_summary(selected_barcodes, transfer_type="移出", distributor=""):
    own_dealer = own_dealer_name()
    wanted = OrderedDict()
    excluded = []
    for barcode in selected_barcodes:
        barcode = _clean_export_value(barcode)
        if is_disassembly_barcode(barcode):
            excluded.append(barcode)
            continue
        if barcode and barcode not in wanted:
            wanted[barcode] = True

    all_items = {item['barcode']: item for item in scan_barcodes()}
    details = []
    missing = []
    incomplete = []
    blocked = []
    grouped = OrderedDict()

    for barcode in wanted:
        item = all_items.get(barcode)
        if not item:
            info = _barcode_product_info_from_library(barcode)
        else:
            info = _barcode_product_info(item)

        if info.get('source') == 'unmatched':
            missing.append(barcode)
            details.append(info)
            continue

        details.append(info)
        if not info['product_code'] or not info['product_name']:
            incomplete.append(barcode)
            continue
        if transfer_type == "移出" and info.get('current_dealer') and info['current_dealer'] != own_dealer:
            blocked.append({
                'barcode': barcode,
                'reason': f"当前所属为 {info['current_dealer']}，不是 {own_dealer}，请确认后再移出",
            })
        if transfer_type == "移入" and distributor and info.get('current_dealer') and info['current_dealer'] != distributor:
            blocked.append({
                'barcode': barcode,
                'reason': f"当前所属为 {info['current_dealer']}，不是所选分销商 {distributor}，请确认后再移入",
            })

        key = f"{info['product_code']}|{info['product_name']}"
        matched_prefix = info.get('matched_prefix') or product_prefix_from_barcode(barcode)
        if key not in grouped:
            grouped[key] = {
                'product_code': info['product_code'],
                'product_name': info['product_name'],
                'product_model': info.get('product_model') or info['product_name'],
                'matched_prefixes': [],
                'quantity': 0,
                'barcodes': [],
            }
        if matched_prefix and matched_prefix not in grouped[key]['matched_prefixes']:
            grouped[key]['matched_prefixes'].append(matched_prefix)
        grouped[key]['quantity'] += 1
        grouped[key]['barcodes'].append(barcode)

    return {
        'total': len(wanted),
        'details': details,
        'groups': list(grouped.values()),
        'missing': missing,
        'incomplete': incomplete,
        'blocked': blocked,
        'excluded': excluded,
    }

def selected_latest_service_orders(selected_barcodes):
    wanted = normalize_input_barcodes(selected_barcodes)
    all_items = {item['barcode']: item for item in scan_barcodes()}
    service_orders = OrderedDict()
    missing = []
    no_service = []
    for barcode in wanted:
        item = all_items.get(barcode)
        if not item:
            missing.append(barcode)
            continue
        latest = _latest_service_record(item.get('fields') or {})
        if not latest or not latest.get("service_no"):
            no_service.append(barcode)
            continue
        service_no = latest["service_no"]
        row = latest.get("row") or {}
        if service_no not in service_orders:
            service_orders[service_no] = {
                "service_no": service_no,
                "barcodes": [],
                "customer_names": [],
                "product_names": [],
                "local_closed": _service_row_is_closed(row),
            }
        service_orders[service_no]["barcodes"].append(barcode)
        customer_name = _clean_export_value(row.get("name1") or row.get("customer1"))
        if customer_name and customer_name not in service_orders[service_no]["customer_names"]:
            service_orders[service_no]["customer_names"].append(customer_name)
        product_name = _clean_export_value(row.get("newproductidName1") or row.get("newproductname1"))
        if product_name and product_name not in service_orders[service_no]["product_names"]:
            service_orders[service_no]["product_names"].append(product_name)
        if not _service_row_is_closed(row):
            service_orders[service_no]["local_closed"] = False
    for row in service_orders.values():
        row["display_label"] = _service_order_display(row)
    return {
        "orders": list(service_orders.values()),
        "missing": missing,
        "no_service": no_service,
    }

def _format_limited_values(values, limit=3):
    cleaned = []
    for value in values or []:
        text = _clean_export_value(value)
        if text and text not in cleaned:
            cleaned.append(text)
    if not cleaned:
        return ""
    shown = cleaned[:limit]
    suffix = f"等{len(cleaned)}个" if len(cleaned) > limit else ""
    return "、".join(shown) + suffix

def _service_order_display(row):
    row = row or {}
    service_no = _clean_export_value(row.get("service_no"))
    barcode_text = _format_limited_values(row.get("barcodes"), 4)
    customer_text = _format_limited_values(row.get("customer_names") or row.get("customers"), 2)
    parts = []
    if barcode_text:
        parts.append(f"条码 {barcode_text}")
    if customer_text:
        parts.append(f"客户 {customer_text}")
    if service_no:
        parts.append(f"服务单 {service_no}")
    return " / ".join(parts) or service_no or "未知服务单"

def queried_dealer_history():
    own_dealer = own_dealer_name()
    dealers = OrderedDict()
    for item in scan_barcodes():
        info = _barcode_product_info(item)
        for key in ("current_dealer", "service_dealer"):
            dealer = _clean_export_value(info.get(key))
            if dealer and dealer != own_dealer:
                dealers[dealer] = True
    return list(dealers.keys())

def load_distributor_history():
    own_dealer = own_dealer_name()
    services = _get_cluster_services()
    if services:
        return [
            row["name"]
            for row in services.catalog.list_distributors()
            if row["name"] != own_dealer
        ]
    if os.path.exists(DISTRIBUTOR_HISTORY_FILE):
        try:
            with open(DISTRIBUTOR_HISTORY_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [
                _clean_export_value(row)
                for row in (data if isinstance(data, list) else [])
                if _clean_export_value(row) and _clean_export_value(row) != own_dealer
            ]
        except Exception:
            pass
    return []

def _save_distributor_history_rows(rows):
    services = _get_cluster_services()
    if services:
        services.catalog.upsert_distributors(rows[:100])
        return
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(DISTRIBUTOR_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows[:100], f, ensure_ascii=False, indent=2)

def load_deleted_distributor_history():
    own_dealer = own_dealer_name()
    services = _get_cluster_services()
    if services:
        return [
            row["name"]
            for row in services.catalog.list_distributors(include_deleted=True)
            if row.get("deleted") and row["name"] != own_dealer
        ]
    if os.path.exists(DISTRIBUTOR_HISTORY_DELETED_FILE):
        try:
            with open(DISTRIBUTOR_HISTORY_DELETED_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [
                _clean_export_value(row)
                for row in (data if isinstance(data, list) else [])
                if _clean_export_value(row) and _clean_export_value(row) != own_dealer
            ]
        except Exception:
            pass
    return []

def save_deleted_distributor_history(rows):
    clean_rows = []
    seen = set()
    own_dealer = own_dealer_name()
    for row in rows:
        row = _clean_export_value(row)
        if row and row != own_dealer and row not in seen:
            clean_rows.append(row)
            seen.add(row)
    services = _get_cluster_services()
    if services:
        services.catalog.set_deleted_distributors(clean_rows[:300])
        return
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(DISTRIBUTOR_HISTORY_DELETED_FILE, 'w', encoding='utf-8') as f:
        json.dump(clean_rows[:300], f, ensure_ascii=False, indent=2)

def restore_deleted_distributor_history(distributor):
    distributor = _clean_export_value(distributor)
    if not distributor:
        return
    deleted = [row for row in load_deleted_distributor_history() if row != distributor]
    save_deleted_distributor_history(deleted)

def import_distributor_history_many(distributors):
    own_dealer = own_dealer_name()
    incoming = []
    seen = set()
    for distributor in distributors:
        distributor = _clean_export_value(distributor)
        if distributor and distributor != own_dealer and distributor not in seen:
            incoming.append(distributor)
            seen.add(distributor)
    if not incoming:
        return load_distributor_history()
    deleted = [row for row in load_deleted_distributor_history() if row not in seen]
    save_deleted_distributor_history(deleted)
    rows = OrderedDict()
    for distributor in incoming:
        rows[distributor] = True
    for distributor in load_distributor_history():
        if distributor and distributor != own_dealer:
            rows[distributor] = True
    _save_distributor_history_rows(list(rows.keys()))
    return load_distributor_history()

def save_distributor_history(distributor):
    own_dealer = own_dealer_name()
    distributor = _clean_export_value(distributor)
    if not distributor or distributor == own_dealer:
        return
    restore_deleted_distributor_history(distributor)
    rows = [distributor] + [row for row in load_distributor_history() if row != distributor]
    _save_distributor_history_rows(rows)

def save_distributor_history_many(distributors):
    own_dealer = own_dealer_name()
    deleted = set(load_deleted_distributor_history())
    rows = OrderedDict()
    for distributor in distributors:
        distributor = _clean_export_value(distributor)
        if distributor and distributor != own_dealer and distributor not in deleted:
            rows[distributor] = True
    for distributor in load_distributor_history():
        if distributor and distributor != own_dealer and distributor not in deleted:
            rows[distributor] = True
    _save_distributor_history_rows(list(rows.keys()))

def delete_distributor_history(distributor):
    own_dealer = own_dealer_name()
    distributor = _clean_export_value(distributor)
    if not distributor or distributor == own_dealer:
        return False
    services = _get_cluster_services()
    if services:
        return services.catalog.delete_distributor(distributor)
    rows = [row for row in load_distributor_history() if row != distributor]
    _save_distributor_history_rows(rows)
    deleted = [distributor] + [row for row in load_deleted_distributor_history() if row != distributor]
    save_deleted_distributor_history(deleted)
    return True

def combined_distributor_history():
    own_dealer = own_dealer_name()
    save_distributor_history_many(queried_dealer_history())
    deleted = set(load_deleted_distributor_history())
    dealers = OrderedDict()
    for dealer in load_distributor_history():
        dealer = _clean_export_value(dealer)
        if dealer and dealer != own_dealer and dealer not in deleted:
            dealers[dealer] = True
    return list(dealers.keys())

def load_data():
    services = _get_cluster_services()
    if services:
        result = {}
        for row in services.catalog.list_barcodes():
            metadata = dict(row.get("metadata") or {})
            metadata.update({
                "remark": row.get("remark") or metadata.get("remark") or "",
                "archived": bool(row.get("archived")),
                "archiveTime": row.get("archive_time") or metadata.get("archiveTime") or "",
                "currentDealerOverride": row.get("current_dealer_override") or metadata.get("currentDealerOverride") or "",
                "transferUpdatedAt": row.get("transfer_updated_at") or metadata.get("transferUpdatedAt") or "",
                "querySlotId": row.get("query_slot_id") or metadata.get("querySlotId") or "",
                "queryUpdatedAt": row.get("query_updated_at") or metadata.get("queryUpdatedAt") or "",
            })
            result[row["barcode"]] = metadata
        return result
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_data(data):
    services = _get_cluster_services()
    if services:
        for barcode, metadata in data.items():
            services.catalog.update_barcode_metadata(barcode, metadata or {})
        return
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _barcode_result_paths(barcode):
    barcode = _clean_export_value(barcode)
    if not barcode:
        return []
    filename = barcode + '.html'
    return [
        os.path.join(BARCODE_DIR, filename),
        os.path.join(ARCHIVE_DIR, filename),
    ]

def _barcode_temp_paths(barcode):
    barcode = _clean_export_value(barcode)
    if not barcode:
        return []
    return [os.path.join(TEMP_QUERY_DIR, barcode + '.html')]

def existing_barcode_result_paths(barcode):
    return {
        os.path.abspath(path)
        for path in _barcode_result_paths(barcode)
        if os.path.exists(path)
    }

def barcode_metadata_exists(barcode):
    barcode = _clean_export_value(barcode)
    return bool(barcode and barcode in load_data())

def delete_temporary_query_result(barcode, log=None, keep_paths=None, keep_metadata=False):
    barcode = _clean_export_value(barcode)
    if not barcode:
        return False
    keep_paths = {os.path.abspath(path) for path in (keep_paths or set())}
    removed_file = False
    paths = _barcode_temp_paths(barcode) + _barcode_result_paths(barcode)
    for path in paths:
        try:
            if os.path.exists(path) and os.path.abspath(path) not in keep_paths:
                os.remove(path)
                removed_file = True
        except Exception as e:
            if log:
                log(f"临时查询文件删除失败：{os.path.basename(path)}，{e}", "warn")
    data = load_data()
    removed_meta = bool(not keep_metadata and barcode in data)
    if removed_meta:
        data.pop(barcode, None)
        save_data(data)
    if log and (removed_file or removed_meta):
        log(f"已删除临时查询结果：{barcode}，不加入结果管理", "dim")
    return removed_file or removed_meta

def get_archived_set():
    data = load_data()
    return {bc for bc, info in data.items() if info.get('archived')}

def get_barcode_info(barcode):
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row:
            return {'remark': '', 'archived': False, 'archiveTime': '', 'archivedBy': ''}
        metadata = dict(row.get("metadata") or {})
        metadata.update({
            "remark": row.get("remark") or metadata.get("remark") or "",
            "archived": bool(row.get("archived")),
            "archiveTime": row.get("archive_time") or metadata.get("archiveTime") or "",
            "currentDealerOverride": row.get("current_dealer_override") or metadata.get("currentDealerOverride") or "",
            "transferUpdatedAt": row.get("transfer_updated_at") or metadata.get("transferUpdatedAt") or "",
            "querySlotId": row.get("query_slot_id") or metadata.get("querySlotId") or "",
            "queryUpdatedAt": row.get("query_updated_at") or metadata.get("queryUpdatedAt") or "",
        })
        return metadata
    data = load_data()
    return data.get(barcode, {'remark': '', 'archived': False, 'archiveTime': '', 'archivedBy': ''})

def update_barcode_info(barcode, info):
    services = _get_cluster_services()
    if services:
        services.catalog.update_barcode_metadata(barcode, info)
        return
    data = load_data()
    data[barcode] = info
    save_data(data)

def update_barcode_query_slot(barcode, slot_id):
    barcode = _clean_export_value(barcode)
    slot_id = _clean_export_value(slot_id)
    if not barcode or not slot_id or not slot_id.startswith('query-'):
        return
    info = get_barcode_info(barcode)
    info['querySlotId'] = slot_id
    info['querySlotLabel'] = _query_slot_label(slot_id)
    info['queryUpdatedAt'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    update_barcode_info(barcode, info)

def load_accounts():
    default_admin = {
        'id': 'admin',
        'username': 'admin',
        'display_name': '管理员',
        'password': '88293529',
        'permissions': ['crm', 'results', 'transfer', 'accounts', 'product-library'],
        'updated_at': '',
    }
    services = _get_cluster_services()
    if services:
        accounts = services.catalog.list_accounts()
        if not accounts:
            services.catalog.replace_accounts([default_admin])
            accounts = services.catalog.list_accounts()
        return accounts
    if os.path.exists(ACCOUNTS_FILE):
        try:
            with open(ACCOUNTS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            accounts = data if isinstance(data, list) else []
            if not any(row.get('username') == 'admin' for row in accounts):
                accounts.insert(0, default_admin)
                save_accounts(accounts)
            return accounts
        except Exception:
            pass
    save_accounts([default_admin])
    return [default_admin]

def save_accounts(accounts):
    services = _get_cluster_services()
    if services:
        services.catalog.replace_accounts(accounts)
        return
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(ACCOUNTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)

def account_public(row):
    return {
        'id': row.get('id', ''),
        'username': row.get('username', ''),
        'display_name': row.get('display_name', ''),
        'permissions': row.get('permissions', []),
        'updated_at': row.get('updated_at', ''),
        'is_admin': row.get('username') == 'admin' or bool(row.get('is_admin')),
    }

def desktop_account_row():
    return {
        'id': 'desktop',
        'username': 'desktop',
        'display_name': '本机管理员',
        'password': '',
        'permissions': ['crm', 'results', 'transfer', 'accounts', 'product-library'],
        'updated_at': '',
        'is_admin': True,
    }

def current_account():
    if IS_DESKTOP_APP:
        return desktop_account_row()
    username = session.get('account_username')
    if not username:
        return None
    return next((row for row in load_accounts() if row.get('username') == username), None)

def current_account_public():
    row = current_account()
    return account_public(row) if row else None

def crm_credentials_owner_key():
    row = current_account()
    if row and row.get("username"):
        return str(row.get("username"))
    return "desktop" if IS_DESKTOP_APP else ""

def load_crm_credentials_store():
    with crm_credentials_lock:
        try:
            if not os.path.exists(CRM_CREDENTIALS_FILE):
                return {}
            with open(CRM_CREDENTIALS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

def save_crm_credentials_store(data):
    with crm_credentials_lock:
        os.makedirs(os.path.dirname(CRM_CREDENTIALS_FILE), exist_ok=True)
        with open(CRM_CREDENTIALS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

def get_remembered_crm_credentials():
    key = crm_credentials_owner_key()
    if not key:
        return {"remember": False, "username": "", "password": ""}
    services = _get_cluster_services()
    if services:
        return services.catalog.get_credentials(key)
    row = load_crm_credentials_store().get(key) or {}
    if not isinstance(row, dict) or not row.get("remember"):
        return {"remember": False, "username": "", "password": ""}
    return {
        "remember": True,
        "username": str(row.get("username") or ""),
        "password": str(row.get("password") or ""),
    }

def save_remembered_crm_credentials(remember, username="", password=""):
    key = crm_credentials_owner_key()
    if not key:
        return False
    services = _get_cluster_services()
    if services:
        services.catalog.save_credentials(key, remember, username, password)
        return True
    data = load_crm_credentials_store()
    if remember:
        data[key] = {
            "remember": True,
            "username": str(username or "").strip(),
            "password": str(password or ""),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    else:
        data.pop(key, None)
    save_crm_credentials_store(data)
    return True


def authenticate_account(username, password):
    services = _get_cluster_services()
    if services:
        return services.catalog.authenticate_account(username, password)
    row = next((item for item in load_accounts() if item.get('username') == username), None)
    if not row or str(row.get('password') or '') != password:
        return None
    return row

PAGE_LINKS = [
    {'permission': 'crm', 'label': '在线查询', 'href': '/crm'},
    {'permission': 'results', 'label': '结果管理', 'href': '/'},
    {'permission': 'transfer', 'label': '移库', 'href': '/transfer'},
    {'permission': 'accounts', 'label': '账号管理', 'href': '/accounts'},
    {'permission': 'product-library', 'label': '条码匹配', 'href': '/product-library'},
]

DESKTOP_PAGE_LINKS = [
    {**link, 'label': '设置'} if link['permission'] == 'accounts' else dict(link)
    for link in PAGE_LINKS
]

def visible_page_links():
    if IS_DESKTOP_APP:
        return DESKTOP_PAGE_LINKS
    row = current_account()
    if not row:
        return []
    if row.get('username') == 'admin':
        return PAGE_LINKS
    permissions = set(row.get('permissions') or [])
    links = [link for link in PAGE_LINKS if link['permission'] in permissions]
    if not any(link['permission'] == 'accounts' for link in links):
        links.append({'permission': 'account-self', 'label': '账号管理', 'href': '/accounts'})
    return links

def is_admin_account():
    row = current_account()
    return bool(row and (row.get('username') == 'admin' or row.get('is_admin')))

def account_has_permission(permission):
    row = current_account()
    if not row:
        return False
    if row.get('username') == 'admin':
        return True
    return permission in (row.get('permissions') or [])

def required_permission_for_path(path):
    if path == "/" or path.startswith("/barcode/"):
        return "results"
    if path == "/crm":
        return "crm"
    if path == "/transfer" or path.startswith("/api/transfer") or path.startswith("/api/crm/transfer"):
        return "transfer"
    if path.startswith("/api/distributor-history"):
        return "transfer"
    if path == "/product-library" or path.startswith("/api/product-library"):
        return "product-library"
    if path == "/accounts":
        return "account-self"
    if path.startswith("/api/barcodes") or path.startswith("/api/filter-options") or path.startswith("/api/export"):
        return "results"
    if path.startswith("/api/service-close"):
        return "results"
    if path == "/api/crm/credentials":
        return "account-self"
    if path.startswith("/api/crm"):
        return "crm"
    return None

@app.before_request
def require_app_login():
    path = request.path
    if IS_DESKTOP_APP:
        return None
    if path.startswith("/api/app-auth"):
        return None
    if path == "/login":
        return None
    if path == "/product-library":
        return None
    if path == "/api/product-library" and request.method == "GET":
        return None
    if path == "/api/product-library/lookup":
        return None
    if path == "/api/product-library/query/start":
        return None
    if path == "/api/product-library/query/status":
        return None
    if path.startswith("/api/accounts"):
        if not current_account():
            return jsonify({'success': False, 'error': '请先登录工具账号'}), 401
        return None

    permission = required_permission_for_path(path)
    if not permission:
        return None

    if not current_account():
        if path.startswith("/api/"):
            return jsonify({'success': False, 'error': '请先登录工具账号'}), 401
        return redirect("/login?next=" + path)
    if permission != "account-self" and not account_has_permission(permission):
        if path.startswith("/api/"):
            return jsonify({'success': False, 'error': '当前账号无权访问该功能'}), 403
        return render_template("no_permission.html", account=current_account_public(), links=visible_page_links()), 403
    return None

@app.context_processor
def inject_app_flags():
    return {'is_desktop_app': IS_DESKTOP_APP}

def archive_barcode(barcode):
    services = _get_cluster_services()
    if services:
        if not services.catalog.get_barcode(barcode):
            return False, '文件不存在'
        info = get_barcode_info(barcode)
        info['archived'] = True
        info['archiveTime'] = datetime.now().astimezone().isoformat()
        services.catalog.update_barcode_metadata(barcode, info)
        return True, '归档成功'
    src = os.path.join(BARCODE_DIR, barcode + '.html')
    if not os.path.exists(src):
        return False, '文件不存在'
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    dst = os.path.join(ARCHIVE_DIR, barcode + '.html')
    try:
        os.rename(src, dst)
        info = get_barcode_info(barcode)
        info['archived'] = True
        info['archiveTime'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        update_barcode_info(barcode, info)
        return True, '归档成功'
    except Exception as e:
        return False, str(e)

def unarchive_barcode(barcode):
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row or not row.get("archived"):
            return False, '归档文件不存在'
        info = get_barcode_info(barcode)
        info['archived'] = False
        info['archiveTime'] = ''
        services.catalog.update_barcode_metadata(barcode, info)
        return True, '取消归档成功'
    src = os.path.join(ARCHIVE_DIR, barcode + '.html')
    if not os.path.exists(src):
        return False, '归档文件不存在'
    dst = os.path.join(BARCODE_DIR, barcode + '.html')
    try:
        os.rename(src, dst)
        info = get_barcode_info(barcode)
        info['archived'] = False
        info['archiveTime'] = ''
        update_barcode_info(barcode, info)
        return True, '取消归档成功'
    except Exception as e:
        return False, str(e)


def _cluster_barcode_to_legacy(row):
    metadata = row.get("metadata") or {}
    time_text = str(row.get("updated_at") or "")
    try:
        mtime = datetime.fromisoformat(time_text.replace("Z", "+00:00")).timestamp()
    except (TypeError, ValueError):
        mtime = 0
    fields = row.get("fields") or {}
    current_override = (
        row.get("current_dealer_override")
        or metadata.get("currentDealerOverride")
        or ""
    )
    fields = _apply_dealer_to_fields(fields, current_override)
    fields = _apply_service_close_overrides(fields, metadata)
    return {
        "barcode": row["barcode"],
        "filename": f"{row['barcode']}.html",
        "time": time_text,
        "mtime": mtime,
        "fields": fields,
        "currentDealerOverride": current_override,
        "transferUpdatedAt": row.get("transfer_updated_at") or metadata.get("transferUpdatedAt") or "",
        "querySlotId": row.get("query_slot_id") or metadata.get("querySlotId") or "",
        "querySlotLabel": metadata.get("querySlotLabel") or "",
        "queryUpdatedAt": row.get("query_updated_at") or metadata.get("queryUpdatedAt") or "",
        "remark": row.get("remark") or metadata.get("remark") or "",
        "archiveTime": row.get("archive_time") or metadata.get("archiveTime") or "",
    }

def scan_barcodes():
    services = _get_cluster_services()
    if services:
        return [
            _cluster_barcode_to_legacy(row)
            for row in services.catalog.list_barcodes()
        ]

    barcodes = []
    seen = set()

    def add_barcode_file(filepath, filename):
        barcode = filename.replace('.html', '')
        if barcode in seen:
            return
        seen.add(barcode)
        mtime = os.path.getmtime(filepath)
        time_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
        fields = extract_fields_from_html(filepath)
        info = get_barcode_info(barcode)
        fields = _apply_dealer_to_fields(fields, info.get('currentDealerOverride', ''))
        fields = _apply_service_close_overrides(fields, info)
        barcodes.append({
            'barcode': barcode,
            'filename': filename,
            'time': time_str,
            'mtime': mtime,
            'fields': fields,
            'currentDealerOverride': info.get('currentDealerOverride', ''),
            'transferUpdatedAt': info.get('transferUpdatedAt', ''),
            'querySlotId': info.get('querySlotId', ''),
            'querySlotLabel': info.get('querySlotLabel', ''),
            'queryUpdatedAt': info.get('queryUpdatedAt', ''),
            'remark': info.get('remark', ''),
        })

    for filename in os.listdir(BARCODE_DIR):
        if filename.endswith('.html'):
            add_barcode_file(os.path.join(BARCODE_DIR, filename), filename)

    if os.path.exists(ARCHIVE_DIR):
        for filename in os.listdir(ARCHIVE_DIR):
            if filename.endswith('.html'):
                add_barcode_file(os.path.join(ARCHIVE_DIR, filename), filename)

    barcodes.sort(key=lambda x: x['mtime'], reverse=True)
    return barcodes

def scan_archived():
    services = _get_cluster_services()
    if services:
        return [
            _cluster_barcode_to_legacy(row)
            for row in services.catalog.list_barcodes()
            if row.get("archived")
        ]

    barcodes = []
    if not os.path.exists(ARCHIVE_DIR):
        return barcodes
    for filename in os.listdir(ARCHIVE_DIR):
        if filename.endswith('.html'):
            barcode = filename.replace('.html', '')
            filepath = os.path.join(ARCHIVE_DIR, filename)
            mtime = os.path.getmtime(filepath)
            time_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            fields = extract_fields_from_html(filepath)
            info = get_barcode_info(barcode)
            fields = _apply_service_close_overrides(fields, info)
            barcodes.append({
                'barcode': barcode,
                'filename': filename,
                'time': time_str,
                'mtime': mtime,
                'fields': fields,
                'currentDealerOverride': info.get('currentDealerOverride', ''),
                'transferUpdatedAt': info.get('transferUpdatedAt', ''),
                'querySlotId': info.get('querySlotId', ''),
                'querySlotLabel': info.get('querySlotLabel', ''),
                'queryUpdatedAt': info.get('queryUpdatedAt', ''),
                'remark': info.get('remark', ''),
                'archiveTime': info.get('archiveTime', ''),
            })
    barcodes.sort(key=lambda x: x['mtime'], reverse=True)
    return barcodes

@app.route("/")
def index():
    return render_template("index.html", nav_links=visible_page_links(), business_config=business_config())

@app.route("/api/barcodes", methods=["GET"])
def api_get_barcodes():
    barcodes = scan_barcodes()
    return jsonify({
        'success': True,
        'total': len(barcodes),
        'barcodes': [{
            'barcode': b['barcode'],
            'filename': b['filename'],
            'time': b['time'],
            'fields': b['fields'],
            'currentDealerOverride': b.get('currentDealerOverride', ''),
            'transferUpdatedAt': b.get('transferUpdatedAt', ''),
            'querySlotId': b.get('querySlotId', ''),
            'querySlotLabel': b.get('querySlotLabel', ''),
            'queryUpdatedAt': b.get('queryUpdatedAt', ''),
            'remark': b.get('remark', ''),
        } for b in barcodes]
    })

@app.route("/api/barcodes/archived", methods=["GET"])
def api_get_archived():
    barcodes = scan_archived()
    return jsonify({
        'success': True,
        'total': len(barcodes),
        'barcodes': [{
            'barcode': b['barcode'],
            'filename': b['filename'],
            'time': b['time'],
            'fields': b['fields'],
            'currentDealerOverride': b.get('currentDealerOverride', ''),
            'transferUpdatedAt': b.get('transferUpdatedAt', ''),
            'querySlotId': b.get('querySlotId', ''),
            'querySlotLabel': b.get('querySlotLabel', ''),
            'queryUpdatedAt': b.get('queryUpdatedAt', ''),
            'remark': b.get('remark', ''),
            'archiveTime': b.get('archiveTime', ''),
        } for b in barcodes]
    })

@app.route("/api/filter-options", methods=["GET"])
def api_get_filter_options():
    barcodes = scan_barcodes()
    options = get_filter_options(barcodes)
    return jsonify({
        'success': True,
        'filters': list(options.values()),
        'total': len(barcodes),
    })

@app.route("/api/barcodes/<barcode>", methods=["GET"])
def api_get_barcode_detail(barcode):
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row:
            return jsonify({'success': False, 'error': '文件不存在'})
        legacy = _cluster_barcode_to_legacy(row)
        return jsonify({
            'success': True,
            'barcode': barcode,
            'time': legacy['time'],
            'fields': legacy['fields'],
        })
    filepath = _barcode_html_path(barcode)
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'})
    
    fields = extract_fields_from_html(filepath)
    info = get_barcode_info(barcode)
    fields = _apply_dealer_to_fields(fields, info.get('currentDealerOverride', ''))
    fields = _apply_service_close_overrides(fields, info)
    
    mtime = os.path.getmtime(filepath)
    time_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
    
    return jsonify({
        'success': True,
        'barcode': barcode,
        'time': time_str,
        'fields': fields
    })

@app.route("/api/export/xlsx", methods=["POST"])
def api_export_xlsx():
    data = request.get_json()
    selected_barcodes = data.get('barcodes', [])

    if not selected_barcodes:
        return jsonify({'success': False, 'error': '没有条码可导出'})

    if not HAS_OPENPYXL:
        return jsonify({
            'success': False,
            'error': '缺少 openpyxl 库，请运行: pip3 install openpyxl'
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "条码查询结果"
    ws.sheet_view.showGridLines = False

    group_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    group_font = Font(color="FFFFFF", bold=True, size=12)
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    export_rows = _prepare_export_rows(selected_barcodes)
    columns = _build_export_columns(export_rows)

    for col_idx, column in enumerate(columns, 1):
        group_cell = ws.cell(row=1, column=col_idx, value=column['group'])
        group_cell.fill = group_fill
        group_cell.font = group_font
        group_cell.alignment = Alignment(horizontal='center', vertical='center')
        group_cell.border = thin_border

        cell = ws.cell(row=2, column=col_idx, value=column['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    merge_start = 1
    for col_idx in range(2, len(columns) + 2):
        if col_idx <= len(columns) and columns[col_idx - 1]['group'] == columns[merge_start - 1]['group']:
            continue
        if col_idx - 1 > merge_start:
            ws.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col_idx - 1)
        merge_start = col_idx

    for row_idx, export_row in enumerate(export_rows, 3):
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_export_cell_value(export_row, column))
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for col_idx, column in enumerate(columns, 1):
        values = [_export_cell_value(row, column) for row in export_rows[:200]]
        ws.column_dimensions[get_column_letter(col_idx)].width = _suggest_column_width(column['label'], values)

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = 'A3'
    if columns:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{len(export_rows) + 2}"

    output_path = os.path.join(BARCODE_DIR, 'export_result.xlsx')
    wb.save(output_path)

    return jsonify({
        'success': True,
        'message': f'已导出 {len(selected_barcodes)} 条记录',
        'filename': 'export_result.xlsx'
    })

@app.route("/api/service-close/start", methods=["POST"])
def api_service_close_start():
    data = request.get_json() or {}
    barcodes = normalize_input_barcodes(data.get("barcodes") or [])
    if not barcodes:
        return jsonify({'success': False, 'error': '请先选择要结单的条码'})

    prepared = selected_latest_service_orders(barcodes)
    orders = prepared.get("orders") or []
    if not orders:
        no_service = prepared.get("no_service") or []
        missing = prepared.get("missing") or []
        reason = "所选条码没有可结单的服务单号"
        if no_service:
            reason += f"，无服务单 {len(no_service)} 个"
        if missing:
            reason += f"，未找到结果 {len(missing)} 个"
        return jsonify({'success': False, 'error': reason, **prepared})

    workers, error = _select_idle_query_workers_desc()
    if error:
        return jsonify({'success': False, 'error': error})
    slot_id, slot_label = workers[0][1], workers[0][2]
    slot_ids = [row[1] for row in workers]
    slot_label_text = "、".join(row[2] for row in workers)

    with service_close_job_lock:
        job = _empty_service_close_job(slot_id, orders)
        job.update({
            'running': True,
            'done': False,
            'success': False,
            'error': '',
            'missing': prepared.get("missing") or [],
            'no_service': prepared.get("no_service") or [],
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'finished_at': '',
        })
        if job['missing']:
            _append_job_log_unlocked(job, f"已跳过未找到结果的条码 {len(job['missing'])} 个：{', '.join(job['missing'][:10])}", "warn", 1000)
        if job['no_service']:
            _append_job_log_unlocked(job, f"已跳过无服务单条码 {len(job['no_service'])} 个：{', '.join(job['no_service'][:10])}", "warn", 1000)
        job['slot_ids'] = slot_ids
        _append_job_log_unlocked(job, f"已分配查询通道：{slot_label_text}", "info", 1000)
        service_close_jobs[job['job_id']] = job
        for selected_slot_id in slot_ids:
            latest_service_close_job_by_slot[selected_slot_id] = job['job_id']

    threading.Thread(target=_run_service_close_job, args=(job['job_id'], workers, orders), daemon=True).start()
    return jsonify({
        'success': True,
        'job_id': job['job_id'],
        'slot_id': slot_id,
        'slot_ids': slot_ids,
        'slot_label': slot_label_text,
        'orders': orders,
        'total': len(orders),
        'missing': prepared.get("missing") or [],
        'no_service': prepared.get("no_service") or [],
        'message': f'批量结单已开始，共 {len(orders)} 个服务单',
    })

@app.route("/api/service-close/status", methods=["GET"])
def api_service_close_status():
    try:
        since = int(request.args.get('since') or 0)
    except (TypeError, ValueError):
        since = 0
    slot_id = crm_pool.normalize_slot(request.args.get("slot_id"), "query")
    job_id = request.args.get("job_id") or _latest_job_id(latest_service_close_job_by_slot, slot_id)
    with service_close_job_lock:
        job = service_close_jobs.get(job_id) or _empty_service_close_job(slot_id)
        return jsonify({
            'success': True,
            'job_id': job.get('job_id') or '',
            'slot_id': job.get('slot_id') or slot_id,
            'slot_ids': job.get('slot_ids') or [job.get('slot_id') or slot_id],
            'running': job['running'],
            'done': job['done'],
            'close_success': job['success'],
            'error': job['error'],
            'total': job['total'],
            'current': job['current'],
            'closed_count': job.get('closed_count') or 0,
            'already_closed_count': job.get('already_closed_count') or 0,
            'failed_count': job.get('failed_count') or 0,
            'orders': job.get('orders') or [],
            'results': job.get('results') or [],
            'missing': job.get('missing') or [],
            'no_service': job.get('no_service') or [],
            'logs': _job_logs_since(job, since),
            'log_seq': job.get('log_seq') or 0,
            'started_at': job['started_at'],
            'finished_at': job['finished_at'],
        })

@app.route("/api/transfer/summary", methods=["POST"])
def api_transfer_summary():
    data = request.get_json() or {}
    slot_id = _request_slot_id("transfer")
    worker = crm_pool.get(slot_id, "transfer")
    barcodes = data.get('barcodes') or []
    barcodes = normalize_input_barcodes(barcodes)
    barcodes, excluded = filter_disassembly_barcodes(barcodes)
    distributor = str(data.get('distributor') or '').strip()
    transfer_type = str(data.get('transfer_type') or '移出').strip()
    if not barcodes:
        return jsonify({'success': False, 'error': '输入的条码都是拆机条码，无需移库' if excluded else '请先选择要移库的条码', 'excluded': excluded})

    auto_library = {'queried': [], 'failed': []}
    representatives = _missing_product_library_representatives(barcodes)
    if representatives:
        ready, ready_message = _crm_ready_for_auto_query(worker)
        if not ready:
            return jsonify({'success': False, 'error': ready_message})
        auto_library = ensure_product_library_for_barcodes(barcodes, None, worker)

    summary = build_transfer_summary(barcodes, transfer_type, distributor)
    summary['excluded'] = excluded
    summary['auto_library'] = auto_library
    _exclude_unmatched_transfer_barcodes(summary)
    if not summary.get('groups'):
        return jsonify({
            'success': False,
            'error': '本次没有可移库条码，未匹配到产品信息的条码已临时排除',
            'summary': summary,
        })
    if summary['missing']:
        return jsonify({
            'success': False,
            'error': '部分条码没有查询结果，也没有匹配到条码前缀，请先维护条码匹配或查询一次该产品条码',
            'summary': summary,
        })
    if summary['incomplete']:
        return jsonify({
            'success': False,
            'error': '部分条码缺少产品名称或产品编码，无法自动汇总',
            'summary': summary,
        })

    return jsonify({'success': True, 'summary': summary})

@app.route("/api/distributor-history", methods=["GET"])
def api_distributor_history():
    return jsonify({
        'success': True,
        'dealers': combined_distributor_history(),
        'can_delete': is_admin_account(),
    })

@app.route("/api/distributor-history", methods=["POST"])
def api_save_distributor_history():
    data = request.get_json() or {}
    distributors = data.get('distributors')
    if isinstance(distributors, list):
        if not is_admin_account():
            return jsonify({'success': False, 'error': '只有管理员可以批量导入目标分销商'}), 403
        import_distributor_history_many(distributors)
    else:
        save_distributor_history(data.get('distributor'))
    return jsonify({
        'success': True,
        'dealers': combined_distributor_history(),
        'can_delete': is_admin_account(),
    })

@app.route("/api/distributor-history", methods=["DELETE"])
def api_delete_distributor_history():
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以删除目标分销商历史'}), 403
    data = request.get_json(silent=True) or {}
    distributor = data.get('distributor')
    if not _clean_export_value(distributor):
        return jsonify({'success': False, 'error': '目标分销商不能为空'}), 400
    delete_distributor_history(distributor)
    return jsonify({
        'success': True,
        'dealers': combined_distributor_history(),
        'can_delete': True,
    })

@app.route("/api/transfer/summary/start", methods=["POST"])
def api_transfer_summary_start():
    data = request.get_json() or {}
    slot_id = _request_slot_id("transfer")
    worker = crm_pool.get(slot_id, "transfer")
    barcodes = data.get('barcodes') or []
    barcodes = normalize_input_barcodes(barcodes)
    barcodes, excluded = filter_disassembly_barcodes(barcodes)
    distributor = str(data.get('distributor') or '').strip()
    transfer_type = str(data.get('transfer_type') or '移出').strip()
    if not barcodes:
        return jsonify({'success': False, 'error': '输入的条码都是拆机条码，无需移库' if excluded else '请先选择要移库的条码', 'excluded': excluded})

    with summary_job_lock:
        running_job_id = latest_summary_job_by_slot.get(slot_id)
        running_job = summary_jobs.get(running_job_id)
        if running_job and running_job.get('running'):
            return jsonify({'success': False, 'error': f'{slot_id} 已有汇总预览正在执行，请等待完成'})
        job = _empty_summary_job(slot_id)
        job.update({
            'running': True,
            'done': False,
            'success': False,
            'error': '',
            'summary': None,
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'finished_at': '',
        })
        summary_jobs[job['job_id']] = job
        latest_summary_job_by_slot[slot_id] = job['job_id']

    threading.Thread(
        target=_run_summary_job,
        args=(job['job_id'], worker, barcodes, transfer_type, distributor, excluded),
        daemon=True
    ).start()
    return jsonify({'success': True, 'job_id': job['job_id'], 'slot_id': slot_id, 'message': '汇总预览已开始'})

@app.route("/api/transfer/summary/status", methods=["GET"])
def api_transfer_summary_status():
    slot_id = crm_pool.normalize_slot(request.args.get("slot_id"), "transfer")
    job_id = request.args.get("job_id") or _latest_job_id(latest_summary_job_by_slot, slot_id)
    with summary_job_lock:
        job = summary_jobs.get(job_id) or _empty_summary_job(slot_id)
        return jsonify({
            'success': True,
            'job_id': job.get('job_id') or '',
            'slot_id': job.get('slot_id') or slot_id,
            'running': job['running'],
            'done': job['done'],
            'summary_success': job['success'],
            'error': job['error'],
            'summary': job['summary'],
            'logs': list(job['logs']),
            'log_seq': job.get('log_seq') or 0,
            'started_at': job['started_at'],
            'finished_at': job['finished_at'],
        })

@app.route("/api/crm/transfer", methods=["POST"])
def api_crm_transfer():
    data = request.get_json() or {}
    slot_id = _request_slot_id("transfer")
    worker = crm_pool.get(slot_id, "transfer")
    barcodes = data.get('barcodes') or []
    barcodes = normalize_input_barcodes(barcodes)
    barcodes, excluded = filter_disassembly_barcodes(barcodes)
    distributor = str(data.get('distributor') or '').strip()
    transfer_type = str(data.get('transfer_type') or '移出').strip()
    remark = str(data.get('remark') or '').strip()

    if not barcodes:
        return jsonify({'success': False, 'error': '输入的条码都是拆机条码，无需移库' if excluded else '请先选择要移库的条码', 'excluded': excluded})
    if not distributor:
        return jsonify({'success': False, 'error': '目标分销商不能为空'})

    representatives = _missing_product_library_representatives(barcodes)
    if representatives:
        ready, ready_message = _crm_ready_for_auto_query(worker)
        if not ready:
            return jsonify({'success': False, 'error': ready_message})
        ensure_product_library_for_barcodes(barcodes, None, worker)

    summary = build_transfer_summary(barcodes, transfer_type, distributor)
    summary['excluded'] = excluded
    _exclude_unmatched_transfer_barcodes(summary)
    if not summary.get('groups'):
        return jsonify({
            'success': False,
            'error': '本次没有可移库条码，未匹配到产品信息的条码已临时排除',
            'summary': summary,
        })
    if summary['missing']:
        return jsonify({'success': False, 'error': '部分条码没有查询结果，也没有匹配到条码前缀，请先维护条码匹配或查询一次该产品条码', 'summary': summary})
    if summary['incomplete']:
        return jsonify({'success': False, 'error': '部分条码缺少产品名称或产品编码，无法自动移库', 'summary': summary})

    with transfer_job_lock:
        running_job_id = latest_transfer_job_by_slot.get(slot_id)
        running_job = transfer_jobs.get(running_job_id)
        if running_job and running_job.get('running'):
            return jsonify({'success': False, 'error': f'{slot_id} 已有移库任务正在执行，请等待完成后再提交'})
        job = _empty_transfer_job(slot_id, summary, distributor, transfer_type, remark)
        job.update({
            'running': True,
            'done': False,
            'success': False,
            'error': '',
            'result': None,
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'finished_at': '',
        })
        transfer_jobs[job['job_id']] = job
        latest_transfer_job_by_slot[slot_id] = job['job_id']

    threading.Thread(
        target=_run_transfer_job,
        args=(job['job_id'], worker, summary, distributor, transfer_type, remark),
        daemon=True
    ).start()

    return jsonify({
        'success': True,
        'started': True,
        'job_id': job['job_id'],
        'slot_id': slot_id,
        'message': '移库任务已开始，请查看日志',
        'summary': summary,
        'transfer': {
            'dealer': own_dealer_name(),
            'distributor': distributor,
            'transfer_type': transfer_type,
            'remark': remark,
        },
    })

@app.route("/api/crm/transfer/status", methods=["GET"])
def api_crm_transfer_status():
    slot_id = crm_pool.normalize_slot(request.args.get("slot_id"), "transfer")
    job_id = request.args.get("job_id") or _latest_job_id(latest_transfer_job_by_slot, slot_id)
    with transfer_job_lock:
        job = transfer_jobs.get(job_id) or _empty_transfer_job(slot_id)
        result = job.get('result') or {}
        order_no = result.get('order_no') if isinstance(result, dict) else ''
        if job['success'] and isinstance(result, dict) and result.get('pending_approval'):
            message = f"移库单已保存待审批：{order_no or '已保存'}"
        elif job['success']:
            message = f"移库单已创建：{order_no or '已保存'}"
        else:
            message = job['error']
        return jsonify({
            'success': True,
            'job_id': job.get('job_id') or '',
            'slot_id': job.get('slot_id') or slot_id,
            'running': job['running'],
            'done': job['done'],
            'transfer_success': job['success'],
            'error': job['error'],
            'message': message,
            'result': job['result'],
            'summary': job['summary'],
            'transfer': {
                'dealer': own_dealer_name(),
                'distributor': job.get('distributor', ''),
                'transfer_type': job.get('transfer_type', ''),
                'remark': job.get('remark', ''),
            },
            'logs': list(job['logs']),
            'log_seq': job.get('log_seq') or 0,
            'started_at': job['started_at'],
            'finished_at': job['finished_at'],
        })

@app.route("/barcode/<filename>")
def serve_barcode(filename):
    barcode = filename.rsplit('.', 1)[0]
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row or not row.get("object_key"):
            return Response("文件不存在", status=404, mimetype="text/plain")
        html = services.objects.get_bytes(row["object_key"]).decode("utf-8", errors="replace")
        info = row.get("metadata") or {}
        info["currentDealerOverride"] = row.get("current_dealer_override") or info.get("currentDealerOverride", "")
        html, _changed = _apply_barcode_html_overrides(html, info)
        return Response(html, mimetype="text/html")
    info = get_barcode_info(barcode)
    filepath = os.path.join(BARCODE_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()
        html, changed = _apply_barcode_html_overrides(html, info)
        if changed:
            return Response(html, mimetype='text/html')
        return send_from_directory(BARCODE_DIR, filename)
    filepath = os.path.join(ARCHIVE_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()
        html, changed = _apply_barcode_html_overrides(html, info)
        if not changed:
            return send_from_directory(ARCHIVE_DIR, filename)
        return Response(html, mimetype='text/html')
    return send_from_directory(ARCHIVE_DIR, filename)

@app.route("/barcode/archived/<filename>")
def serve_archived(filename):
    barcode = filename.rsplit('.', 1)[0]
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row or not row.get("archived") or not row.get("object_key"):
            return Response("文件不存在", status=404, mimetype="text/plain")
        html = services.objects.get_bytes(row["object_key"]).decode("utf-8", errors="replace")
        info = row.get("metadata") or {}
        info["currentDealerOverride"] = row.get("current_dealer_override") or info.get("currentDealerOverride", "")
        html, _changed = _apply_barcode_html_overrides(html, info)
        return Response(html, mimetype="text/html")
    info = get_barcode_info(barcode)
    filepath = os.path.join(ARCHIVE_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()
        html, changed = _apply_barcode_html_overrides(html, info)
        if not changed:
            return send_from_directory(ARCHIVE_DIR, filename)
        return Response(html, mimetype='text/html')
    return send_from_directory(ARCHIVE_DIR, filename)

@app.route("/api/barcodes/<barcode>", methods=["DELETE"])
def api_delete_barcode(barcode):
    services = _get_cluster_services()
    if services:
        row = services.catalog.get_barcode(barcode)
        if not row:
            return jsonify({'success': False, 'error': '文件不存在'})
        try:
            if row.get("object_key"):
                services.objects.delete(row["object_key"])
            services.catalog.delete_barcode(barcode)
            return jsonify({'success': True, 'message': f'已删除 {barcode}'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    filepath = os.path.join(BARCODE_DIR, barcode + '.html')
    if not os.path.exists(filepath):
        filepath = os.path.join(ARCHIVE_DIR, barcode + '.html')

    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'})

    try:
        os.remove(filepath)
        data = load_data()
        if barcode in data:
            del data[barcode]
            save_data(data)
        return jsonify({'success': True, 'message': f'已删除 {barcode}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route("/api/barcodes/<barcode>/remark", methods=["POST"])
def api_update_remark(barcode):
    data = request.get_json()
    remark = data.get('remark', '')
    info = get_barcode_info(barcode)
    info['remark'] = remark
    update_barcode_info(barcode, info)
    return jsonify({'success': True, 'remark': remark})

@app.route("/api/barcodes/<barcode>/archive", methods=["POST"])
def api_archive_barcode(barcode):
    success, msg = archive_barcode(barcode)
    return jsonify({'success': success, 'message': msg})

@app.route("/api/barcodes/<barcode>/unarchive", methods=["POST"])
def api_unarchive_barcode(barcode):
    success, msg = unarchive_barcode(barcode)
    return jsonify({'success': success, 'message': msg})

@app.route("/crm")
def crm_page():
    return render_template("crm.html", nav_links=visible_page_links())

@app.route("/transfer")
def transfer_page():
    return render_template(
        "transfer.html",
        nav_links=visible_page_links(),
        business_config=business_config(),
        account=current_account_public(),
    )

@app.route("/product-library")
def product_library_page():
    return render_template("product_library.html", nav_links=visible_page_links(), account=current_account_public())

@app.route("/accounts")
def accounts_page():
    return render_template("accounts.html", nav_links=visible_page_links())

@app.route("/login")
def login_page():
    if IS_DESKTOP_APP:
        return redirect("/product-library")
    return render_template("login.html")

@app.route("/api/product-library", methods=["GET"])
def api_product_library():
    rows = sorted(load_product_library().values(), key=lambda row: row.get('prefix', ''))
    return jsonify({'success': True, 'products': rows, 'can_edit': is_admin_account(), 'account': current_account_public()})

@app.route("/api/product-library/lookup")
def api_product_library_lookup():
    barcode = str(request.args.get('barcode') or '').strip()
    if not barcode:
        return jsonify({'success': False, 'error': '请输入条码'})
    info = lookup_product_by_barcode(barcode)
    if not info or info.get('source') == 'unmatched':
        return jsonify({
            'success': False,
            'error': '条码匹配没有匹配到该条码前缀',
            'barcode': barcode,
            'prefix': product_prefix_from_barcode(barcode),
            'need_query': True,
        })
    return jsonify({'success': True, 'info': info})

@app.route("/api/product-library/query/start", methods=["POST"])
def api_product_library_query_start():
    data = request.get_json() or {}
    barcode = str(data.get('barcode') or '').strip()
    if not barcode:
        return jsonify({'success': False, 'error': '请输入条码'})
    if is_disassembly_barcode(barcode):
        return jsonify({'success': False, 'error': '这是拆机条码，CRM 不查询，也不写入条码匹配'})
    with library_query_lock:
        if library_query_job['running']:
            return jsonify({'success': False, 'error': '已有条码匹配查询正在执行'})

    worker, slot_id, slot_label, error = _select_idle_query_worker_desc()
    if error:
        return jsonify({'success': False, 'error': error})

    with library_query_lock:
        if library_query_job['running']:
            return jsonify({'success': False, 'error': '已有条码匹配查询正在执行'})
        library_query_job.update({
            'running': True,
            'done': False,
            'success': False,
            'barcode': barcode,
            'error': '',
            'logs': [],
            'slot_id': slot_id,
            'slot_label': slot_label,
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'finished_at': '',
        })
    _library_query_log("准备启动 CRM 查询...", 'info')
    _library_query_log(f"已分配查询通道：{slot_label}", 'info')
    threading.Thread(target=_run_library_query_job, args=(barcode, worker, slot_id, slot_label), daemon=True).start()
    return jsonify({'success': True, 'slot_id': slot_id, 'slot_label': slot_label, 'message': '条码查询已开始'})

@app.route("/api/product-library/query/status")
def api_product_library_query_status():
    with library_query_lock:
        return jsonify({
            'success': True,
            'running': library_query_job['running'],
            'done': library_query_job['done'],
            'query_success': library_query_job['success'],
            'barcode': library_query_job['barcode'],
            'slot_id': library_query_job.get('slot_id', ''),
            'slot_label': library_query_job.get('slot_label', ''),
            'error': library_query_job['error'],
            'logs': list(library_query_job['logs']),
            'started_at': library_query_job['started_at'],
            'finished_at': library_query_job['finished_at'],
        })

@app.route("/api/product-library", methods=["POST"])
def api_product_library_save():
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以修改条码匹配'})
    data = request.get_json() or {}
    prefix = str(data.get('prefix') or '').strip()
    product_code = str(data.get('product_code') or '').strip()
    product_name = str(data.get('product_name') or '').strip()
    if not prefix or not product_code or not product_name:
        return jsonify({'success': False, 'error': '前缀、产品编码、产品名称都不能为空'})
    upsert_product_library(prefix, product_code, product_name, data.get('source_barcode') or '')
    return jsonify({'success': True})

@app.route("/api/product-library/<prefix>", methods=["DELETE"])
def api_product_library_delete(prefix):
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以删除条码匹配规则'})
    data = load_product_library()
    if prefix in data:
        del data[prefix]
        save_product_library(data)
    return jsonify({'success': True})

@app.route("/api/accounts", methods=["GET"])
def api_accounts():
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以查看账号列表', 'account': current_account_public(), 'accounts': []})
    return jsonify({'success': True, 'account': current_account_public(), 'accounts': [account_public(row) for row in load_accounts()]})

@app.route("/api/accounts", methods=["POST"])
def api_accounts_save():
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以维护账号'})
    data = request.get_json() or {}
    account_id = str(data.get('id') or '').strip()
    username = str(data.get('username') or '').strip()
    display_name = str(data.get('display_name') or '').strip()
    password = str(data.get('password') or '').strip()
    permissions = data.get('permissions') or []
    if not username:
        return jsonify({'success': False, 'error': '账号不能为空'})
    if not isinstance(permissions, list):
        permissions = []
    allowed = {'crm', 'results', 'transfer', 'accounts', 'product-library'}
    permissions = [p for p in permissions if p in allowed]
    accounts = load_accounts()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if account_id:
        for row in accounts:
            if row.get('id') == account_id:
                row.update({
                    'username': username,
                    'display_name': display_name,
                    'permissions': permissions,
                    'updated_at': now,
                })
                if password:
                    row['password'] = password
                save_accounts(accounts)
                return jsonify({'success': True})
    if any(row.get('username') == username for row in accounts):
        return jsonify({'success': False, 'error': '账号已存在'})
    accounts.append({
        'id': uuid.uuid4().hex,
        'username': username,
        'display_name': display_name,
        'password': password,
        'permissions': permissions,
        'updated_at': now,
    })
    save_accounts(accounts)
    return jsonify({'success': True})

@app.route("/api/accounts/<account_id>", methods=["DELETE"])
def api_accounts_delete(account_id):
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以删除账号'})
    if account_id == 'admin':
        return jsonify({'success': False, 'error': '默认管理员账号不能删除'})
    accounts = [row for row in load_accounts() if row.get('id') != account_id]
    save_accounts(accounts)
    return jsonify({'success': True})

@app.route("/api/app-auth/status")
def api_app_auth_status():
    return jsonify({'success': True, 'account': current_account_public(), 'is_admin': is_admin_account()})

@app.route("/api/runtime-config", methods=["GET"])
def api_runtime_config():
    config = load_runtime_config()
    return jsonify({
        'success': True,
        'config': config,
        'active': {
            'query_workers': len(crm_pool.query_slots),
            'transfer_workers': len(crm_pool.transfer_slots),
        },
        'can_edit': is_admin_account(),
    })

@app.route("/api/runtime-config", methods=["POST"])
def api_runtime_config_save():
    if not (is_admin_account() or _cluster_admin_authorized()):
        return jsonify({'success': False, 'error': '只有管理员可以修改系统配置'})
    data = request.get_json() or {}
    config, slots = _save_local_runtime_config_from_payload(data)
    return jsonify({
        'success': True,
        'config': config,
        'active': {
            'query_workers': len(crm_pool.query_slots),
            'transfer_workers': len(crm_pool.transfer_slots),
        },
        'slots': slots,
    })

@app.route("/api/cluster/nodes")
def api_cluster_nodes():
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以查看服务器配置'}), 403
    nodes = load_cluster_nodes()
    return jsonify({
        'success': True,
        'local_node_id': _node_identity().get("id"),
        'nodes': [_cluster_node_status(node) for node in nodes],
    })

@app.route("/api/cluster/nodes/<node_id>/runtime-config", methods=["POST"])
def api_cluster_node_runtime_config_save(node_id):
    if not is_admin_account():
        return jsonify({'success': False, 'error': '只有管理员可以修改服务器配置'}), 403
    data = request.get_json() or {}
    node = next((row for row in load_cluster_nodes() if row.get("id") == node_id), None)
    if not node:
        return jsonify({'success': False, 'error': '未找到这个服务器节点'}), 404
    if node.get("id") == _node_identity().get("id"):
        config, slots = _save_local_runtime_config_from_payload(data)
        return jsonify({
            'success': True,
            'node': _cluster_node_status(_current_cluster_node() or node),
            'config': config,
            'slots': slots,
        })
    ok, error_message, result = _post_node_runtime_config(node, data)
    if not ok:
        return jsonify({'success': False, 'error': error_message or '远程服务器保存失败', 'node': node}), 502
    return jsonify({
        'success': True,
        'node': _cluster_node_status(node),
        'remote': result,
    })

@app.route("/api/app-auth/login", methods=["POST"])
def api_app_auth_login():
    data = request.get_json() or {}
    username = str(data.get('username') or '').strip()
    password = str(data.get('password') or '')
    row = authenticate_account(username, password)
    if not row:
        return jsonify({'success': False, 'error': '账号或密码错误'})
    session['account_username'] = username
    return jsonify({'success': True, 'account': account_public(row)})

@app.route("/api/app-auth/logout", methods=["POST"])
def api_app_auth_logout():
    session.pop('account_username', None)
    return jsonify({'success': True})

@app.route("/logout")
def app_logout_page():
    session.pop('account_username', None)
    if IS_DESKTOP_APP:
        return redirect("/product-library")
    return redirect("/login")

@app.route("/api/app-auth/password", methods=["POST"])
def api_app_auth_password():
    row = current_account()
    if not row:
        return jsonify({'success': False, 'error': '请先登录工具账号'})
    data = request.get_json() or {}
    old_password = str(data.get('old_password') or '')
    new_password = str(data.get('new_password') or '')
    if not authenticate_account(str(row.get('username') or ''), old_password):
        return jsonify({'success': False, 'error': '原密码不正确'})
    if not new_password:
        return jsonify({'success': False, 'error': '新密码不能为空'})
    accounts = load_accounts()
    for item in accounts:
        if item.get('username') == row.get('username'):
            item['password'] = new_password
            item['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            break
    save_accounts(accounts)
    return jsonify({'success': True})

@app.route("/api/crm/status")
def api_crm_status():
    kind = request.args.get("kind") or "query"
    slot_id = _request_slot_id(kind)
    worker = crm_pool.get(slot_id, kind)
    return jsonify({
        'slot_id': worker.slot_id,
        'browser_running': worker.is_alive(),
        'logged_in': worker.logged_in,
        'remembered_logged_in': worker.remembered_logged_in,
    })

@app.route("/api/crm/slots")
def api_crm_slots():
    return jsonify({'success': True, **crm_pool.slots_payload()})

@app.route("/healthz")
def healthz():
    return jsonify({
        "ok": True,
        "node": _node_identity(),
    })

@app.route("/readyz")
def readyz():
    payload = _node_status_payload(include_checks=True)
    status = 200 if payload.get("ready") else 503
    return jsonify(payload), status

@app.route("/api/node/status")
def api_node_status():
    include_checks = request.args.get("checks") in ("1", "true", "yes")
    return jsonify(_node_status_payload(include_checks=include_checks))

@app.route("/api/crm/credentials", methods=["GET", "POST"])
def api_crm_credentials():
    if request.method == "GET":
        return jsonify({"success": True, **get_remembered_crm_credentials()})
    data = request.get_json(silent=True) or {}
    remember = bool(data.get("remember"))
    username = str(data.get("username") or "").strip()
    password = str(data.get("password") or "")
    if remember and (not username or not password):
        return jsonify({"success": False, "error": "请输入 CRM 账号和密码"})
    if not save_remembered_crm_credentials(remember, username, password):
        return jsonify({"success": False, "error": "保存失败"})
    return jsonify({"success": True, "remember": remember})

@app.route("/api/crm/transfer/slots-status")
def api_crm_transfer_slots_status():
    payload = crm_pool.slots_payload()
    rows = payload.get("transfer", [])
    with transfer_job_lock:
        transfer_snapshot = {
            slot_id: dict(transfer_jobs.get(job_id) or {})
            for slot_id, job_id in latest_transfer_job_by_slot.items()
        }
    with summary_job_lock:
        summary_snapshot = {
            slot_id: dict(summary_jobs.get(job_id) or {})
            for slot_id, job_id in latest_summary_job_by_slot.items()
        }
    for row in rows:
        slot_id = row.get("id")
        transfer_job_row = transfer_snapshot.get(slot_id) or {}
        summary_job_row = summary_snapshot.get(slot_id) or {}
        row["transfer_running"] = bool(transfer_job_row.get("running"))
        row["summary_running"] = bool(summary_job_row.get("running"))
        row["transfer_job_id"] = transfer_job_row.get("job_id", "")
        row["summary_job_id"] = summary_job_row.get("job_id", "")
    return jsonify({'success': True, 'transfer': rows})

@app.route("/api/crm/login", methods=["POST"])
def api_crm_login():
    """统一登录接口：自动填账号密码 → 点登录 → 点发送验证码 → 等待用户输入验证码填入 → 点确定"""
    data = request.get_json()
    slot_id = _request_slot_id(data.get("kind") or "query")
    worker = crm_pool.get(slot_id, data.get("kind") or "query")
    username = data.get('username', '').strip()
    password = data.get('password', '')
    captcha = data.get('captcha', '').strip()

    if captcha:
        success, msg = worker.login_step2(captcha)
    else:
        success, msg = worker.login(username, password)
    return jsonify({'success': success, 'message': msg, 'slot_id': worker.slot_id})

@app.route("/api/crm/login-step1", methods=["POST"])
def api_crm_login_step1():
    """Step1: 填账号密码 → 点登录 → 点发送验证码 → 返回给前端"""
    data = request.get_json()
    slot_id = _request_slot_id(data.get("kind") or "query")
    worker = crm_pool.get(slot_id, data.get("kind") or "query")
    username = data.get('username', '').strip()
    password = data.get('password', '')
    success, msg = worker.login_step1(username, password)
    return jsonify({'success': success, 'message': msg, 'slot_id': worker.slot_id})

@app.route("/api/crm/login-step2", methods=["POST"])
def api_crm_login_step2():
    """Step2: 收到验证码后提交"""
    data = request.get_json()
    slot_id = _request_slot_id(data.get("kind") or "query")
    worker = crm_pool.get(slot_id, data.get("kind") or "query")
    captcha = data.get('captcha', '').strip()
    success, msg = worker.login_step2(captcha)
    return jsonify({'success': success, 'message': msg, 'slot_id': worker.slot_id})

def _bulk_login_status_payload(job):
    slots, waiting, active, success_count, failed_count = _bulk_login_slot_snapshot(job)
    return {
        'success': True,
        'job_id': job.get('job_id') or '',
        'scope': job.get('scope') or '',
        'running': bool(job.get('running')),
        'done': bool(job.get('done')),
        'login_success': bool(job.get('success')),
        'error': job.get('error') or '',
        'waiting_captcha': bool(waiting),
        'captcha_received': bool(job.get('captcha')),
        'pending_slots': waiting,
        'active_slots': active,
        'slots': slots,
        'total': len(slots),
        'success_count': success_count,
        'failed_count': failed_count,
        'logs': list(job.get('logs') or []),
        'log_seq': job.get('log_seq') or 0,
        'started_at': job.get('started_at') or '',
        'finished_at': job.get('finished_at') or '',
    }

@app.route("/api/crm/bulk-login/start", methods=["POST"])
def api_crm_bulk_login_start():
    data = request.get_json() or {}
    scope = str(data.get('scope') or data.get('kind') or 'query').strip() or 'query'
    username = str(data.get('username') or '').strip()
    password = str(data.get('password') or '')
    if not username or not password:
        return jsonify({'success': False, 'error': '请输入 CRM 账号和密码'})
    scope, slots = _bulk_login_slots_for_scope(scope)
    with bulk_login_job_lock:
        running_id = latest_bulk_login_job_by_scope.get(scope)
        running_job = bulk_login_jobs.get(running_id)
        if running_job and running_job.get('running'):
            return jsonify(_bulk_login_status_payload(running_job))
        job = _empty_bulk_login_job(scope, slots)
        job.update({
            'running': bool(slots),
            'done': not bool(slots),
            'success': not bool(slots),
            'username': username,
            'password': password,
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'finished_at': '' if slots else datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        if not slots:
            _append_job_log_unlocked(job, '所有 CRM 通道都已登录', 'success', 1000)
        bulk_login_jobs[job['job_id']] = job
        latest_bulk_login_job_by_scope[scope] = job['job_id']
        payload = _bulk_login_status_payload(job)
    if slots:
        threading.Thread(target=_run_bulk_login_job, args=(job['job_id'], username, password), daemon=True).start()
    return jsonify(payload)

@app.route("/api/crm/bulk-login/status")
def api_crm_bulk_login_status():
    scope = str(request.args.get('scope') or 'query').strip() or 'query'
    job_id = request.args.get('job_id') or latest_bulk_login_job_by_scope.get(scope)
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return jsonify({'success': False, 'error': '没有批量登录任务'})
        return jsonify(_bulk_login_status_payload(job))

@app.route("/api/crm/bulk-login/captcha", methods=["POST"])
def api_crm_bulk_login_captcha():
    data = request.get_json() or {}
    scope = str(data.get('scope') or 'query').strip() or 'query'
    job_id = data.get('job_id') or latest_bulk_login_job_by_scope.get(scope)
    captcha = str(data.get('captcha') or '').strip()
    if not captcha:
        return jsonify({'success': False, 'error': '请输入验证码'})
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return jsonify({'success': False, 'error': '没有批量登录任务'})
        job['captcha'] = captcha
        _append_job_log_unlocked(job, '已收到验证码，正在同步提交到等待中的通道', 'info', 1000)
        payload = _bulk_login_status_payload(job)
    threading.Thread(target=_submit_bulk_login_pending, args=(job_id,), daemon=True).start()
    return jsonify(payload)

@app.route("/api/crm/bulk-login/cancel", methods=["POST"])
def api_crm_bulk_login_cancel():
    data = request.get_json() or {}
    scope = str(data.get('scope') or 'query').strip() or 'query'
    job_id = data.get('job_id') or latest_bulk_login_job_by_scope.get(scope)
    with bulk_login_job_lock:
        job = bulk_login_jobs.get(job_id)
        if not job:
            return jsonify({'success': True})
        job['stop_requested'] = True
        job['running'] = False
        job['done'] = True
        job['error'] = '批量登录已取消'
        job['finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _append_job_log_unlocked(job, '批量登录已取消', 'warn', 1000)
        return jsonify(_bulk_login_status_payload(job))

@app.route("/api/crm/logout", methods=["POST"])
def api_crm_logout():
    kind = request.args.get("kind") or "query"
    slot_id = _request_slot_id(kind)
    crm_pool.get(slot_id, kind).logout()
    return jsonify({'success': True, 'slot_id': slot_id})

@app.route("/api/crm/login/cancel", methods=["POST"])
def api_crm_login_cancel():
    kind = request.args.get("kind") or "query"
    slot_id = _request_slot_id(kind)
    crm_pool.get(slot_id, kind).cancel_login()
    return jsonify({'success': True, 'slot_id': slot_id})

@app.route("/api/crm/check-login", methods=["POST"])
def api_crm_check_login():
    """手动完成登录后调用此接口，后端检查浏览器状态并更新登录状态"""
    kind = request.args.get("kind") or "query"
    slot_id = _request_slot_id(kind)
    worker = crm_pool.get(slot_id, kind)
    success, msg = worker.check_login_status()
    return jsonify({'success': success, 'message': msg, 'logged_in': worker.logged_in, 'slot_id': worker.slot_id})

@app.route("/api/crm/query", methods=["POST"])
def api_crm_query():
    data = request.get_json()
    slot_id = _request_slot_id(data.get("kind") or "query")
    worker = crm_pool.get(slot_id, data.get("kind") or "query")
    barcode = data.get('barcode', '').strip()
    if not barcode:
        return jsonify({'success': False, 'error': '条码不能为空'})
    if is_disassembly_barcode(barcode):
        return jsonify({'success': False, 'error': '这是拆机条码，CRM 不查询'})
    success, result = worker.query_barcode(barcode)
    if success:
        success, publish_error = publish_cluster_query_result(barcode, worker.slot_id)
        if not success:
            return jsonify({'success': False, 'error': publish_error})
        update_barcode_query_slot(barcode, worker.slot_id)
        return jsonify({
            'success': True,
            'slot_id': worker.slot_id,
            'barcode': result,
            'view_url': f'/barcode/{barcode}.html'
        })
    else:
        return jsonify({'success': False, 'error': result})

@app.route("/api/crm/batch/start", methods=["POST"])
def api_crm_batch_start():
    data = request.get_json()
    slot_id = _request_slot_id("query")
    worker = crm_pool.get(slot_id, "query")
    barcodes = data.get('barcodes') or []
    barcodes = normalize_input_barcodes(barcodes)
    barcodes, excluded = filter_disassembly_barcodes(barcodes)
    retry_limit = _normalize_retry_limit(data.get('retry_limit', DEFAULT_BATCH_RETRY_LIMIT))
    if not barcodes:
        return jsonify({'success': False, 'error': '输入的条码都是拆机条码，无需查询' if excluded else '条码不能为空', 'excluded': excluded})

    with batch_job_lock:
        running_job_id = latest_batch_job_by_slot.get(slot_id)
        running_job = batch_jobs.get(running_job_id)
        if running_job and running_job.get('running'):
            return jsonify({'success': False, 'error': f'{slot_id} 已有批量查询正在运行'})
        job = _empty_batch_job(slot_id, barcodes, retry_limit)
        job.update({
            'running': True,
            'started_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        batch_jobs[job['job_id']] = job
        latest_batch_job_by_slot[slot_id] = job['job_id']

    t = threading.Thread(target=_run_batch_job, args=(job['job_id'], worker, barcodes, retry_limit, excluded), daemon=True)
    t.start()
    return jsonify({'success': True, 'job_id': job['job_id'], 'slot_id': slot_id, 'total': len(barcodes), 'retry_limit': retry_limit, 'excluded': excluded})

@app.route("/api/crm/batch/status")
def api_crm_batch_status():
    try:
        since = int(request.args.get('since') or 0)
    except (TypeError, ValueError):
        since = 0
    include_results = request.args.get('include_results') in ('1', 'true', 'yes')
    slot_id = crm_pool.normalize_slot(request.args.get("slot_id"), "query")
    job_id = request.args.get("job_id") or _latest_job_id(latest_batch_job_by_slot, slot_id)
    with batch_job_lock:
        job = batch_jobs.get(job_id) or _empty_batch_job(slot_id)
        logs = _job_logs_since(job, since)
        payload = {
            'success': True,
            'job_id': job.get('job_id') or '',
            'slot_id': job.get('slot_id') or slot_id,
            'running': job['running'],
            'stop_requested': job['stop_requested'],
            'total': job['total'],
            'current': job['current'],
            'success_count': job['success'],
            'failed_count': job['failed'],
            'retry_limit': job['retry_limit'],
            'log_seq': job.get('log_seq') or 0,
            'logs': logs,
            'results_count': len(job['results']),
            'started_at': job.get('started_at', ''),
            'finished_at': job.get('finished_at', ''),
        }
        if include_results:
            payload['results'] = list(job['results'])
        return jsonify(payload)

@app.route("/api/crm/batch/stop", methods=["POST"])
def api_crm_batch_stop():
    data = request.get_json(silent=True) or {}
    slot_id = crm_pool.normalize_slot(request.args.get("slot_id") or data.get("slot_id"), "query")
    job_id = request.args.get("job_id") or data.get("job_id") or _latest_job_id(latest_batch_job_by_slot, slot_id)
    with batch_job_lock:
        job = batch_jobs.get(job_id)
        if job and job['running']:
            job['stop_requested'] = True
            return jsonify({'success': True, 'job_id': job_id, 'slot_id': slot_id})
    return jsonify({'success': False, 'error': '没有正在运行的批量查询'})

if __name__ == "__main__":
    print("=" * 60)
    print("怡口 CRM 条码查询结果页面")
    print("=" * 60)
    print("请访问: http://localhost:5001")
    print("=" * 60)

    os.makedirs(BARCODE_DIR, exist_ok=True)

    app.run(host="0.0.0.0", port=5001, debug=False)
