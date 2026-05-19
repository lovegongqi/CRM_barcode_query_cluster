#!/usr/bin/env python3
"""HTML 转 Excel 转换器 - 使用 BeautifulSoup 解析"""
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_subreports(html_content):
    """解析所有 Subreport 区块"""
    soup = BeautifulSoup(html_content, 'lxml')

    # 找到所有 Subreport div
    subreports = soup.find_all('div', id=re.compile(r'^Subreport\d+$'))

    results = []
    for i, sub in enumerate(subreports):
        # Subreport div 后面的那个 div 包含实际内容
        next_sibling = sub.find_next_sibling()

        if next_sibling:
            # 解析这个内容块
            result = parse_single_block(next_sibling)
            if result['title']:
                results.append(result)

    return results

def get_span_text_and_class(div):
    """获取 div 中所有 span 的文本和类"""
    spans = div.find_all('span')
    texts = []
    classes = []
    for span in spans:
        text = span.get_text().strip()
        if text:
            texts.append(text)
        span_class = span.get('class', [])
        if span_class:
            classes.extend(span_class)
    return texts, classes

def parse_single_block(block):
    """解析单个区块内容"""
    result = {
        'title': None,
        'labels': {},   # left -> label_text
        'data': {}       # left -> value
    }

    # 找标题 (class='f-0' 的 span) - 新版 CSS 类名
    all_spans = block.find_all('span', class_=re.compile(r'f-0'))
    if all_spans:
        # 取第一个 b-0 span 的文本作为标题
        result['title'] = all_spans[0].get_text().strip()

    # 找所有带 style 属性的 div
    divs = block.find_all('div', style=True)

    for div in divs:
        style = div.get('style', '')
        left_match = re.search(r'left:(\d+)px', style)
        top_match = re.search(r'top:(\d+)px', style)

        if not left_match or not top_match:
            continue

        left = int(left_match.group(1))
        top = int(top_match.group(1))

        div_id = div.get('id', '')

        # 获取 div 内所有 span 的文本和类
        texts, classes = get_span_text_and_class(div)
        if not texts:
            continue

        text = texts[0]  # 取第一个非空文本

        if not text or text == '&nbsp;' or text.startswith('.'):
            continue

        # 跳过无 id 的 div（通常是容器）
        if not div_id:
            continue

        # 检查是否是标签 (有 f-2 或 f-3 class) - 新版 CSS 类名
        is_label = any('f-2' in c or 'f-3' in c for c in classes)
        is_title = any('f-0' in c for c in classes)

        # 标签行判断：id 以 Text 开头，且有 b-2 或 b-3 class
        if div_id.startswith('Text') and is_label and not is_title:
            result['labels'][left] = text
        # 数据行判断：id 不以 Text 开头
        elif not div_id.startswith('Text'):
            result['data'][left] = text

    return result

def html_to_excel(html_file, excel_file):
    """将 Crystal Reports HTML 转换为 Excel"""
    with open(html_file, 'r', encoding='utf-8') as f:
        html_content = f.read()

    # 解析区块
    blocks = parse_subreports(html_content)
    print(f"解析到 {len(blocks)} 个区块:")

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "报表数据"

    # 样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    row = 1
    headers = ["区块", "字段", "值"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40

    row = 2

    # 定义区块顺序
    block_order = {
        '设备档案': 1,
        '库存条码明细': 2,
        '装箱单': 3,
        '库存调整单': 4,
        '调拨单': 5,
        '移库单': 6,
        '服务单': 7,
        '保卡扫描': 8,
        '押金返还': 9,
        '移机单': 10,
    }

    # 按指定顺序排序
    blocks.sort(key=lambda b: block_order.get(b['title'], 99))

    # 处理每个区块
    for block in blocks:
        title = block['title']
        labels = block['labels']
        data = block['data']

        print(f"  {title}: {len(labels)} 标签, {len(data)} 数据")

        # 合并所有 left 位置
        all_lefts = sorted(set(labels.keys()) | set(data.keys()))

        for left in all_lefts:
            label = labels.get(left, '')
            value = data.get(left, '')

            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=value)
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3).border = thin_border
            row += 1

        # 区块之间空一行
        row += 1

    wb.save(excel_file)
    print(f"\n已保存到: {excel_file}")

if __name__ == "__main__":
    import sys
    html_file = sys.argv[1] if len(sys.argv) > 1 else "debug_372509240205_iframe.html"
    excel_file = html_file.replace('.html', '_converted.xlsx')
    html_to_excel(html_file, excel_file)