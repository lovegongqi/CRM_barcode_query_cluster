#!/usr/bin/env python3
"""
怡口 CRM 条码批量查询脚本
使用 Playwright 自动化浏览器操作
"""

import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, Browser, Page, BrowserContext


# ============ 配置区 ============
CONFIG_PATH = "config.json"
ACCOUNTS_FILE = "accounts.json"
# ================================


def print_menu():
    """打印主菜单"""
    print("\n" + "="*50)
    print("  怡口 CRM 条码查询系统")
    print("="*50)
    print("  1. 单个查询")
    print("  2. 批量查询（前台模式）")
    print("  3. 登录账号")
    print("  4. 添加账号")
    print("  5. 后台批量查询（不影响电脑使用）⭐")
    print("  6. 退出脚本")
    print("="*50)


def load_accounts() -> dict:
    """加载账号列表"""
    if os.path.exists(ACCOUNTS_FILE):
        with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_accounts(accounts: dict):
    """保存账号列表"""
    with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)


def add_account_interactive():
    """添加账号"""
    print("\n--- 添加账号 ---")
    username = input("用户名: ").strip()
    if not username:
        print("用户名不能为空")
        return None

    password = input("密码: ").strip()
    if not password:
        print("密码不能为空")
        return None

    accounts = load_accounts()
    accounts[username] = {"password": password}
    save_accounts(accounts)
    print(f"账号 {username} 已保存！")
    return username


class BarcodeQueryApp:
    """条码查询应用"""

    def __init__(self, config_path: str = CONFIG_PATH):
        self.config = self._load_config(config_path)
        self.playwright = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None

        # 确保目录存在
        Path(self.config["session"]["state_path"]).mkdir(parents=True, exist_ok=True)
        Path(self.config["output"]["results_dir"]).mkdir(parents=True, exist_ok=True)

    def _load_config(self, config_path: str) -> dict:
        """加载配置文件"""
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _gracefully_close_chrome(self, session_dir: str):
        """优雅关闭占用 session 目录的 Chrome 进程，确保数据写入磁盘"""
        import subprocess
        import signal

        try:
            result = subprocess.run(["lsof", "-t", "+D", session_dir], capture_output=True, text=True)
            pids = [line.strip() for line in result.stdout.strip().split("\n") if line.strip()]

            if not pids:
                return

            for pid in pids:
                try:
                    pid_int = int(pid)
                    os.kill(pid_int, signal.SIGTERM)
                    print(f"  已发送优雅关闭信号到进程 {pid}，等待数据写入磁盘...")
                except ProcessLookupError:
                    pass
                except Exception:
                    pass

            for _ in range(10):
                time.sleep(0.5)
                still_running = False
                for pid in pids:
                    try:
                        os.kill(int(pid), 0)
                        still_running = True
                    except (ProcessLookupError, ValueError):
                        pass
                if not still_running:
                    break

            result2 = subprocess.run(["lsof", "-t", "+D", session_dir], capture_output=True, text=True)
            remaining_pids = [line.strip() for line in result2.stdout.strip().split("\n") if line.strip()]
            if remaining_pids:
                for pid in remaining_pids:
                    try:
                        subprocess.run(["kill", "-9", pid], capture_output=True)
                        print(f"  进程 {pid} 未响应，已强制终止")
                    except:
                        pass
            else:
                print("  Chrome 进程已正常退出，会话数据已安全保存")

        except Exception as e:
            print(f"  关闭 Chrome 进程时出错: {e}")

    def cleanup_session_lock(self, session_dir: str):
        """清理会话锁文件和残留进程"""
        self._gracefully_close_chrome(session_dir)

        lock_file = os.path.join(session_dir, "SingletonLock")
        if os.path.exists(lock_file):
            os.remove(lock_file)
            print("  已清理会话锁文件")

    def _save_storage_state_backup(self):
        """将当前 cookies 保存为 JSON 备份文件（persistent_context 的补充保险）"""
        if not self.context:
            return
        try:
            state_file = os.path.join(self.config["session"]["state_path"], "storage_state.json")
            self.context.storage_state(path=state_file)
            print(f"  cookies 已备份到: {state_file}")
        except Exception as e:
            print(f"  cookies 备份失败（非致命）: {e}")

    def _save_session(self):
        """保存浏览器会话"""
        self._save_storage_state_backup()

    def _close_browser_gracefully(self):
        """统一的安全关闭浏览器方法：先备份cookies → 关闭context → 等待 → 停止playwright"""
        if self.context:
            try:
                self._save_storage_state_backup()
                self.context.close()
                time.sleep(3)
            except Exception as e:
                print(f"  关闭浏览器上下文出错: {e}")
        if self.playwright:
            try:
                self.playwright.stop()
            except:
                pass

    def init_browser(self):
        """初始化浏览器"""
        self.playwright = sync_playwright().start()

        # 检查是否有保存的会话
        session_dir = self.config["session"]["state_path"]
        if os.path.exists(session_dir):
            print(f"发现保存的会话: {session_dir}")
            self.context = self.playwright.chromium.launch_persistent_context(
                user_data_dir=session_dir,
                headless=self.config["browser"]["headless"],
                viewport=self.config["browser"]["viewport"]
            )
        else:
            print("未发现保存的会话，将打开浏览器进行首次登录...")
            # 由于 Playwright 的 persistent_context 需要指定 user_data_dir
            # 我们改用普通浏览器然后手动保存
            self.browser = self.playwright.chromium.launch(
                headless=self.config["browser"]["headless"]
            )
            self.context = self.browser.contexts[0] if self.context is None else self.context
            self.page = self.context.new_page()
            self.page.set_viewport_size(self.config["browser"]["viewport"])

    def wait_for_manual_login(self) -> bool:
        """等待用户手动登录（首次登录需要验证码）"""
        url = self.config["website"]["url"]
        print(f"\n{'='*60}")
        print("首次登录需要手动操作：")
        print("1. 浏览器已打开，请完成登录流程")
        print("2. 如果需要验证码，请输入手机收到的验证码")
        print("3. 登录成功后，返回此窗口按 Enter 继续...")
        print(f"4. 目标网址: {url}")
        print(f"{'='*60}\n")

        input("登录成功后，按 Enter 继续...")
        return True

    def check_login_status(self) -> bool:
        """检查是否已登录"""
        try:
            # 尝试访问主页检查是否需要登录
            self.page.goto(self.config["website"]["url"], timeout=30000)
            time.sleep(2)

            # 检查是否跳转到登录页
            if "login" in self.page.url.lower() or "登录" in self.page.content():
                return False
            return True
        except Exception as e:
            print(f"检查登录状态失败: {e}")
            return False

    def setup_browser(self):
        """设置浏览器（首次登录或加载已有会话）"""
        self.playwright = sync_playwright().start()

        session_dir = self.config["session"]["state_path"]

        # 确保目录存在
        os.makedirs(session_dir, exist_ok=True)

        # 清理可能存在的锁文件
        lock_file = os.path.join(session_dir, "SingletonLock")
        if os.path.exists(lock_file):
            os.remove(lock_file)
            print("  已清理残留的会话锁文件")

        # 检查是否已有持久化的浏览器数据
        session_files = [f for f in os.listdir(session_dir) if f not in ['SingletonLock', 'lockfile']]
        has_existing_session = len(session_files) > 0

        if has_existing_session:
            print(f"发现保存的会话（{len(session_files)} 个文件），正在加载...")
            try:
                # 使用已存在的 user_data_dir，自动恢复登录状态
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=self.config["browser"]["headless"],
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
                print("  会话加载成功")
            except Exception as e:
                print(f"  加载会话失败: {e}，将创建新会话")
                has_existing_session = False
        else:
            print("未发现保存的会话，将打开浏览器进行首次登录...")

        if not has_existing_session:
            try:
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=self.config["browser"]["headless"],
                    viewport=self.config["browser"]["viewport"]
                )
            except Exception as e:
                print(f"  创建会话失败: {e}")
                raise

        self.page = self.context.pages[0]

        # 打开网站
        self.page.goto(self.config["website"]["url"], timeout=30000)

        # 检查是否已登录
        time.sleep(2)
        if "login" not in self.page.url.lower():
            print("会话有效，已自动登录！")
            return True

        # 需要登录
        print(f"\n{'='*60}")
        print("请在浏览器中完成登录：")
        print("1. 如果需要验证码，请输入手机收到的验证码")
        print("2. 登录成功后，返回此窗口按 Enter 继续...")
        print(f"{'='*60}\n")
        input("登录成功后，按 Enter 继续...")

        # 登录成功后，等待页面完全加载
        time.sleep(3)
        print("会话已保存，下次运行时将自动登录")

        print("会话已保存，下次运行时将自动登录")
        return True

    def extract_report_url(self) -> str:
        """从水晶报表页面提取报表链接"""
        print("\n提取报表 URL...")

        try:
            # 获取当前页面所有链接
            links = self.page.query_selector_all("a[href]")
            for link in links:
                try:
                    href = link.get_attribute("href")
                    text = link.inner_text()
                    # 查找包含 EcoCrystalRp 的链接
                    if href and "EcoCrystalRp" in href:
                        print(f"  找到报表链接: {href}")
                        return href
                    # 查找"查询条码所有信息"的链接
                    if text and "查询条码所有信息" in text:
                        print(f"  找到报表链接: {href}")
                        return href
                except:
                    pass
        except Exception as e:
            print(f"  提取报表链接失败: {e}")

        return ""

    def navigate_to_report(self):
        """导航到报表查询页面"""
        print("\n========== 开始导航 ==========")

        # 等待页面加载
        print("  [等待] 页面加载中...")
        time.sleep(5)
        print(f"  [OK] 当前URL: {self.page.url}")

        # 如果当前不在 CRM 主页，先跳回去
        if "crmportal.ecowaterchina" not in self.page.url.lower():
            print("  [跳转] 当前不在 CRM 主页，正在跳转...")
            self.page.goto(self.config["website"]["url"], timeout=30000)
            time.sleep(3)

        # 1. 点击报表管理
        print("\n[步骤1] 点击'报表管理'")
        报表管理成功 = False

        selectors = ["text=报表管理", "a:has-text('报表管理')"]
        for selector in selectors:
            try:
                self.page.click(selector, timeout=5000)
                print("  [等待] 等待页面加载...")
                time.sleep(3)  # 等待页面加载
                # 验证
                page_text = self.page.inner_text("body")
                if "报表管理" in page_text:
                    print(f"  [成功✓] 点击'报表管理'成功")
                    报表管理成功 = True
                    break
            except Exception as e:
                print(f"  [失败✗] selector={selector}, 错误: {e}")

        if not 报表管理成功:
            print("  [失败✗] 点击'报表管理'失败")
            input("  [手动] 请手动点击后按 Enter...")

        # 2. 点击水晶报表查看
        print("\n[步骤2] 点击'水晶报表查看'")
        水晶报表成功 = False
        selectors = ["text=水晶报表查看", "a:has-text('水晶报表查看')"]
        url_before = self.page.url
        print("  [等待] 等待'水晶报表查看'元素可见...")

        # 等待元素可见，最多等10秒
        for wait_i in range(10):
            try:
                # 检查元素是否可见
                visible = self.page.is_visible("text=水晶报表查看")
                if visible:
                    print(f"  [OK] 元素已可见，开始点击...")
                    break
                else:
                    print(f"  [等待] 第{wait_i+1}次，元素尚未可见，继续等待...")
                    time.sleep(1)
            except:
                print(f"  [等待] 第{wait_i+1}次检查出错...")
                time.sleep(1)

        for selector in selectors:
            try:
                # 使用 force=True 强制点击，或者等待元素可见
                self.page.click(selector, timeout=10000)
                time.sleep(3)  # 等待页面响应
                # 验证
                url_after = self.page.url
                page_text = self.page.inner_text("body")
                if url_after != url_before:
                    print(f"  [成功✓] 点击'水晶报表查看'成功，URL变化: {url_before} -> {url_after}")
                    水晶报表成功 = True
                    break
                elif "水晶报表" in page_text or "Crystal" in page_text:
                    print(f"  [成功✓] 点击'水晶报表查看'成功")
                    水晶报表成功 = True
                    break
            except Exception as e:
                print(f"  [失败✗] selector={selector}, 错误: {e}")

        if not 水晶报表成功:
            print("  [失败✗] 点击'水晶报表查看'失败")
            input("  [手动] 请手动点击后按 Enter...")

        time.sleep(2)

        # 3. 翻到第2页
        print("\n[步骤3] 翻到第2页（查找'查询条码所有信息'）")
        print("  [等待] 等待页面加载完成...")
        time.sleep(3)  # 等待水晶报表页面加载

        翻页成功 = False

        # 检查是否已经在第2页
        try:
            content = self.page.inner_text("body")
            if "查询条码所有信息" in content:
                print(f"  [成功✓] 已在第2页，找到'查询条码所有信息'")
                翻页成功 = True
        except:
            pass

        # 如果不在，尝试翻页
        if not 翻页成功:
            print("  [尝试] 查找页码输入框...")
            try:
                spinbuttons = self.page.query_selector_all("spinbutton, input[type='number']")
                for sb in spinbuttons:
                    try:
                        val = sb.get_attribute("value")
                        if val == "1":
                            sb.fill("2")
                            sb.press("Enter")
                            time.sleep(2)
                            if "查询条码所有信息" in self.page.inner_text("body"):
                                print(f"  [成功✓] 输入页码2成功")
                                翻页成功 = True
                                break
                    except:
                        pass
            except Exception as e:
                print(f"  [失败✗] 页码输入框方式失败: {e}")

        if not 翻页成功:
            print("  [尝试] 查找'下一页'按钮...")
            try:
                next_btn = self.page.query_selector("button:has-text('下一页'), a:has-text('下一页')")
                if next_btn:
                    next_btn.click()
                    time.sleep(2)
                    if "查询条码所有信息" in self.page.inner_text("body"):
                        print(f"  [成功✓] 点击'下一页'成功")
                        翻页成功 = True
            except Exception as e:
                print(f"  [失败✗] '下一页'按钮方式失败: {e}")

        if not 翻页成功:
            print("  [尝试] 查找页码'2'链接...")
            try:
                page2_link = self.page.get_by_text("2", exact=True).first
                if page2_link:
                    page2_link.click()
                    time.sleep(2)
                    if "查询条码所有信息" in self.page.inner_text("body"):
                        print(f"  [成功✓] 点击页码'2'成功")
                        翻页成功 = True
            except Exception as e:
                print(f"  [失败✗] 页码'2'链接方式失败: {e}")

        if not 翻页成功:
            print("  [失败✗] 无法自动翻页")
            input("  [手动] 请手动翻到第2页后按 Enter...")

        # 4. 双击打开报表
        print("\n[步骤4] 双击打开'查询条码所有信息'报表")
        pages_before = len(self.context.pages)
        print(f"  [INFO] 打开前有 {pages_before} 个标签页")

        # 方法1: JS dblclick 事件
        print("  [尝试] JS dblclick...")
        try:
            result = self.page.evaluate("""() => {
                const elements = document.querySelectorAll('*');
                for (const el of elements) {
                    if (el.textContent.trim() === '查询条码所有信息') {
                        const dblclickEvent = new MouseEvent('dblclick', {
                            bubbles: true,
                            cancelable: true,
                            view: window
                        });
                        el.dispatchEvent(dblclickEvent);
                        return true;
                    }
                }
                return false;
            }""")
            if result:
                print("  [触发] JS dblclick 事件已触发，等待并检测...")
                for i in range(8):
                    time.sleep(1)
                    pages_now = len(self.context.pages)
                    if pages_now > pages_before:
                        print(f"  [成功✓] JS dblclick 打开新标签页（现在共{pages_now}个）")
                        break
                    print(f"  [等待] 第{i+1}秒，等待标签页打开...")
            else:
                print("  [失败✗] JS dblclick 未找到元素")
        except Exception as e:
            print(f"  [失败✗] JS dblclick 出错: {e}")

        # 如果 JS 没效果，用 mouse.dblclick
        if len(self.context.pages) == pages_before:
            print("  [尝试] mouse.dblclick...")
            try:
                report_link = self.page.get_by_text("查询条码所有信息", exact=True).first
                if report_link:
                    box = report_link.bounding_box()
                    if box:
                        x = box['x'] + box['width'] / 2
                        y = box['y'] + box['height'] / 2
                        self.page.mouse.dblclick(x, y)
                        print("  [触发] mouse.dblclick 已触发，等待并检测...")
                        for i in range(8):
                            time.sleep(1)
                            pages_now = len(self.context.pages)
                            if pages_now > pages_before:
                                print(f"  [成功✓] mouse.dblclick 打开新标签页（现在共{pages_now}个）")
                                break
                            print(f"  [等待] 第{i+1}秒，等待标签页打开...")
            except Exception as e:
                print(f"  [失败✗] mouse.dblclick 出错: {e}")

        # 检查最终状态
        pages_now = len(self.context.pages)
        if pages_now == pages_before:
            print("  [失败✗] 未能打开新标签页")
            input("  [手动] 请手动双击打开后按 Enter...")
        else:
            print(f"  [成功✓] 报表已在新标签页中打开（现在共{pages_now}个）")

        print("\n========== 导航完成 ==========")

    def _open_report_in_new_tab(self):
        """直接双击打开'查询条码所有信息'报表（不开新标签页时调用）"""
        print("\n[步骤] 双击打开'查询条码所有信息'报表")
        pages_before = len(self.context.pages)
        print(f"  [INFO] 打开前有 {pages_before} 个标签页")

        # 方法1: JS dblclick 事件
        print("  [尝试] JS dblclick...")
        try:
            result = self.page.evaluate("""() => {
                const elements = document.querySelectorAll('*');
                for (const el of elements) {
                    if (el.textContent.trim() === '查询条码所有信息') {
                        const dblclickEvent = new MouseEvent('dblclick', {
                            bubbles: true,
                            cancelable: true,
                            view: window
                        });
                        el.dispatchEvent(dblclickEvent);
                        return true;
                    }
                }
                return false;
            }""")
            if result:
                print("  [触发] JS dblclick 事件已触发，等待并检测...")
                # 增加等待时间，确保新标签页完全打开
                for i in range(5):
                    time.sleep(1)
                    pages_now = len(self.context.pages)
                    if pages_now > pages_before:
                        print(f"  [成功✓] JS dblclick 打开新标签页（现在共{pages_now}个）")
                        return  # 成功了就直接返回
                    print(f"  [等待] 第{i+1}秒，等待标签页打开...")
            else:
                print("  [失败✗] JS dblclick 未找到元素")
        except Exception as e:
            print(f"  [失败✗] JS dblclick 出错: {e}")

        # 如果 JS 没效果，用 mouse.dblclick
        if len(self.context.pages) == pages_before:
            print("  [尝试] mouse.dblclick...")
            try:
                report_link = self.page.get_by_text("查询条码所有信息", exact=True).first
                if report_link:
                    box = report_link.bounding_box()
                    if box:
                        x = box['x'] + box['width'] / 2
                        y = box['y'] + box['height'] / 2
                        self.page.mouse.dblclick(x, y)
                        print("  [触发] mouse.dblclick 已触发，等待并检测...")
                        # 增加等待时间，确保新标签页完全打开
                        for i in range(5):
                            time.sleep(1)
                            pages_now = len(self.context.pages)
                            if pages_now > pages_before:
                                print(f"  [成功✓] mouse.dblclick 打开新标签页（现在共{pages_now}个）")
                                return  # 成功了就直接返回
                            print(f"  [等待] 第{i+1}秒，等待标签页打开...")
            except Exception as e:
                print(f"  [失败✗] mouse.dblclick 出错: {e}")

        # 检查最终状态
        pages_now = len(self.context.pages)
        if pages_now == pages_before:
            print("  [失败✗] 未能打开新标签页")
            input("  [手动] 请手动双击打开后按 Enter...")
        else:
            print(f"  [成功✓] 报表已在新标签页中打开（现在共{pages_now}个）")

    def prepare_next_report(self):
        """准备下一个报表：检查是否已在本页面，直接打开报表"""
        print("\n========== 准备下一个报表 ==========")

        # 检查当前页面URL是否是CRM报表列表页
        try:
            current_url = self.page.url
            if "crmportal.ecowaterchina" in current_url and "/report/reportlist" in current_url:
                print("  [OK] 已在CRM报表列表页，翻到第2页并打开报表")
                time.sleep(2)

                # 翻到第2页
                翻页成功 = False
                try:
                    page2_link = self.page.get_by_text("2", exact=True).first
                    if page2_link:
                        page2_link.click()
                        time.sleep(2)
                        翻页成功 = True
                        print("  [OK] 已翻到第2页")
                except:
                    pass

                if not 翻页成功:
                    try:
                        next_btn = self.page.query_selector("button:has-text('下一页'), a:has-text('下一页')")
                        if next_btn:
                            next_btn.click()
                            time.sleep(2)
                            翻页成功 = True
                            print("  [OK] 已点击下一页")
                    except:
                        pass

                if 翻页成功:
                    # 直接双击打开报表
                    self._open_report_in_new_tab()
                    self.switch_to_report_tab()
                    return
        except Exception as e:
            print(f"  [检查] 页面检查失败: {e}")

        # 如果不在报表列表页，做完整导航
        print("  [提示] 需要完整导航到报表页面")
        self.navigate_to_report()
        self.switch_to_report_tab()

    def switch_to_report_tab(self) -> bool:
        """切换到报表标签页"""
        print("\n========== 切换到报表标签页 ==========")

        # 获取当前所有页面
        pages = self.context.pages
        print(f"  [INFO] 当前共 {len(pages)} 个标签页")

        for i, p in enumerate(pages):
            try:
                url = p.url
                print(f"    标签页{i+1}: {url}")
            except:
                print(f"    标签页{i+1}: (无法获取URL)")

        # 查找包含 EcoCrystalRp 的标签页（找最后一个，即最新的）
        report_page = None
        for i, p in enumerate(pages):
            try:
                if "/EcoCrystalReports/EcoCrystalRp" in p.url:
                    report_page = p
                    print(f"  [找到] 标签页{i+1} 包含报表 URL")
            except:
                pass

        if report_page:
            self.page = report_page
            self.page.bring_to_front()
            print(f"  [切换] 已切换到报表标签页")
            time.sleep(2)
            print("\n========== 切换完成 ==========")
            return True

        # 如果只有1个标签页，说明没打开新标签
        if len(pages) == 1:
            print("  [失败✗] 没有新标签页被打开")
            print("\n========== 切换完成 ==========")
            return False

        # 如果没找到报表页，切换到最后一个标签页
        self.page = pages[-1]
        self.page.bring_to_front()
        print(f"  [切换] 已切换到最后一个标签页")
        time.sleep(2)

        # 检查是否是报表页
        try:
            content = self.page.inner_text("body")
            if "输入 barcode" in content or "确定" in content or "EcoCrystalRp" in self.page.url:
                print("  [成功✓] 当前标签页是报表页")
            else:
                print(f"  [警告] 当前标签页可能不是报表，URL: {self.page.url}")
        except Exception as e:
            print(f"  [错误] 检查标签页内容失败: {e}")

        print("\n========== 切换完成 ==========")
        return True

    def query_barcode(self, barcode: str) -> dict:
        """查询单个条码"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "状态": "待查询"
        }

        try:
            print(f"\n查询条码: {barcode}")

            # 等待输入框出现
            time.sleep(3)

            # 查找条码输入框
            input_box = self.page.query_selector("input[name='CrystalReportViewer1_p0DiscreteValue']")
            if not input_box:
                print("  未找到条码输入框，尝试重新切换到报表标签页...")
                # 尝试重新切换到报表标签页
                switched = self.switch_to_report_tab()
                if switched:
                    # 再次查找输入框
                    time.sleep(2)
                    input_box = self.page.query_selector("input[name='CrystalReportViewer1_p0DiscreteValue']")
                    if input_box:
                        print("  重新切换标签页后找到输入框")
                    else:
                        print("  重新切换标签页后仍未找到输入框，请手动操作...")
                        input("  按 Enter 继续...")
                        result["状态"] = "手动取消"
                        return result
                else:
                    print("  切换标签页失败，请手动操作...")
                    input("  按 Enter 继续...")
                    result["状态"] = "手动取消"
                    return result

            # 等待 Crystal Report 的加载动画（wait01.gif）消失
            for _ in range(30):
                try:
                    loading = self.page.evaluate("""() => {
                        const imgs = document.querySelectorAll('img[src*="wait"]');
                        for (const img of imgs) {
                            if (img.offsetParent !== null) return true;
                        }
                        return false;
                    }""")
                    if not loading:
                        break
                except:
                    break
                time.sleep(1)

            # 输入条码
            input_box.click()
            time.sleep(0.3)
            input_box.press("Control+a")
            time.sleep(0.1)
            input_box.press("Backspace")
            time.sleep(0.1)
            input_box.type(barcode, delay=100)
            print(f"  已输入条码: {barcode}")

            # 点击确定按钮提交查询
            print("  提交查询...")
            确定成功 = False
            try:
                self.page.evaluate("if(typeof CrystalReportViewer1_submit === 'function') { CrystalReportViewer1_submit(); }")
                确定成功 = True
                time.sleep(1)
            except:
                pass

            if not 确定成功:
                try:
                    confirm_button = self.page.get_by_text("确定", exact=True).first
                    if confirm_button:
                        confirm_button.click()
                        确定成功 = True
                except:
                    pass

            if not 确定成功:
                try:
                    links = self.page.query_selector_all("a")
                    for link in links:
                        try:
                            if "确定" in link.inner_text().strip():
                                link.click()
                                确定成功 = True
                                break
                        except:
                            pass
                except:
                    pass

            if not 确定成功:
                print("  提交失败，请手动操作...")
                input("  按 Enter 继续...")
                result["状态"] = "手动取消"
                return result

            # 等待报表处理（智能等待：循环检测直到数据出现）
            print("  等待报表处理...")
            max_wait = 60  # 最多等60秒
            data_ready = False
            for wait_i in range(max_wait):
                time.sleep(1)
                try:
                    # 检查 iframe 中是否有实际数据
                    js_check = self.page.evaluate("""() => {
                        const iframes = document.querySelectorAll('iframe');
                        for (const iframe of iframes) {
                            try {
                                const doc = iframe.contentDocument || iframe.contentWindow?.document;
                                if (doc && doc.body) {
                                    const text = doc.body.innerText || '';
                                    // 排除"正在处理"的加载状态
                                    if (text.length > 500 && !text.includes('正在处理') && !text.includes('请稍候')) {
                                        return { ready: true, length: text.length };
                                    }
                                }
                            } catch(e) {}
                        }
                        return { ready: false, length: 0 };
                    }""")
                    if js_check.get('ready'):
                        data_ready = True
                        print(f"  报表已加载完成（{js_check['length']} 字符）")
                        break
                    elif (wait_i + 1) % 10 == 0:
                        print(f"  ...已等待 {wait_i + 1} 秒，继续等待...")
                except:
                    pass

            if not data_ready:
                print("  ⚠️ 等待超时，尝试提取可能的数据...")

            # 提取报表数据
            print("  提取报表数据...")
            result = self.extract_report_data(barcode)
            result["状态"] = "成功"

        except Exception as e:
            print(f"  查询失败: {e}")
            import traceback
            traceback.print_exc()
            result["状态"] = f"失败: {str(e)}"

        return result

    def _export_report(self, barcode: str) -> dict:
        """导出报表文件"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "状态": "待导出"
        }

        try:
            print(f"\n  开始导出报表...")

            # 查找导出按钮
            # Crystal Report 的导出按钮通常在 toolbar 上
            export_button = None
            print("  调试 - 查找导出按钮:")

            # 先打印所有可见图片，了解 toolbar 结构
            try:
                all_imgs = self.page.query_selector_all("img")
                print(f"    页面共有 {len(all_imgs)} 个图片")
                for i, img in enumerate(all_imgs[:20]):  # 只打印前20个
                    try:
                        if img.is_visible():
                            src = img.get_attribute("src") or ""
                            alt = img.get_attribute("alt") or ""
                            title = img.get_attribute("title") or ""
                            # 只打印 toolbar 相关的图片
                            if 'crystalreport' in src.lower() or 'toolbar' in src.lower() or 'dhtmllib' in src.lower():
                                print(f"    img[{i}]: src='{src[:60]}' alt='{alt}' title='{title}'")
                    except:
                        pass
            except Exception as e:
                print(f"    打印图片失败: {e}")

            try:
                # 方法1: 通过 alt 文本找到导出按钮并点击
                export_result = self.page.evaluate("""() => {
                    // 查找 alt='导出此报表' 的 img 元素
                    const imgs = document.querySelectorAll('img');
                    for (const img of imgs) {
                        if (img.getAttribute('alt') === '导出此报表') {
                            img.click();
                            return 'clicked img[alt=导出此报表]';
                        }
                    }
                    return null;
                }""")
                if export_result:
                    print(f"    {export_result}")
                    export_button = True

                # 方法2: 如果方法1失败，尝试 onclick 包含 exportto
                if not export_button:
                    export_elements = self.page.evaluate("""() => {
                        const results = [];
                        const allEls = document.querySelectorAll('*');
                        for (const el of allEls) {
                            const onclick = el.getAttribute('onclick') || '';
                            const title = el.getAttribute('title') || '';
                            if (onclick.toLowerCase().includes('exportto') || title.toLowerCase().includes('export')) {
                                results.push({
                                    tag: el.tagName,
                                    onclick: onclick.substring(0, 80),
                                    title: title
                                });
                            }
                        }
                        return results;
                    }""")
                    if export_elements:
                        print(f"    找到 {len(export_elements)} 个导出相关元素:")
                        for el in export_elements[:5]:
                            print(f"      {el}")
                        export_button = True

                # 方法3: 调用 CrystalReportViewer 的导出方法
                if not export_button:
                    toolbar_result = self.page.evaluate("""() => {
                        if (typeof CrystalReportViewer1 !== 'undefined') {
                            if (typeof CrystalReportViewer1.exportTo === 'function') {
                                CrystalReportViewer1.exportTo(1);
                                return 'exportTo(1) called';
                            }
                        }
                        return null;
                    }""")
                    if toolbar_result:
                        print(f"    CrystalReportViewer: {toolbar_result}")
                        export_button = True

            except Exception as e:
                print(f"    查找导出按钮失败: {e}")

            # 点击导出按钮
            print("  点击导出按钮...")
            export_clicked = False
            try:
                # 方法1: 直接通过 alt 文本找到按钮并点击
                export_click_result = self.page.evaluate("""() => {
                    const imgs = document.querySelectorAll('img');
                    for (const img of imgs) {
                        if (img.getAttribute('alt') === '导出此报表') {
                            img.click();
                            return 'clicked img[alt=导出此报表]';
                        }
                    }
                    return null;
                }""")
                if export_click_result:
                    print(f"    {export_click_result}")
                    export_clicked = True

                # 方法2: 如果方法1失败，尝试查找父元素的 onclick
                if not export_clicked:
                    export_click_result = self.page.evaluate("""() => {
                        // 查找 img 元素然后找其父元素的 onclick
                        const imgs = document.querySelectorAll('img');
                        for (const img of imgs) {
                            if (img.getAttribute('alt') === '导出此报表') {
                                // 尝试从父元素开始向上查找有 onclick 的
                                let parent = img.parentElement;
                                while (parent) {
                                    const onclick = parent.getAttribute('onclick');
                                    if (onclick) {
                                        parent.click();
                                        return 'clicked parent: ' + parent.tagName;
                                    }
                                    parent = parent.parentElement;
                                }
                                // 尝试直接调用 CrystalReportViewer 的方法
                                if (typeof CrystalReportViewer1 !== 'undefined') {
                                    if (typeof CrystalReportViewer1.exportTo === 'function') {
                                        CrystalReportViewer1.exportTo(1);
                                        return 'exportTo(1)';
                                    }
                                }
                            }
                        }
                        return null;
                    }""")
                    if export_click_result:
                        print(f"    {export_click_result}")
                        export_clicked = True

            except Exception as e:
                print(f"    点击导出按钮失败: {e}")

            # 等待导出对话框出现
            print("  等待导出对话框...")
            time.sleep(3)

            # 查找并操作对话框
            export_success = False
            try:
                print("    调试 - 选择导出格式...")

                # 方法: 直接调用 ExportUI 的 setExportFormat 方法
                select_result = self.page.evaluate("""() => {
                    // 从页面源码可知 bobj.crv.ExportUI 有 setExportFormat 方法
                    // 格式值: 'RecordToMSExcel' 对应 "Microsoft Excel (97-2003) 仅限数据"

                    // 尝试通过 window 对象找到 ExportUI 实例
                    try {
                        // bobj 对象在全局作用域
                        if (typeof(bobj) !== 'undefined' && bobj.crv) {
                            // 尝试找到 ExportUI widget
                            // 从页面元素中找到 viewer
                            const viewerEl = document.querySelector('[id*="CrystalReportViewer"]');
                            if (viewerEl && viewerEl.widget) {
                                const viewer = viewerEl.widget;
                                if (viewer.exportUI) {
                                    viewer.exportUI.setExportFormat('RecordToMSExcel');
                                    return '方法1成功: viewer.exportUI.setExportFormat';
                                }
                            }
                        }
                    } catch(e) {
                        // 忽略错误
                    }

                    // 尝试通过其他方式
                    try {
                        // 直接调用 CrystalReportViewer1 的方法（如果存在）
                        if (typeof(CrystalReportViewer1) !== 'undefined') {
                            // 查找_export 或类似方法
                            for (const key in CrystalReportViewer1) {
                                if (key.toLowerCase().includes('export')) {
                                    return '找到: ' + key;
                                }
                            }
                        }
                    } catch(e) {}

                    // 最后尝试：手动设置 select 的值
                    const selects = document.querySelectorAll('select');
                    for (const sel of selects) {
                        for (let i = 0; i < sel.options.length; i++) {
                            if (sel.options[i].text.includes('仅限数据')) {
                                sel.selectedIndex = i;
                                sel.dispatchEvent(new Event('change', { bubbles: true }));
                                return '方法2成功: 设置 select 为 ' + sel.options[i].text;
                            }
                        }
                    }

                    return null;
                }""")
                print(f"    {select_result}")

                if select_result and select_result.startswith(('方法1', '方法2')):
                    export_success = True
                else:
                    # 如果上面的方法都不行，让用户手动选择
                    print("    请手动选择格式: Microsoft Excel (97-2003) 仅限数据")

            except Exception as e:
                print(f"    选择格式失败: {e}")

            # 点击确定按钮下载
            if export_success:
                print("  点击确定按钮...")
                time.sleep(1)
                try:
                    # 查找确定按钮
                    confirm_btns = self.page.query_selector_all("button, input[type='button']")
                    for btn in confirm_btns:
                        try:
                            if btn.is_visible():
                                text = btn.inner_text().strip()
                                if "确" in text or "定" in text:
                                    btn.click()
                                    print(f"    已点击: {text}")
                                    break
                        except:
                            pass
                except Exception as e:
                    print(f"    点击确定失败: {e}")

                # 等待下载
                print("  等待文件下载...")
                time.sleep(5)

            # 如果没找到对话框，让用户手动操作
            if not export_success:
                print("\n  请手动操作导出：")
                print("  1. 点击导出按钮")
                print("  2. 在下拉菜单选择: Microsoft Excel (97-2003) 仅限数据")
                print("  3. 点击确定下载")
                print(f"  4. 保存文件名为: {barcode}.xls")
                input("  下载完成后按 Enter 继续...")
            else:
                print(f"\n  导出完成，文件应保存为: {barcode}.xls")

            result["状态"] = "导出完成"

        except Exception as e:
            print(f"  导出失败: {e}")
            result["状态"] = f"导出失败: {str(e)}"

        return result

    def extract_report_data(self, barcode: str) -> dict:
        """从报表页面提取数据 - 保存 HTML 到文件"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        try:
            # 获取 iframe 的完整 innerHTML
            js_result = self.page.evaluate("""() => {
                const iframes = document.querySelectorAll('iframe');
                let html = '';
                for (const iframe of iframes) {
                    try {
                        const iframeDoc = iframe.contentDocument || iframe.contentWindow?.document;
                        if (iframeDoc && iframeDoc.body) {
                            html += iframeDoc.body.innerHTML;
                        }
                    } catch (e) {}
                }
                return html;
            }""")

            # 检查内容是否有效（排除空白或无效页面）
            if js_result and len(js_result.strip()) > 1000:
                # 保存 HTML 文件到 barcode 目录
                html_dir = os.path.join(os.path.dirname(__file__), "barcode")
                Path(html_dir).mkdir(exist_ok=True)
                output_file = os.path.join(html_dir, f"{barcode}.html")
                # 添加返回结果页面的按钮
                back_button = '''<div style="position:fixed;top:10px;right:10px;z-index:9999;"><a href="http://us.mlmll.cn:5001" style="display:inline-block;padding:10px 20px;background:#4472C4;color:#fff;text-decoration:none;border-radius:4px;font-size:14px;">← 返回结果列表</a></div>'''
                wrapped_html = js_result + back_button
                with open(output_file, "w", encoding="utf-8") as f:
                    f.write(wrapped_html)
                print(f"\n  HTML 已保存到: {output_file} ({len(wrapped_html)} 字符)")
                result["状态"] = "成功"
            else:
                print("  查询结果为空，请核对条码是否正确")
                result["状态"] = "查询结果为空，请核对条码"

        except Exception as e:
            print(f"  数据提取失败: {e}")
            import traceback
            traceback.print_exc()
            result["状态"] = f"提取失败: {str(e)}"

        return result

    def _parse_html_with_colons(self, barcode: str, html_content: str) -> dict:
        """从 HTML 内容中解析带冒号的字段"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        import re

        # 区块标题
        block_titles = [
            "设备档案", "库存条码明细", "装箱单", "库存调整单",
            "调拨单", "移库单", "服务单", "保卡扫描", "押金返还", "移机单"
        ]

        # 字段名映射
        field_map = {
            # 设备档案
            "条码": "设备档案_条码",
            "物料编码": "设备档案_物料编码",
            "物料描述": "设备档案_物料描述",
            "代码": "设备档案_代码",
            "所属经销商": "设备档案_所属经销商",
            "客户": "设备档案_客户",
            "电话": "设备档案_电话",
            "地址": "设备档案_地址",
            "安装日期": "设备档案_安装日期",
            "安装": "设备档案_安装",
            "返修": "设备档案_返修",
            "押金经销商": "设备档案_押金经销商",
            "押金": "设备档案_押金",
            "发货日期": "设备档案_发货日期",
            "线上产品": "设备档案_线上产品",
            # 库存条码明细
            "产品条码": "库存条码明细_产品条码",
            "产品编码": "库存条码明细_产品编码",
            "产品名称": "库存条码明细_产品名称",
            "产品状态": "库存条码明细_产品状态",
            "是否可用": "库存条码明细_是否可用",
            # 装箱单
            "装箱单号": "装箱单_装箱单号",
            "发货单号": "装箱单_发货单号",
            "订单号": "装箱单_订单号",
            "物料名称": "装箱单_物料名称",
            "装箱单状态": "装箱单_状态",
            "类型": "装箱单_类型",
            "经销商": "装箱单_经销商",
            "订单经销商": "装箱单_订单经销商",
            "押金": "装箱单_押金",
            # 库存调整单
            "库存调整单号": "库存调整单_单号",
            "调整日期": "库存调整单_日期",
            "交易方向": "库存调整单_交易方向",
            # 调拨单
            "调拨单号": "调拨单_单号",
            "调入代码": "调拨单_调入代码",
            "调入经销商名称": "调拨单_调入经销商",
            "调出代码": "调拨单_调出代码",
            "调出经销商名称": "调拨单_调出经销商",
            # 移库单
            "移库单号": "移库单_单号",
            "移库日期": "移库单_日期",
            "移库类型": "移库单_类型",
            "经销商代码": "移库单_经销商代码",
            "经销商名称": "移库单_经销商名称",
            "分销商代码": "移库单_分销商代码",
            "分销商名称": "移库单_分销商名称",
            # 服务单
            "服务单号": "服务单_单号",
            "服务类型": "服务单_类型",
            "售前经销商": "服务单_售前经销商",
            "服务经销商": "服务单_服务经销商",
            "服务站": "服务单_服务站",
            "物料描述": "服务单_物料描述",
            "服务状态": "服务单_状态",
            "线下带货上门": "服务单_线下带货",
            "是否结单": "服务单_是否结单",
            # 保卡扫描
            "扫描时间": "保卡扫描_时间",
            # 押金返还
            "返还日期": "押金返还_日期",
            "押金单号": "押金返还_单号",
            "押金经销商": "押金返还_押金经销商",
            "服务单日期": "押金返还_服务单日期",
            "服务经销商": "押金返还_服务经销商",
            "押金金额": "押金返还_金额",
            "物料": "押金返还_物料",
            # 移机单
            "移机单号": "移机单_单号",
            "移机日期": "移机单_日期",
            "设备号": "移机单_设备号",
            "原联系人": "移机单_原联系人",
            "原联系电话": "移机单_原联系电话",
            "新联系人": "移机单_新联系人",
            "新联系电话": "移机单_新联系电话",
            "新联系地址": "移机单_新联系地址",
            "关联经销商": "移机单_关联经销商",
            "原经销商": "移机单_原经销商",
            "原联系地址": "移机单_原联系地址",
        }

        # 解析 "字段名: 值" 格式
        found_count = 0
        lines = html_content.split('\\n')
        current_block = None

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # 检查是否是区块标题
            for title in block_titles:
                if title in line and len(line) < 20:
                    current_block = title
                    break

            # 查找 "字段名: 值" 格式
            if ':' in line:
                parts = line.split(':', 1)
                if len(parts) == 2:
                    key = parts[0].strip()
                    value = parts[1].strip()

                    # 清理 HTML 标签
                    import re
                    key = re.sub(r'<[^>]+>', '', key)
                    value = re.sub(r'<[^>]+>', '', value)
                    key = key.strip()
                    value = value.strip()

                    # 跳过空值和无关内容
                    if not key or not value:
                        continue
                    skip_values = ['¥', '.00', '主报表', '组树', '正在处理', '条码详细信息', 'pePromptFieldset']
                    if any(v in value for v in skip_values) or any(v in key for v in skip_values):
                        continue

                    # 匹配字段名
                    if key in field_map:
                        result_key = field_map[key]
                        if result.get(result_key, "（空）") == "（空）":
                            result[result_key] = value
                            found_count += 1
                            print(f"    [{current_block}] {key} = {value}")

        print(f"\n  HTML 解析找到 {found_count} 个字段")
        return result

    def _parse_block_structure(self, barcode: str, lines: list) -> dict:
        """按区块结构解析：字段名和值分开，先列字段名再列值"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # 区块标题
        block_titles = [
            "设备档案", "库存条码明细", "装箱单", "库存调整单",
            "调拨单", "移库单", "服务单", "保卡扫描", "押金返还", "移机单"
        ]

        # 全部字段名（用于识别）
        all_field_names = [
            # 设备档案
            "条码", "物料编码", "物料描述", "代码", "所属经销商", "客户",
            "电话", "地址", "安装日期", "安装", "返修", "押金经销商",
            "押金", "发货日期", "线上产品",
            # 库存条码明细
            "产品条码", "产品编码", "产品名称", "产品状态", "是否可用",
            # 装箱单
            "装箱单号", "发货单号", "订单号", "物料名称", "装箱单状态",
            "类型", "经销商", "订单经销商",
            # 库存调整单
            "库存调整单号", "调整日期", "交易方向",
            # 调拨单
            "调拨单号", "调入代码", "调入经销商名称", "调出代码", "调出经销商名称",
            # 移库单
            "移库单号", "移库日期", "移库类型", "经销商代码", "经销商名称",
            "分销商代码", "分销商名称",
            # 服务单
            "服务单号", "服务类型", "售前经销商", "服务经销商", "服务站",
            "物料描述", "服务状态", "线下带货上门", "是否结单",
            # 保卡扫描
            "扫描时间",
            # 押金返还
            "返还日期", "押金单号", "押金经销商", "服务单日期", "服务经销商", "押金金额", "物料",
            # 移机单
            "移机单号", "移机日期", "设备号", "原联系人", "原联系电话",
            "新联系人", "新联系电话", "新联系地址", "关联经销商", "原经销商", "原联系地址",
        ]

        # 字段名 -> 结果键
        field_map = {
            # 设备档案
            "条码": "设备档案_条码",
            "物料编码": "设备档案_物料编码",
            "物料描述": "设备档案_物料描述",
            "代码": "设备档案_代码",
            "所属经销商": "设备档案_所属经销商",
            "客户": "设备档案_客户",
            "电话": "设备档案_电话",
            "地址": "设备档案_地址",
            "安装日期": "设备档案_安装日期",
            "安装": "设备档案_安装",
            "返修": "设备档案_返修",
            "押金经销商": "设备档案_押金经销商",
            "押金": "设备档案_押金",
            "发货日期": "设备档案_发货日期",
            "线上产品": "设备档案_线上产品",
            # 库存条码明细
            "产品条码": "库存条码明细_产品条码",
            "产品编码": "库存条码明细_产品编码",
            "产品名称": "库存条码明细_产品名称",
            "产品状态": "库存条码明细_产品状态",
            "是否可用": "库存条码明细_是否可用",
            # 装箱单
            "装箱单号": "装箱单_装箱单号",
            "发货单号": "装箱单_发货单号",
            "订单号": "装箱单_订单号",
            "物料名称": "装箱单_物料名称",
            "装箱单状态": "装箱单_状态",
            "类型": "装箱单_类型",
            "经销商": "装箱单_经销商",
            "订单经销商": "装箱单_订单经销商",
            # 库存调整单
            "库存调整单号": "库存调整单_单号",
            "调整日期": "库存调整单_日期",
            "交易方向": "库存调整单_交易方向",
            # 调拨单
            "调拨单号": "调拨单_单号",
            "调入代码": "调拨单_调入代码",
            "调入经销商名称": "调拨单_调入经销商",
            "调出代码": "调拨单_调出代码",
            "调出经销商名称": "调拨单_调出经销商",
            # 移库单
            "移库单号": "移库单_单号",
            "移库日期": "移库单_日期",
            "移库类型": "移库单_类型",
            "经销商代码": "移库单_经销商代码",
            "经销商名称": "移库单_经销商名称",
            "分销商代码": "移库单_分销商代码",
            "分销商名称": "移库单_分销商名称",
            # 服务单
            "服务单号": "服务单_单号",
            "服务类型": "服务单_类型",
            "售前经销商": "服务单_售前经销商",
            "服务经销商": "服务单_服务经销商",
            "服务站": "服务单_服务站",
            "物料描述": "服务单_物料描述",
            "服务状态": "服务单_状态",
            "线下带货上门": "服务单_线下带货",
            "是否结单": "服务单_是否结单",
            # 保卡扫描
            "扫描时间": "保卡扫描_时间",
            # 押金返还
            "返还日期": "押金返还_日期",
            "押金单号": "押金返还_单号",
            "押金经销商": "押金返还_押金经销商",
            "服务单日期": "押金返还_服务单日期",
            "服务经销商": "押金返还_服务经销商",
            "押金金额": "押金返还_金额",
            "物料": "押金返还_物料",
            # 移机单
            "移机单号": "移机单_单号",
            "移机日期": "移机单_日期",
            "设备号": "移机单_设备号",
            "原联系人": "移机单_原联系人",
            "原联系电话": "移机单_原联系电话",
            "新联系人": "移机单_新联系人",
            "新联系电话": "移机单_新联系电话",
            "新联系地址": "移机单_新联系地址",
            "关联经销商": "移机单_关联经销商",
            "原经销商": "移机单_原经销商",
            "原联系地址": "移机单_原联系地址",
        }

        # 跳过无关关键词
        skip_keywords = ['主报表', '组树', '正在处理', '¥', '.00', '条码详细信息']

        current_block = None
        current_field_names = []  # 当前区块收集的字段名
        current_values = []        # 当前区块收集的值
        phase = "fields"  # "fields" = 收集字段名阶段, "values" = 收集值阶段
        found_count = 0

        # 调试：打印所有实际处理的行
        print(f"\n  === 实际处理 {len(lines)} 行 ===")
        for i, line in enumerate(lines[:100]):
            print(f"    [{i}] {line}")
        if len(lines) > 100:
            print(f"    ... 还有 {len(lines)-100} 行")
        print(f"  ============================")

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # 跳过空行和无关内容
            if not line or any(kw in line for kw in skip_keywords):
                i += 1
                continue

            # 检查是否是区块标题
            is_block_title = False
            for title in block_titles:
                if title in line and len(line) < 15:
                    # 先处理上一个区块的值
                    if current_block and current_field_names and current_values:
                        # 对应填充：字段名和值按顺序对应
                        for fi, fname in enumerate(current_field_names):
                            if fname in field_map and fi < len(current_values):
                                result_key = field_map[fname]
                                if result.get(result_key, "（空）") == "（空）":
                                    result[result_key] = current_values[fi]
                                    found_count += 1
                                    print(f"    [{current_block}] {fname} = {current_values[fi]}")

                    # 开始新区块
                    current_block = title
                    current_field_names = []
                    current_values = []
                    phase = "fields"
                    print(f"  [区块] {current_block}")
                    is_block_title = True

            if is_block_title:
                i += 1
                continue

            # 判断当前是字段名还是值
            if current_block:
                if line in all_field_names:
                    # 是已知字段名，添加到字段名列表
                    current_field_names.append(line)
                    phase = "fields"
                else:
                    # 不是已知字段名，应该是值
                    if phase == "fields" and current_field_names:
                        # 开始进入值阶段
                        phase = "values"
                    if phase == "values" or (not current_field_names and line):
                        current_values.append(line)

            i += 1

        # 处理最后一个区块
        if current_block and current_field_names and current_values:
            for fi, fname in enumerate(current_field_names):
                if fname in field_map and fi < len(current_values):
                    result_key = field_map[fname]
                    if result.get(result_key, "（空）") == "（空）":
                        result[result_key] = current_values[fi]
                        found_count += 1
                        print(f"    [{current_block}] {fname} = {current_values[fi]}")

        print(f"\n  解析完成，找到 {found_count} 个字段")
        return result

    def _parse_html_tables(self, barcode: str, tables: list) -> dict:
        """解析 HTML 表格结构提取字段"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # 字段映射：报表字段名 -> 结果键
        field_map = {
            # 设备档案
            "条码": "设备档案_条码",
            "物料编码": "设备档案_物料编码",
            "物料描述": "设备档案_物料描述",
            "代码": "设备档案_代码",
            "所属经销商": "设备档案_所属经销商",
            "客户": "设备档案_客户",
            "电话": "设备档案_电话",
            "地址": "设备档案_地址",
            "安装日期": "设备档案_安装日期",
            "安装": "设备档案_安装",
            "返修": "设备档案_返修",
            "押金经销商": "设备档案_押金经销商",
            "押金": "设备档案_押金",
            "发货日期": "设备档案_发货日期",
            "线上产品": "设备档案_线上产品",
            # 库存条码明细
            "产品条码": "库存条码明细_产品条码",
            "产品编码": "库存条码明细_产品编码",
            "产品名称": "库存条码明细_产品名称",
            "产品状态": "库存条码明细_产品状态",
            "是否可用": "库存条码明细_是否可用",
            # 装箱单
            "装箱单号": "装箱单_装箱单号",
            "发货单号": "装箱单_发货单号",
            "订单号": "装箱单_订单号",
            "物料名称": "装箱单_物料名称",
            "装箱单状态": "装箱单_状态",
            "类型": "装箱单_类型",
            "经销商": "装箱单_经销商",
            "订单经销商": "装箱单_订单经销商",
            # 移库单
            "移库单号": "移库单_单号",
            "移库日期": "移库单_日期",
            "移库类型": "移库单_类型",
            "经销商代码": "移库单_经销商代码",
            "经销商名称": "移库单_经销商名称",
            "分销商代码": "移库单_分销商代码",
            "分销商名称": "移库单_分销商名称",
            # 服务单
            "服务单号": "服务单_单号",
            "服务类型": "服务单_类型",
            "售前经销商": "服务单_售前经销商",
            "服务经销商": "服务单_服务经销商",
            "服务站": "服务单_服务站",
            "物料描述": "服务单_物料描述",
            "服务状态": "服务单_状态",
            "线下带货上门": "服务单_线下带货",
            "是否结单": "服务单_是否结单",
            # 库存调整单
            "库存调整单号": "库存调整单_单号",
            "调整日期": "库存调整单_日期",
            "交易方向": "库存调整单_交易方向",
            # 调拨单
            "调拨单号": "调拨单_单号",
            "调入代码": "调拨单_调入代码",
            "调入经销商名称": "调拨单_调入经销商",
            "调出代码": "调拨单_调出代码",
            "调出经销商名称": "调拨单_调出经销商",
            # 保卡扫描
            "扫描时间": "保卡扫描_时间",
            # 押金返还
            "返还日期": "押金返还_日期",
            "押金单号": "押金返还_单号",
            "押金经销商": "押金返还_押金经销商",
            "服务单日期": "押金返还_服务单日期",
            "服务经销商": "押金返还_服务经销商",
            "押金金额": "押金返还_金额",
            "物料": "押金返还_物料",
        }

        found_count = 0
        debug_printed = set()  # 避免重复打印
        
        for tbl in tables:
            headers = tbl.get('headers', [])
            rows = tbl.get('rows', [])
            
            # 调试：打印前5个有内容的表格原始数据
            tbl_idx = tbl.get('index', -1)
            if tbl_idx not in debug_printed and len(rows) > 0:
                debug_printed.add(tbl_idx)
                print(f"\n  调试表格 index={tbl_idx}: headers={headers[:5]}, 前3行={rows[:3]}")

            # 调试每个表格的原始 HTML
            if tbl_idx < 3 and len(rows) > 0:
                raw_html = self.page.evaluate(("""
                    () => {
                        const iframes = document.querySelectorAll('iframe');
                        for (const iframe of iframes) {
                            try {
                                const iframeDoc = iframe.contentDocument || iframe.contentWindow?.document;
                                if (iframeDoc) {
                                    const tables = iframeDoc.querySelectorAll('table');
                                    if (tables[%d]) {
                                        return tables[%d].outerHTML.substring(0, 2000);
                                    }
                                }
                            } catch(e) {}
                        }
                        return 'not found';
                    }
                """ % (tbl_idx, tbl_idx)))
                print(f"    原始HTML[{tbl_idx}]: {raw_html}")
            
            # 方式1：表头 + 数据行配对
            if headers:
                print(f"    表头: {headers[:10]}")
                # 遍历每一行
                for row_idx, row in enumerate(rows):
                    # 跳过列标题行（如果首行像表头）
                    if row_idx == 0 and len(row) == len(headers) and row == headers:
                        continue
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            # 精确匹配字段名
                            if header in field_map:
                                result_key = field_map[header]
                                if result.get(result_key, "（空）") == "（空）":
                                    result[result_key] = cell
                                    found_count += 1
                                    print(f"    提取 [{row_idx}][{col_idx}] {header} = {cell}")
                        # 也可以检查 cell 是否是字段名，下一个 cell 是值
                        if cell in field_map and col_idx + 1 < len(row):
                            result_key = field_map[cell]
                            if result.get(result_key, "（空）") == "（空）":
                                result[result_key] = row[col_idx + 1]
                                found_count += 1
                                print(f"    提取 [{row_idx}][{col_idx}] {cell} -> {row[col_idx + 1]}")

            # 方式2：纯数据行，两列且第一列像字段名
            for row in rows:
                if len(row) >= 2:
                    col0, col1 = row[0].strip(), row[1].strip()
                    # 检查第一列是否是字段名
                    if col0 in field_map:
                        result_key = field_map[col0]
                        if result.get(result_key, "（空）") == "（空）":
                            result[result_key] = col1
                            found_count += 1
                            print(f"    提取(2列) {col0} = {col1}")

        print(f"\n  HTML 表格解析完成，找到 {found_count} 个字段")
        return result

    def _parse_text_content(self, barcode: str, lines: list) -> dict:
        """解析文本内容提取字段 - 根据用户提供的字段列表（回退方案）"""
        result = {
            "条码": barcode,
            "查询时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # 用户提供的字段列表（按区块组织）
        sections = [
            # 一、设备档案
            ("条码", "设备档案_条码"),
            ("物料编码", "设备档案_物料编码"),
            ("物料描述", "设备档案_物料描述"),
            ("代码", "设备档案_代码"),
            ("所属经销商", "设备档案_所属经销商"),
            ("客户", "设备档案_客户"),
            ("电话", "设备档案_电话"),
            ("地址", "设备档案_地址"),
            ("安装日期", "设备档案_安装日期"),
            ("安装", "设备档案_安装"),
            ("返修", "设备档案_返修"),
            ("押金经销商", "设备档案_押金经销商"),
            ("押金", "设备档案_押金"),
            ("发货日期", "设备档案_发货日期"),
            ("线上产品", "设备档案_线上产品"),
            # 二、库存条码明细
            ("产品条码", "库存条码明细_产品条码"),
            ("产品编码", "库存条码明细_产品编码"),
            ("产品名称", "库存条码明细_产品名称"),
            ("产品状态", "库存条码明细_产品状态"),
            ("是否可用", "库存条码明细_是否可用"),
            # 三、装箱单
            ("装箱单号", "装箱单_装箱单号"),
            ("发货单号", "装箱单_发货单号"),
            ("订单号", "装箱单_订单号"),
            ("物料名称", "装箱单_物料名称"),
            ("装箱单状态", "装箱单_状态"),
            ("类型", "装箱单_类型"),
            ("经销商", "装箱单_经销商"),
            ("订单经销商", "装箱单_订单经销商"),
            ("押金", "装箱单_押金"),
            # 四、库存调整单
            ("库存调整单号", "库存调整单_单号"),
            ("调整日期", "库存调整单_日期"),
            ("交易方向", "库存调整单_交易方向"),
            # 五、调拨单
            ("调拨单号", "调拨单_单号"),
            ("调入代码", "调拨单_调入代码"),
            ("调入经销商名称", "调拨单_调入经销商"),
            ("调出代码", "调拨单_调出代码"),
            ("调出经销商名称", "调拨单_调出经销商"),
            # 六、移库单
            ("移库单号", "移库单_单号"),
            ("移库日期", "移库单_日期"),
            ("移库类型", "移库单_类型"),
            ("经销商代码", "移库单_经销商代码"),
            ("经销商名称", "移库单_经销商名称"),
            ("分销商代码", "移库单_分销商代码"),
            ("分销商名称", "移库单_分销商名称"),
            # 七、服务单
            ("服务单号", "服务单_单号"),
            ("服务类型", "服务单_类型"),
            ("售前经销商", "服务单_售前经销商"),
            ("服务经销商", "服务单_服务经销商"),
            ("服务站", "服务单_服务站"),
            ("物料编码", "服务单_物料编码"),
            ("物料描述", "服务单_物料描述"),
            ("服务状态", "服务单_状态"),
            ("客户", "服务单_客户"),
            ("电话", "服务单_电话"),
            ("线下带货上门", "服务单_线下带货"),
            ("是否结单", "服务单_是否结单"),
            ("地址", "服务单_地址"),
            # 八、保卡扫描
            ("扫描时间", "保卡扫描_时间"),
            ("条码", "保卡扫描_条码"),
            ("物料", "保卡扫描_物料"),
            ("订单号码", "保卡扫描_订单号"),
            ("服务单号", "保卡扫描_服务单号"),
            ("服务单日期", "保卡扫描_服务单日期"),
            # 九、押金返还
            ("返还日期", "押金返还_日期"),
            ("押金单号", "押金返还_单号"),
            ("押金经销商", "押金返还_押金经销商"),
            ("订单号码", "押金返还_订单号"),
            ("服务单日期", "押金返还_服务单日期"),
            ("服务单号", "押金返还_服务单号"),
            ("服务经销商", "押金返还_服务经销商"),
            ("押金金额", "押金返还_金额"),
            ("物料", "押金返还_物料"),
        ]

        # 过滤行
        clean_lines = []
        skip_keywords = ['主报表', '组树', '正在处理文档，请稍候。', '条码详细信息']
        for line in lines:
            line = line.strip()
            if not line or line in skip_keywords:
                continue
            line = line.replace('\xa0', '').strip()
            if line:
                clean_lines.append(line)

        print(f"  清理后共 {len(clean_lines)} 行")

        # 逐行扫描，匹配字段名（基于 HTML 表格结构的 2 列配对）
        found = []
        for i, line in enumerate(clean_lines):
            for field_name, result_key in sections:
                if field_name == line:
                    # 找到字段名，下一行应该是值（但要确认值不是另一个字段名）
                    if i + 1 < len(clean_lines):
                        value = clean_lines[i + 1]
                        is_also_field = any(fn == value for fn, _ in sections)
                        if not is_also_field and value and len(value) < 200:
                            if result.get(result_key, "（空）") == "（空）":
                                result[result_key] = value
                                found.append(f"{field_name}={value}")

        found_count = sum(1 for v in result.values() if v != "（空）" and v != barcode and v != "查询时间")
        print(f"  文本解析找到 {found_count} 个字段")
        if found:
            print(f"  提取的字段: {found[:20]}")

        return result

    def close_report_tab(self):
        """关闭报表标签页，只保留CRM列表页"""
        try:
            pages = self.context.pages
            print(f"  [DEBUG] 关闭前共有 {len(pages)} 个标签页")

            # 关闭所有非CRM列表页的标签页
            pages_to_close = []
            crm_list_page_idx = None

            for i, p in enumerate(pages):
                try:
                    url = p.url
                    # 只保留 CRM 报表列表页
                    if "crmportal.ecowaterchina" in url and "/report/reportlist" in url:
                        crm_list_page_idx = i
                    else:
                        pages_to_close.append(i)
                except:
                    pages_to_close.append(i)

            # 关闭多余的标签页（从后往前关，避免索引变化）
            pages_to_close.sort(reverse=True)
            for idx in pages_to_close:
                try:
                    pages[idx].close()
                    print(f"  [DEBUG] 已关闭标签页 {idx}")
                except:
                    pass
            time.sleep(1)

            # 切换到CRM列表页
            if crm_list_page_idx is not None:
                self.page = self.context.pages[crm_list_page_idx]
                self.page.bring_to_front()
                print(f"  [DEBUG] 已切换到CRM列表页（标签页{crm_list_page_idx}）")
                time.sleep(2)
            else:
                print("  [WARN] 未找到CRM列表页")
        except Exception as e:
            print(f"关闭标签页失败: {e}")

    def load_barcodes(self) -> list:
        """从 Excel 文件加载条码清单"""
        barcodes_file = self.config["barcodes"]["input_file"]

        # 如果文件不存在，创建一个示例文件
        if not os.path.exists(barcodes_file):
            print(f"条码文件不存在，创建示例文件: {barcodes_file}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.config["barcodes"]["sheet_name"]
            ws["A1"] = "条码号"
            ws["B1"] = "备注"
            ws["A2"] = "142112090153"
            ws["B2"] = "示例条码1"
            ws["A3"] = "5322508110583"
            ws["B3"] = "示例条码2"
            wb.save(barcodes_file)
            print(f"示例文件已创建，请编辑添加实际条码后重新运行")
            return []

        # 读取条码
        wb = openpyxl.load_workbook(barcodes_file)
        ws = wb.active

        barcodes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                barcodes.append(str(row[0]).strip())

        print(f"从 {barcodes_file} 加载了 {len(barcodes)} 个条码")
        return barcodes

    def save_results(self, results: list):
        """保存结果到 Excel 文件"""
        if not results:
            print("没有结果需要保存")
            return

        # 生成文件名（格式：202604061104）
        now = datetime.now().strftime("%Y%m%d%H%M")
        output_file = os.path.join(
            self.config["output"]["results_dir"],
            f"{now}.xlsx"
        )

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "条码查询结果"

        # 设置样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 获取所有字段名
        all_fields = list(results[0].keys())

        # 写入表头
        for col, field in enumerate(all_fields, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 高亮关键字段
        key_fields = ["设备档案_客户", "设备档案_电话", "设备档案_地址", "服务单_服务站"]
        key_columns = []
        for col, field in enumerate(all_fields, 1):
            if field in key_fields:
                key_columns.append(col)

        # 写入数据
        for row_idx, result in enumerate(results, 2):
            for col, field in enumerate(all_fields, 1):
                cell = ws.cell(row=row_idx, column=col, value=result.get(field, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # 关键字段高亮
                if col in key_columns:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 调整列宽
        for col in range(1, len(all_fields) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # 保存文件
        wb.save(output_file)
        print(f"\n结果已保存到: {output_file}")
        return output_file

    def login_and_save_session(self, username=None):
        """登录并保存会话 - 自动填写账号密码、发送验证码、登录"""
        print("\n--- 登录 ---")

        # 加载账号
        accounts = load_accounts()
        if not accounts:
            print("没有找到保存的账号，请先添加账号！")
            return

        # 如果没指定账号，取第一个
        if not username:
            username = list(accounts.keys())[0]

        if username not in accounts:
            print(f"账号 {username} 不存在！")
            return

        password = accounts[username]["password"]
        print(f"使用账号: {username}")

        session_dir = self.config["session"]["state_path"]
        os.makedirs(session_dir, exist_ok=True)

        # 优雅关闭可能占用 session 目录的 Chrome 进程
        self.cleanup_session_lock(session_dir)

        self.playwright = sync_playwright().start()

        # 启动浏览器
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=session_dir,
            headless=False,
            viewport=self.config["browser"]["viewport"]
        )
        self.page = self.context.pages[0]

        # 打开网站
        self.page.goto(self.config["website"]["url"], timeout=30000)
        print("等待登录页面加载...")
        time.sleep(5)  # 等待页面完全加载

        # 先检查是否已经登录
        url = self.page.url.lower()
        page_text = self.page.inner_text("body") if self.page else ""

        if "login" not in url and ("退出" in page_text or "首页" in page_text or "报表" in page_text):
            print("\n[已登录] 检测到网站已登录，会话有效！")
            print(f"  URL: {self.page.url}")
            input("\n按 Enter 保存会话并关闭浏览器...")
            print("正在关闭浏览器，备份 cookies 并保存会话数据...")
            self._close_browser_gracefully()
            print("会话已保存！")
            return

        # 未登录，开始登录流程
        print("未检测到登录，开始登录流程...")

        # 查找并填写用户名
        print("步骤1: 填写用户名...")
        try:
            for selector in [
                "input[name='username']",
                "input[name='user']",
                "input[name='account']",
                "input[id*='username']",
                "input[id*='user']",
                "input[type='text']"
            ]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.fill(username)
                        print(f"  [OK] 用户名已填写: {username}")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 填写用户名: {e}")

        # 填写密码
        print("步骤2: 填写密码...")
        try:
            for selector in [
                "input[name='password']",
                "input[name='pwd']",
                "input[type='password']"
            ]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.fill(password)
                        print("  [OK] 密码已填写")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 填写密码: {e}")

        # 点击登录按钮
        print("步骤3: 点击'登录'按钮...")
        try:
            for selector in [
                "button:has-text('登录')",
                "a:has-text('登录')",
                "input[value*='登录']",
                "text=登录"
            ]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.click()
                        print("  [OK] 已点击'登录'")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击登录: {e}")

        # 等待页面响应（可能需要输入验证码）
        print("等待页面响应...")
        time.sleep(2)

        # 点击发送验证码
        print("步骤4: 点击'发送验证码'...")
        try:
            for selector in [
                "button:has-text('发送验证码')",
                "button:has-text('获取验证码')",
                "a:has-text('发送验证码')",
                "a:has-text('获取验证码')",
                "text=发送验证码"
            ]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.click()
                        print("  [OK] 已点击'发送验证码'")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击发送验证码: {e}")

        # 等待验证码发送
        print("等待验证码发送...")
        time.sleep(2)

        # 填写验证码
        print("步骤5: 填写验证码...")
        captcha = input("\n请输入收到的验证码: ").strip()

        if not captcha:
            print("验证码不能为空，登录取消")
            self.context.close()
            self.playwright.stop()
            return

        # 填写验证码 - 点击输入框，然后用键盘输入
        try:
            captcha_input = self.page.query_selector("input[placeholder='验证码']")
            if captcha_input:
                # 点击输入框
                captcha_input.click()
                time.sleep(0.3)
                # 全选并删除
                captcha_input.press("Control+a")
                time.sleep(0.1)
                captcha_input.press("Backspace")
                time.sleep(0.1)
                # 逐字输入验证码
                captcha_input.type(captcha, delay=150)
                print(f"  [OK] 验证码已输入: {captcha}")
                # 确认输入成功
                val = captcha_input.get_attribute("value")
                print(f"  [调试] 验证码输入框值: '{val}'")
            else:
                print("  [失败] 没有找到验证码输入框")
        except Exception as e:
            print(f"  [失败] 填写验证码: {e}")

        # 点击确定按钮
        print("步骤6: 点击'确定'按钮...")
        try:
            # 用 placeholder="验证码" 的父元素下的按钮，或者直接用索引
            # 根据调试，按钮索引是 [6]
            buttons = self.page.query_selector_all("button")
            for i, btn in enumerate(buttons):
                try:
                    if btn.is_visible():
                        text = btn.inner_text().strip()
                        if "确" in text and "定" in text:
                            btn.click()
                            print(f"  [OK] 已点击'确定' (索引{i}, 文本='{text}')")
                            break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击确定: {e}")

        # 等待登录完成
        print("等待登录完成...")
        time.sleep(5)

        # 检查是否登录成功 - 多种方式检测
        url = self.page.url.lower()
        page_text = self.page.inner_text("body") if self.page else ""

        # 检测成功：URL不包含login 或者 页面包含退出/用户名等
        login_success = (
            "login" not in url or
            "退出" in page_text or
            "首页" in page_text or
            "报表" in page_text or
            "home" in url
        )

        # 如果登录失败且页面显示验证码错误，重试
        retry_count = 0
        max_retries = 3
        while not login_success and retry_count < max_retries:
            # 检查是否验证码错误（多种可能的错误提示）
            error_keywords = ["验证失败", "验证码错误", "验证码不正确", "验证码有误", "错误", "失败"]
            has_error = any(kw in page_text for kw in error_keywords)
            if has_error:
                print(f"\n[验证码错误] 第 {retry_count + 1} 次尝试")
                retry_count += 1

                # 先点击重新获取验证码
                print("重新获取验证码...")
                try:
                    for selector in [
                        "button:has-text('发送验证码')",
                        "button:has-text('获取验证码')",
                        "button:has-text('重新获取')"
                    ]:
                        try:
                            elem = self.page.query_selector(selector)
                            if elem and elem.is_visible():
                                elem.click()
                                print("  [OK] 已点击发送验证码")
                                time.sleep(3)
                                break
                        except:
                            pass
                except Exception as e:
                    print(f"  获取验证码失败: {e}")

                print("请重新输入验证码...")

                # 重新获取验证码输入框
                try:
                    captcha_input = self.page.query_selector("input[placeholder='验证码']")
                    if captcha_input:
                        captcha_input.click()
                        time.sleep(0.3)
                        captcha_input.press("Control+a")
                        time.sleep(0.1)
                        captcha_input.press("Backspace")
                        time.sleep(0.1)
                        new_code = input("请输入新的验证码: ").strip()
                        if not new_code:
                            print("验证码不能为空")
                            break
                        captcha_input.type(new_code, delay=150)
                        print(f"  [OK] 新验证码已输入: {new_code}")

                        # 点击确定
                        buttons = self.page.query_selector_all("button")
                        for i, btn in enumerate(buttons):
                            try:
                                if btn.is_visible():
                                    text = btn.inner_text().strip()
                                    if "确" in text and "定" in text:
                                        btn.click()
                                        break
                            except:
                                pass

                        print("等待验证结果...")
                        time.sleep(5)

                        # 重新检查
                        url = self.page.url.lower()
                        page_text = self.page.inner_text("body") if self.page else ""
                        login_success = (
                            "login" not in url or
                            "退出" in page_text or
                            "首页" in page_text or
                            "报表" in page_text or
                            "home" in url
                        )
                        if login_success:
                            print("\n[成功] 登录成功！")
                            print(f"  URL: {self.page.url}")
                            break
                    else:
                        print("找不到验证码输入框")
                        break
                except Exception as e:
                    print(f"重试失败: {e}")
                    break
            else:
                break

        if not login_success:
            print(f"\n[警告] 登录可能失败")
            print(f"  URL: {self.page.url}")
            print(f"  页面内容: {page_text[:200]}")

        input("\n按 Enter 保存会话并关闭浏览器...")

        print("正在关闭浏览器，备份 cookies 并保存会话数据...")
        self._close_browser_gracefully()

        print("会话已保存！")

    def single_query(self, barcode: str) -> dict:
        """单个条码查询"""
        print(f"\n--- 单个查询: {barcode} ---")

        # 检查会话
        session_dir = self.config["session"]["state_path"]

        # 优雅关闭可能占用 session 目录的 Chrome 进程
        self.cleanup_session_lock(session_dir)

        # 加载浏览器
        self.playwright = sync_playwright().start()

        try:
            # 检查是否存在有效会话
            session_files = []
            if os.path.exists(session_dir):
                session_files = [f for f in os.listdir(session_dir) if f not in ['SingletonLock', 'lockfile']]

            if not session_files:
                print("未登录，开始登录流程...")
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)

                # 执行登录
                self._do_login()

                # 保存会话
                print("保存会话...")
                time.sleep(2)
            else:
                # 加载已有会话 - persistent_context 会自动从 user_data_dir 恢复会话
                print(f"加载已有会话（{len(session_files)} 个文件）...")

                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)

                # 检查登录状态
                if "login" in self.page.url.lower():
                    print("会话已过期，重新登录...")
                    # 关闭旧context
                    try:
                        self.context.close()
                    except:
                        pass
                    time.sleep(1)

                    # 重新创建context
                    self.context = self.playwright.chromium.launch_persistent_context(
                        user_data_dir=session_dir,
                        headless=False,
                        viewport=self.config["browser"]["viewport"]
                    )
                    self.page = self.context.pages[0]
                    self.page.goto(self.config["website"]["url"], timeout=30000)
                    time.sleep(2)

                    # 执行登录
                    self._do_login()

            # 导航到报表
            self.navigate_to_report()
            self.switch_to_report_tab()

            # 查询
            result = self.query_barcode(barcode)

            # 关闭浏览器（备份 cookies + persistent_context 自动保存）
            self._close_browser_gracefully()
            print("  浏览器已关闭，会话数据已保存")
            return result

        except Exception as e:
            print(f"查询出错: {e}")
            import traceback
            traceback.print_exc()
            if self.context:
                try:
                    self.context.close()
                except:
                    pass
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass
            return None

    def _do_login(self):
        """执行登录流程"""
        print("\n--- 登录 ---")

        # 加载账号
        accounts = load_accounts()
        if not accounts:
            print("没有找到保存的账号，请先使用菜单4添加账号！")
            input("按 Enter 退出...")
            sys.exit(1)

        # 选择账号
        usernames = list(accounts.keys())
        if len(usernames) == 1:
            username = usernames[0]
        else:
            print("\n--- 选择登录账号 ---")
            for i, u in enumerate(usernames, 1):
                print(f"  {i}. {u}")
            print("-" * 30)
            while True:
                try:
                    idx = int(input("请选择账号: ").strip())
                    if 1 <= idx <= len(usernames):
                        username = usernames[idx - 1]
                        break
                    print("无效选择，请重新输入")
                except ValueError:
                    print("请输入数字")

        password = accounts[username]["password"]
        print(f"使用账号: {username}")

        # 填写用户名
        print("步骤1: 填写用户名...")
        try:
            for selector in ["input[name='username']", "input[name='user']", "input[type='text']"]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.fill(username)
                        print(f"  [OK] 用户名已填写: {username}")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 填写用户名: {e}")

        # 填写密码
        print("步骤2: 填写密码...")
        try:
            for selector in ["input[name='password']", "input[type='password']"]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.fill(password)
                        print("  [OK] 密码已填写")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 填写密码: {e}")

        # 点击登录
        print("步骤3: 点击'登录'...")
        try:
            for selector in ["button:has-text('登录')", "a:has-text('登录')"]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.click()
                        print("  [OK] 已点击'登录'")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击登录: {e}")

        time.sleep(2)

        # 点击发送验证码
        print("步骤4: 点击'发送验证码'...")
        try:
            for selector in ["button:has-text('发送验证码')", "a:has-text('发送验证码')"]:
                try:
                    elem = self.page.query_selector(selector)
                    if elem and elem.is_visible():
                        elem.click()
                        print("  [OK] 已点击'发送验证码'")
                        break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击发送验证码: {e}")

        time.sleep(2)

        # 输入验证码
        captcha = input("\n请输入手机收到的验证码: ").strip()
        if not captcha:
            print("验证码不能为空，登录取消")
            self.context.close()
            self.playwright.stop()
            sys.exit(1)

        print("步骤5: 填写验证码...")
        try:
            captcha_input = self.page.query_selector("input[placeholder='验证码']")
            if captcha_input:
                captcha_input.click()
                time.sleep(0.3)
                captcha_input.press("Control+a")
                time.sleep(0.1)
                captcha_input.press("Backspace")
                time.sleep(0.1)
                captcha_input.type(captcha, delay=150)
                print(f"  [OK] 验证码已输入: {captcha}")
        except Exception as e:
            print(f"  [失败] 填写验证码: {e}")

        # 点击确定
        print("步骤6: 点击'确定'...")
        try:
            buttons = self.page.query_selector_all("button")
            for btn in buttons:
                try:
                    if btn.is_visible():
                        text = btn.inner_text().strip()
                        if "确" in text and "定" in text:
                            btn.click()
                            print(f"  [OK] 已点击'确定'")
                            break
                except:
                    pass
        except Exception as e:
            print(f"  [失败] 点击确定: {e}")

        print("等待登录完成...")
        time.sleep(5)

    def _save_single_result(self, result: dict, output_file: str):
        """保存单个查询结果到 Excel"""
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "条码查询结果"

        # 样式
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        key_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # 关键字段
        key_fields = ["设备档案_客户", "设备档案_电话", "设备档案_地址", "服务单_服务站"]

        # 写入数据
        row = 1
        for key, value in result.items():
            col = 1
            # 字段名
            cell = ws.cell(row=row, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            col += 1
            # 值
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if key in key_fields:
                cell.fill = key_fill
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

        wb.save(output_file)
        print(f"\n结果已保存到: {output_file}")

    def run_batch(self, barcodes_file: str = None):
        """批量查询"""
        if barcodes_file:
            self.config["barcodes"]["input_file"] = barcodes_file

        print("\n--- 批量查询 ---")

        # 加载条码
        barcodes = self.load_barcodes()
        if not barcodes:
            print("没有条码可查询")
            return

        # 检查会话
        session_dir = self.config["session"]["state_path"]

        # 优雅关闭可能占用 session 目录的 Chrome 进程
        self.cleanup_session_lock(session_dir)

        # 加载浏览器
        self.playwright = sync_playwright().start()

        try:
            # 检查是否存在有效会话
            session_files = []
            if os.path.exists(session_dir):
                session_files = [f for f in os.listdir(session_dir) if f not in ['SingletonLock', 'lockfile']]

            if not session_files:
                print("未登录，开始登录流程...")
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)
                self._do_login()
                print("保存会话...")
                time.sleep(2)
            else:
                print(f"加载已有会话（{len(session_files)} 个文件）...")

                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)

                if "login" in self.page.url.lower():
                    print("会话已过期，重新登录...")
                    # 直接在当前登录页面执行登录
                    self._do_login()

            # 导航到报表
            self.navigate_to_report()
            self.switch_to_report_tab()

            # 批量查询
            results = []
            for i, barcode in enumerate(barcodes, 1):
                print(f"\n[{i}/{len(barcodes)}] 查询: {barcode}")
                result = self.query_barcode(barcode)
                results.append(result)

                if i < len(barcodes):
                    self.close_report_tab()
                    time.sleep(2)
                    try:
                        self.prepare_next_report()
                    except:
                        print("  重新准备报表页面失败，跳过...")
                        break

            # 批量导出不合并，每个文件单独保存
            print(f"\n批量导出完成！请将下载的文件重命名为对应条码号。")
            print(f"本次查询了 {len(results)} 个条码。")

            # 关闭浏览器（备份 cookies + persistent_context 自动保存）
            self._close_browser_gracefully()
            print("  浏览器已关闭，会话数据已保存")

        except Exception as e:
            print(f"批量查询出错: {e}")
            if self.context:
                try:
                    self.context.close()
                except:
                    pass
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass

    def run_batch_inline(self, barcodes: list):
        """直接输入条码列表进行批量查询"""
        if not barcodes:
            print("没有条码可查询")
            return

        session_dir = self.config["session"]["state_path"]

        # 优雅关闭可能占用 session 目录的 Chrome 进程
        self.cleanup_session_lock(session_dir)

        self.playwright = sync_playwright().start()

        try:
            # 检查是否存在有效会话
            session_files = []
            if os.path.exists(session_dir):
                session_files = [f for f in os.listdir(session_dir) if f not in ['SingletonLock', 'lockfile']]

            if not session_files:
                print("未登录，开始登录流程...")
                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)
                self._do_login()
                print("保存会话...")
                time.sleep(2)
            else:
                print(f"加载已有会话（{len(session_files)} 个文件）...")

                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)

                if "login" in self.page.url.lower():
                    print("会话已过期，重新登录...")
                    # 直接在当前登录页面执行登录
                    self._do_login()

            # 导航到报表
            self.navigate_to_report()
            self.switch_to_report_tab()

            # 批量查询
            results = []
            for i, barcode in enumerate(barcodes, 1):
                print(f"\n[{i}/{len(barcodes)}] 查询: {barcode}")
                result = self.query_barcode(barcode)
                results.append(result)

                if i < len(barcodes):
                    self.close_report_tab()
                    time.sleep(2)
                    try:
                        self.prepare_next_report()
                    except:
                        print("  重新准备报表页面失败，跳过...")
                        break

            print(f"\n批量查询完成！共查询 {len(results)} 个条码。")

            # 关闭浏览器（备份 cookies + persistent_context 自动保存）
            self._close_browser_gracefully()
            print("  浏览器已关闭，会话数据已保存")

        except Exception as e:
            print(f"批量查询出错: {e}")
            if self.context:
                try:
                    self.context.close()
                except:
                    pass
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass

    def run_batch_headless(self, barcodes: list):
        """后台批量查询（无头模式，不影响电脑使用）"""
        import signal

        print("\n" + "="*50)
        print("  ⭐ 后台批量查询模式")
        print("  浏览器将在后台运行，不会弹出窗口")
        print("="*50)

        if not barcodes:
            print("没有条码需要查询")
            return

        print(f"\n共 {len(barcodes)} 个条码，开始后台查询...\n")

        try:
            session_dir = self.config["session"]["state_path"]

            self.cleanup_session_lock(session_dir)

            self.playwright = sync_playwright().start()

            os.makedirs(session_dir, exist_ok=True)

            session_files = []
            if os.path.exists(session_dir):
                session_files = [f for f in os.listdir(session_dir) if f not in ['SingletonLock', 'lockfile']]
            has_existing_session = len(session_files) > 0

            if has_existing_session:
                print(f"发现保存的会话（{len(session_files)} 个文件），正在加载...")
                try:
                    self.context = self.playwright.chromium.launch_persistent_context(
                        user_data_dir=session_dir,
                        headless=True,
                        viewport=self.config["browser"]["viewport"]
                    )
                    self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
                    print("  会话加载成功（无头模式）")
                except Exception as e:
                    print(f"  加载会话失败: {e}")
                    has_existing_session = False
            else:
                print("未发现保存的会话")

            if not has_existing_session:
                print("\n⚠️ 首次使用后台模式，需要先登录一次")
                print("  将临时启动可见浏览器完成登录，之后即可纯后台运行\n")

                self.context = self.playwright.chromium.launch_persistent_context(
                    user_data_dir=session_dir,
                    headless=False,
                    viewport=self.config["browser"]["viewport"]
                )
                self.page = self.context.pages[0]
                self.page.goto(self.config["website"]["url"], timeout=30000)
                time.sleep(2)

                accounts = load_accounts()
                if accounts:
                    username = list(accounts.keys())[0]
                    print(f"使用账号登录: {username}")
                    self._do_login()
                    input("\n登录完成！按 Enter 切换到后台模式继续查询...")

                    self._save_storage_state_backup()
                    self.context.close()
                    time.sleep(3)
                    self.playwright.stop()

                    self.cleanup_session_lock(session_dir)
                    self.playwright = sync_playwright().start()
                    self.context = self.playwright.chromium.launch_persistent_context(
                        user_data_dir=session_dir,
                        headless=True,
                        viewport=self.config["browser"]["viewport"]
                    )
                    self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
                    print("  已切换到无头模式 ✅\n")
                else:
                    print("没有已保存的账号，请先用选项4添加账号，再用选项3登录一次")
                    self.context.close()
                    self.playwright.stop()
                    return

            self.page.goto(self.config["website"]["url"], timeout=60000)
            time.sleep(5)

            url = self.page.url
            page_text = ""
            try:
                page_text = self.page.inner_text("body").strip()[:1000]
            except:
                pass

            print(f"  当前URL: {url}")

            needs_login = (
                "login" in url.lower() or
                "用户名" in page_text or
                "密码" in page_text or
                "登录" in page_text or
                "Login" in url
            )

            if needs_login:
                print("\n⚠️ 检测到会话已过期，需要重新登录")
                print("  正在尝试自动登录...")

                accounts = load_accounts()
                if not accounts:
                    print("  ❌ 没有已保存的账号，无法自动登录")
                    print("  请用选项3重新登录后再试")
                    self.context.close()
                    self.playwright.stop()
                    return

                username = list(accounts.keys())[0]
                password = accounts[username]["password"]

                time.sleep(2)

                username_input = None
                password_input = None

                selectors_to_try = [
                    "#txtUserName",
                    "input[name*='user']",
                    "input[name*='User']",
                    "input[type='text']",
                    "input[id*='user']",
                    "input[id*='User']",
                ]
                for sel in selectors_to_try:
                    username_input = self.page.query_selector(sel)
                    if username_input:
                        break

                pwd_selectors = [
                    "#txtPassword",
                    "input[name*='pass']",
                    "input[name*='Pass']",
                    "input[type='password']",
                    "input[id*='pass']",
                    "input[id*='Pass']",
                ]
                for sel in pwd_selectors:
                    password_input = self.page.query_selector(sel)
                    if password_input:
                        break

                if not username_input or not password_input:
                    print("  ⚠️ 未找到登录表单，尝试截图诊断...")
                    try:
                        self.page.screenshot(path="debug_login_page.png")
                        print(f"  已保存页面截图: debug_login_page.png")
                        all_inputs = self.page.query_selector_all("input")
                        print(f"  页面上共有 {len(all_inputs)} 个 input 元素")
                        for inp in all_inputs[:10]:
                            inp_type = inp.get_attribute("type") or ""
                            inp_name = inp.get_attribute("name") or ""
                            inp_id = inp.get_attribute("id") or ""
                            print(f"    input: type={inp_type}, name={inp_name}, id={inp_id}")
                    except Exception as e:
                        print(f"  截图失败: {e}")

                    print("\n  💡 建议：请先用「选项3 登录账号」更新会话，再使用后台模式")
                    self.context.close()
                    self.playwright.stop()
                    return

                username_input.fill(username)
                password_input.fill(password)
                print(f"  已填写账号: {username}")

                send_btn = self.page.query_selector("#btnSendCode")
                if send_btn:
                    send_btn.click()
                    print("  验证码已发送到手机，请在下方输入：")
                    code = input("  验证码: ").strip()
                    code_input = self.page.query_selector("#txtCode")
                    if code_input:
                        code_input.fill(code)
                        login_btn = self.page.query_selector("#btnLogin")
                        if login_btn:
                            login_btn.click()
                            print("  登录中...")
                            time.sleep(5)
                            print("  登录成功！切换到无头模式继续...")
                            self._save_storage_state_backup()
                    else:
                        print("  ❌ 找不到验证码输入框")
                        self.context.close()
                        self.playwright.stop()
                        return
                else:
                    print("  ❌ 找不到验证码按钮，请手动登录后重试")
                    self.context.close()
                    self.playwright.stop()
                    return

            self.navigate_to_report()
            self.switch_to_report_tab()

            results = []
            success_count = 0
            fail_count = 0

            for i, barcode in enumerate(barcodes, 1):
                print(f"\n[{i}/{len(barcodes)}] 查询: {barcode}")

                try:
                    result = self.query_barcode(barcode)
                    results.append(result)
                    if result.get("状态") == "成功":
                        success_count += 1
                    else:
                        fail_count += 1
                except Exception as e:
                    print(f"  ❌ 查询出错: {e}")
                    results.append({"条码": barcode, "状态": "错误", "错误信息": str(e)})
                    fail_count += 1

                if i < len(barcodes):
                    self.close_report_tab()
                    time.sleep(2)
                    try:
                        self.prepare_next_report()
                    except Exception as e:
                        print(f"  重新准备报表失败: {e}，跳过...")
                        break

            self.save_results(results)

            print(f"\n{'='*50}")
            print(f"  后台批量查询完成！")
            print(f"  总计: {len(barcodes)} 条 | 成功: {success_count} | 失败: {fail_count}")
            print(f"{'='*50}")

            self._close_browser_gracefully()
            print("  浏览器已关闭，会话数据已保存")

        except KeyboardInterrupt:
            print("\n\n⚠️ 用户中断查询")
            try:
                if self.context:
                    self._close_browser_gracefully()
            except:
                pass
        except Exception as e:
            print(f"\n❌ 后台查询出错: {e}")
            import traceback
            traceback.print_exc()
            if self.context:
                try:
                    self.context.close()
                except:
                    pass
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass

    def run(self):
        """运行主流程"""
        print("="*60)
        print("怡口 CRM 条码批量查询系统")
        print("="*60)

        # 1. 加载条码清单
        barcodes = self.load_barcodes()
        if not barcodes:
            print("没有条码可查询，退出")
            return

        # 2. 初始化浏览器
        self.setup_browser()

        # 3. 导航到报表页面
        self.navigate_to_report()

        # 4. 切换到报表标签页
        if not self.switch_to_report_tab():
            input("未能打开报表标签页，请检查后按 Enter 重试...")

        # 5. 批量查询
        results = []
        for i, barcode in enumerate(barcodes, 1):
            print(f"\n[{i}/{len(barcodes)}] 正在查询...")

            # 查询单个条码
            result = self.query_barcode(barcode)
            results.append(result)

            # 如果不是最后一个，关闭报表标签页并重新打开
            if i < len(barcodes):
                self.close_report_tab()

                # 等待页面稳定
                time.sleep(2)

                # 导航回报表列表页
                print("  导航回报表列表页...")
                try:
                    # 检查当前页面状态
                    content = self.page.inner_text("body")
                    if "查询条码所有信息" not in content:
                        # 需要导航回报表列表
                        print("  当前不在报表列表页，尝试导航...")
                        self.page.goto(self.config["website"]["url"], timeout=30000)
                        time.sleep(3)

                        # 点击报表管理
                        try:
                            self.page.click("text=报表管理", timeout=5000)
                            time.sleep(2)
                        except:
                            pass

                        # 点击水晶报表查看
                        try:
                            self.page.click("text=水晶报表查看", timeout=5000)
                            time.sleep(2)
                        except:
                            pass
                except Exception as e:
                    print(f"  导航失败: {e}")

                # 翻到第2页
                print("  翻到第2页...")
                翻页成功 = False
                try:
                    # 尝试多种方式翻页
                    page2 = self.page.get_by_text("2", exact=True).first
                    if page2:
                        page2.click()
                        time.sleep(2)
                        print("  已翻到第2页")
                        翻页成功 = True
                except:
                    pass

                if not 翻页成功:
                    try:
                        # 尝试查找 spinbutton
                        spinbuttons = self.page.query_selector_all("spinbutton, input[type='number']")
                        for sb in spinbuttons:
                            try:
                                val = sb.get_attribute("value")
                                if val == "1":
                                    sb.fill("2")
                                    sb.press("Enter")
                                    time.sleep(2)
                                    print("  已翻到第2页（通过spinbutton）")
                                    翻页成功 = True
                                    break
                            except:
                                pass
                    except:
                        pass

                # 用 JS dblclick 重新打开报表
                print("  尝试重新打开报表...")
                pages_before = len(self.context.pages)

                # 使用 JS dblclick
                try:
                    result = self.page.evaluate("""() => {
                        const elements = document.querySelectorAll('*');
                        for (const el of elements) {
                            if (el.textContent.trim() === '查询条码所有信息') {
                                const dblclickEvent = new MouseEvent('dblclick', {
                                    bubbles: true,
                                    cancelable: true,
                                    view: window
                                });
                                el.dispatchEvent(dblclickEvent);
                                return true;
                            }
                        }
                        return false;
                    }""")
                    if result:
                        print("  -> JS dblclick 事件已触发，等待新标签页...")
                        time.sleep(8)  # 增加等待时间

                        pages_now = len(self.context.pages)
                        if pages_now > pages_before:
                            print("  -> 成功打开新标签页")
                        else:
                            print("  -> JS dblclick 没有打开新标签，尝试 mouse.dblclick...")
                            report_link = self.page.get_by_text("查询条码所有信息", exact=True).first
                            if report_link:
                                box = report_link.bounding_box()
                                if box:
                                    x = box['x'] + box['width'] / 2
                                    y = box['y'] + box['height'] / 2
                                    self.page.mouse.dblclick(x, y)
                                    print("  -> 执行 mouse.dblclick")
                                    time.sleep(8)
                    else:
                        print("  -> 未找到'查询条码所有信息'元素")
                        input("  请手动双击后按 Enter 继续...")
                except Exception as e:
                    print(f"  -> JS dblclick 失败: {e}")
                    input("  请手动双击后按 Enter 继续...")

                # 切换到新标签页
                time.sleep(2)
                self.switch_to_report_tab()

        # 6. 保存结果
        self.save_results(results)

        # 7. 清理 - 关闭浏览器并保存会话
        print("\n查询完成！")
        try:
            if self.context:
                # 关闭所有页面
                for page in self.context.pages:
                    try:
                        page.close()
                    except:
                        pass
                # 关闭 context（persistent_context 会自动保存会话）
                self.context.close()
                print("  会话已保存")
        except Exception as e:
            print(f"  关闭浏览器时出错: {e}")
        finally:
            if self.playwright:
                try:
                    self.playwright.stop()
                except:
                    pass

        print("\n再见！")


def main():
    """交互式主入口"""
    while True:
        print_menu()
        choice = input("\n请选择 (1-6): ").strip()

        if choice == "1":
            # 查询单个条码
            barcode = input("\n请输入条码号: ").strip()
            if barcode:
                app = BarcodeQueryApp()
                result = app.single_query(barcode)
                if result:
                    print("\n查询结果:")
                    for k, v in result.items():
                        if v and v != "（空）":
                            print(f"  {k}: {v}")
                else:
                    print("查询失败")
            input("\n按 Enter 继续...")

        elif choice == "2":
            # 批量查询（前台模式）
            print("\n--- 批量查询（前台模式）---")
            print("请输入条码号，每行一个，输入空行结束：")
            barcodes = []
            while True:
                line = input().strip()
                if not line:
                    break
                barcodes.append(line)
            if not barcodes:
                print("没有输入条码")
            else:
                print(f"已输入 {len(barcodes)} 个条码，开始查询...")
                app = BarcodeQueryApp()
                app.run_batch_inline(barcodes)
            input("\n按 Enter 继续...")

        elif choice == "3":
            # 登录账号 - 选择已添加的账号
            accounts = load_accounts()
            if not accounts:
                print("\n没有已保存的账号，请先添加账号（选项4）")
            else:
                print("\n--- 选择登录账号 ---")
                usernames = list(accounts.keys())
                for i, username in enumerate(usernames, 1):
                    print(f"  {i}. {username}")
                print(f"  0. 取消")
                print("-" * 30)

                try:
                    idx = int(input("请选择账号: ").strip())
                    if idx == 0:
                        pass
                    elif 1 <= idx <= len(usernames):
                        username = usernames[idx - 1]
                        app = BarcodeQueryApp()
                        app.login_and_save_session(username)
                    else:
                        print("无效选择")
                except ValueError:
                    print("请输入数字")

            input("\n按 Enter 继续...")

        elif choice == "4":
            # 添加账号
            add_account_interactive()
            input("\n按 Enter 继续...")

        elif choice == "5":
            # 后台批量查询（无头模式）
            print("\n--- 后台批量查询（无头模式）---")
            print("浏览器将在后台运行，不会弹出窗口，不影响你使用电脑 ✅\n")
            print("请输入条码号，每行一个，输入空行结束：")
            barcodes = []
            while True:
                line = input().strip()
                if not line:
                    break
                barcodes.append(line)
            if not barcodes:
                print("没有输入条码")
            else:
                print(f"\n共 {len(barcodes)} 个条码，开始后台查询...")
                print("💡 提示：你可以随时切换到其他窗口工作，查询在后台进行\n")
                app = BarcodeQueryApp()
                app.run_batch_headless(barcodes)
            input("\n按 Enter 继续...")

        elif choice == "6":
            print("\n再见！")
            break

        else:
            print("\n无效选择，请重新输入")


if __name__ == "__main__":
    import sys

    # 支持命令行参数模式
    if len(sys.argv) > 1:
        if sys.argv[1] == "--query-single" and len(sys.argv) > 2:
            barcode = sys.argv[2]
            app = BarcodeQueryApp()
            result = app.single_query(barcode)
            if result and result.get("状态") == "成功":
                print(f"查询完成，HTML已保存到 barcode/{barcode}.html")
            else:
                print(f"查询失败: {result.get('状态', '未知') if result else '无结果'}")
            sys.exit(0)
        else:
            print("用法: python main.py [--query-single 条码]")
            sys.exit(1)

    main()
